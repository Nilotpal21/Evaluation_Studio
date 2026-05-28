#!/usr/bin/env python3
"""
sizing-live-xlsx.py — Live Google-Sheets-ready .xlsx for sizing-and-cost.

Reads:
  - benchmarks/config/sizing/infra-snapshot.json     (LIVE kubectl, refresh first)
  - benchmarks/config/sizing/rate-card.json          (user-maintained, hash-stamped)
  - benchmarks/config/sizing/scenarios/<id>.json     (scenario definition)
  - benchmarks/config/sizing/measurements/<id>/*.json (optional — skip math if absent)

Produces a workbook with:
  Summary        — headline + cost breakdown + binding layer + sanity check
  Warnings       — live contradiction flags
  Controls       — target msg/s, p95, utilization target (editable)
  RateCard       — unit prices (editable, Original column for reference)
  Scenario       — LLM profile, session pattern, per-msg cost (editable)
  Infra          — NODE INVENTORY + runtime/Mongo/Redis with requests+limits+placement
  Measurement    — workloadIntrinsic ratios + perPodCapacity + multiPodScaling
  TestEvidence   — saturation-run scorecard (per-step VUs/msg-s/p95/CPU%) + component RAG
  Provenance     — live age checks vs every source file
  Baseline       — F-7 with all formulas
  Scaling        — F-1..F-8 with formulas
  Bottleneck     — per-layer util % with conditional formatting
  WhatIf         — target sweep 0.25× / 0.5× / 1× / 2× / 5×

Usage:
  source benchmarks/scripts/.venv-sizing/bin/activate
  python3 benchmarks/scripts/sizing-live-xlsx.py --target 35 --out <path>
"""

import argparse
import json
import re
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

# ============================================================================
# CLI
# ============================================================================

parser = argparse.ArgumentParser()
parser.add_argument("--scenario", default="chat-agent-mock")
parser.add_argument("--target", type=int, default=None,
                    help="Target msg/s. If omitted AND --measurement given, defaults to the "
                         "measurement's safe-fleet-msg/s (maxSafeMsgPerSec × max-measured-pods). "
                         "If both omitted, defaults to 35.")
parser.add_argument("--p95", type=int, default=1300)
parser.add_argument("--util", type=float, default=0.70)
parser.add_argument("--measurement", default=None,
                    help="Path to saturation-measurement JSON. Omit for baseline-only mode.")
parser.add_argument("--sat-report", default=None,
                    help="Path to saturation-run markdown report. If omitted and --measurement given, "
                         "the script will try measurement.saturationReportPath.")
parser.add_argument("--rate-card", default="benchmarks/config/sizing/rate-card.json")
parser.add_argument("--infra", default="benchmarks/config/sizing/infra-snapshot.json")
parser.add_argument("--baseline-tiers", default="benchmarks/config/sizing/baseline-tiers.json",
                    help="Path to baseline-tiers.json (tiered HA sizing catalog for Mongo/Redis).")
parser.add_argument("--scenario-file", default=None)
parser.add_argument("--sat-run-dir", default=None,
                    help="Path to saturation run directory (benchmarks/results/runs/<label>/). "
                         "When set, reads poll JSONs for datastoreOps data and run config/summary. "
                         "Auto-discovers --sat-report from the run dir's markdown report.")
parser.add_argument("--env", default=None,
                    help="Environment (dev/qa/staging). When set, auto-refreshes infra-snapshot.json "
                         "from kubectl before generating the report. Overrides --infra with fresh data.")
parser.add_argument("--out", required=True)
args = parser.parse_args()

if args.scenario_file is None:
    args.scenario_file = f"benchmarks/config/sizing/scenarios/{args.scenario}.json"


def load_json(path):
    return json.loads(Path(path).read_text())


# --- Auto-discover from --sat-run-dir ----------------------------------------
sat_run_polls_data = None   # list of poll dicts with datastoreOps
sat_run_config = None       # run config.json
sat_run_summary = None      # run summary.json

if args.sat_run_dir:
    run_dir = Path(args.sat_run_dir)
    # Auto-discover sat-report from the run dir
    if args.sat_report is None:
        config_path = run_dir / "config.json"
        if config_path.exists():
            sat_run_config = load_json(config_path)
            k6_id = sat_run_config.get("k6RunId") or sat_run_config.get("runLabel", "")
            # Look for matching markdown report
            for md in sorted(Path("benchmarks/docs").glob("saturation-run-*.md"), reverse=True):
                if k6_id and k6_id in md.name:
                    args.sat_report = str(md)
                    print(f"[sat-run-dir] Auto-discovered sat-report: {md}")
                    break
    # Load summary
    summary_path = run_dir / "summary.json"
    if summary_path.exists():
        sat_run_summary = load_json(summary_path)
    # Load poll JSONs for datastoreOps
    polls_dir = run_dir / "polls"
    if polls_dir.is_symlink():
        polls_dir = polls_dir.resolve()
    if polls_dir.exists():
        import glob as _glob
        poll_files = sorted(_glob.glob(str(polls_dir / "poll-*.json")))
        sat_run_polls_data = []
        for pf in poll_files:
            try:
                poll = load_json(pf)
                ds = poll.get("datastoreOps")
                if ds and ds.get("delta"):
                    sat_run_polls_data.append({
                        "pollNumber": poll.get("pollNumber"),
                        "epoch": poll.get("epoch"),
                        "datastoreOps": ds,
                        "k6": poll.get("k6", {}),
                    })
            except Exception:
                pass
        if sat_run_polls_data:
            print(f"[sat-run-dir] Loaded {len(sat_run_polls_data)} polls with datastoreOps deltas")
        else:
            print("[sat-run-dir] No polls with datastoreOps deltas found")


# --- Auto-refresh infra-snapshot from kubectl when --env is set ---
if args.env:
    import subprocess
    refresh_script = Path(__file__).parent / "refresh-infra-snapshot.sh"
    if refresh_script.exists():
        print(f"[infra] Refreshing infra-snapshot.json from live cluster (env={args.env})...")
        result = subprocess.run(
            ["bash", str(refresh_script)],
            env={**__import__("os").environ, "ENV": args.env},
            capture_output=True, text=True, cwd=str(Path(__file__).parent.parent.parent)
        )
        if result.returncode == 0:
            print(result.stdout.strip())
        else:
            print(f"[infra] WARNING: refresh failed (rc={result.returncode}): {result.stderr.strip()}")
            print("[infra] Falling back to existing infra-snapshot.json")
    else:
        print(f"[infra] WARNING: refresh-infra-snapshot.sh not found at {refresh_script}")

rate = load_json(args.rate_card)
infra = load_json(args.infra)
scenario = load_json(args.scenario_file)
meas = load_json(args.measurement) if args.measurement else None
baseline_tiers_catalog = load_json(args.baseline_tiers) if Path(args.baseline_tiers).exists() else None

today = datetime.now(timezone.utc).strftime("%Y-%m-%d")

# Derive msgs_per_session early (needed by Controls tab before Scenario tab)
_turn = scenario["requestShape"].get("turnStructure", {})
msgs_per_session = _turn.get("creates", 1) + _turn.get("followups", 4)

# Resolve target: if --target omitted AND measurement present, default to the
# measurement's proven safe-fleet-msg/s (max measured point in multiPodScaling[]).
# This is the number the saturation test ACTUALLY demonstrated healthy at p95 <= target.
if args.target is None:
    if meas is not None:
        mps = meas.get("multiPodScaling", [])
        if mps:
            max_point = max(mps, key=lambda p: p["pods"])
            args.target = int(round(max_point["totalMsgPerSec"]))
            target_source = (
                f"measurement-proven safe fleet msg/s "
                f"(multiPodScaling[pods={max_point['pods']}] = "
                f"{max_point['totalMsgPerSec']:.2f} msg/s; p95 at max "
                f"= {meas['perPodCapacity'].get('p95AtMaxMs', '?')}ms)"
            )
        else:
            args.target = 35
            target_source = "fallback (no multiPodScaling data in measurement)"
    else:
        args.target = 35
        target_source = "default (no measurement provided)"
else:
    target_source = "user-supplied via --target"

print(f"[target] {args.target} msg/s — {target_source}")


# ----------------------------------------------------------------------------
# Baseline tier selection (F-7.T.1, F-7.T.2)
# ----------------------------------------------------------------------------
# Pick the smallest Mongo/Redis tier whose msgPerSecCeiling >= targetMsgPerSec.
# This is calculation-only — it does NOT mutate any deployed infra. The live
# infra-snapshot stays read-only and is used for measurement hash-binding and
# bottleneck checks, not for baseline cost sizing.
def _pick_tier(tiers_dict, target):
    """Return (tier_name, tier_spec) for smallest tier covering target msg/s.
    Raises ValueError if target exceeds all tiers (BASELINE-TIER-UNAVAILABLE).
    """
    # Sort by ceiling asc and pick first that covers target
    sorted_tiers = sorted(tiers_dict.items(),
                          key=lambda kv: kv[1]["msgPerSecCeiling"])
    for name, spec in sorted_tiers:
        if spec["msgPerSecCeiling"] >= target:
            return name, spec
    raise ValueError(
        f"BASELINE-TIER-UNAVAILABLE: target={target} msg/s exceeds largest tier "
        f"ceiling ({sorted_tiers[-1][1]['msgPerSecCeiling']}). "
        f"Extend baseline-tiers.json with a larger tier.")


if baseline_tiers_catalog is None:
    print("[baseline] WARN: baseline-tiers.json missing; baseline will use live infra-snapshot "
          "(this tends to over-price baseline if live cluster is dev-scale).")
    mongo_tier_name = None
    mongo_tier = None
    redis_tier_name = None
    redis_tier = None
    mongo_tier_table_loaded = False
    redis_tier_table_loaded = False
    runtime_tier_table_loaded = False
else:
    mongo_tier_table_loaded = True
    redis_tier_table_loaded = True
    runtime_tier_table_loaded = "runtime_tiers" in baseline_tiers_catalog
    # fall through to original else block
    mongo_tier_name, mongo_tier = _pick_tier(
        baseline_tiers_catalog["mongodb_tiers"], args.target)
    redis_tier_name, redis_tier = _pick_tier(
        baseline_tiers_catalog["redis_tiers"], args.target)
    print(f"[baseline] Mongo tier: {mongo_tier_name}  "
          f"({mongo_tier['replicas']} × {mongo_tier['cpuLimitCores']}c / "
          f"{mongo_tier['memoryLimitGi']}Gi, {mongo_tier['pvcSizeGi']}Gi PVC, "
          f"ceiling {mongo_tier['msgPerSecCeiling']} msg/s)")
    print(f"[baseline] Redis tier: {redis_tier_name}  "
          f"(master {redis_tier['master']['cpuLimitCores']}c/{redis_tier['master']['memoryLimitGi']}Gi "
          f"+ {redis_tier['replicaReplicas']} × {redis_tier['replica']['cpuLimitCores']}c/"
          f"{redis_tier['replica']['memoryLimitGi']}Gi, "
          f"ceiling {redis_tier['msgPerSecCeiling']} msg/s)")


# ----------------------------------------------------------------------------
# Saturation report parser
# ----------------------------------------------------------------------------
# Looks for a scorecard-style markdown table with columns
#   Step | VUs | Msg/s ... | p95 ... | CPU% ... | Status/Decision
# Handles the common shapes we emit in benchmarks/docs/saturation-run-*.md.
# Returns a list of dicts. If nothing parses, returns []. Never raises.
def _strip_md(cell):
    """Remove markdown emphasis (**bold**, *italic*, backticks) and whitespace."""
    s = str(cell).strip()
    s = re.sub(r"\*+", "", s)
    s = s.replace("`", "")
    return s.strip()


def parse_saturation_report(md_path):
    try:
        text = Path(md_path).read_text()
    except Exception:
        return {"header": {}, "steps": [], "parseError": f"could not read {md_path}"}

    steps = []
    header = {
        "reportPath": str(md_path),
    }
    # Run ID and date from filename: saturation-run-<RUN_ID>-<DATE>[-tag].md
    m = re.search(r"saturation-run-(\d+)-(\d{4}-\d{2}-\d{2})", Path(md_path).name)
    if m:
        header["runId"] = m.group(1)
        header["runDate"] = m.group(2)

    # Extract first markdown table whose header contains both "VUs" and "p95"
    lines = text.splitlines()
    i = 0
    while i < len(lines):
        line = lines[i]
        if line.strip().startswith("|") and "VUs" in line and "p95" in line:
            # Check next line is separator (| --- | --- | ...)
            if i + 1 < len(lines) and set(lines[i + 1].replace("|", "").strip()) <= set("-: "):
                header_cells = [_strip_md(c) for c in line.strip("|").split("|")]
                header["columns"] = header_cells
                # Index of known columns (best-effort, case-insensitive)
                col_idx = {c.lower(): j for j, c in enumerate(header_cells)}
                # Read rows until a non-table line
                j = i + 2
                while j < len(lines) and lines[j].strip().startswith("|"):
                    row_cells = [_strip_md(c) for c in lines[j].strip("|").split("|")]
                    if len(row_cells) != len(header_cells):
                        j += 1
                        continue
                    step = {}
                    for k, v in enumerate(row_cells):
                        step[header_cells[k]] = v
                    steps.append(step)
                    j += 1
                break  # use the first matching table only
        i += 1

    if not steps:
        header["parseError"] = "no matching table found"
    return {"header": header, "steps": steps}

# ============================================================================
# Derive values from infra-snapshot (new schema with requests+limits+placement)
# ============================================================================

infra_hash = infra.get("hash", "unknown")
infra_captured = infra.get("capturedAt", "unknown")[:10]

# Runtime
rt = infra["runtimeDefaults"]
runtime_cpu_lim = rt.get("cpuLimitCores") or 4
runtime_cpu_req = rt.get("cpuRequestCores") or 1
runtime_mem_lim = rt.get("memoryLimitGi") or 4
runtime_mem_req = rt.get("memoryRequestGi") or 2
runtime_image = rt.get("image", "unknown")
runtime_node_pool = rt.get("placement", {}).get("nodePool", "unknown")
runtime_node_instance = rt.get("placement", {}).get("instanceType", "unknown")
runtime_hpa_min = rt.get("hpaMinReplicas", 1)
runtime_hpa_max = rt.get("hpaMaxReplicas", 1)

# Mongo
mg = infra["mongodb"]
mongo_replicas = mg["replicas"]
mongo_cpu_req = mg.get("cpuRequestCores") or 0
# If no CPU limit set, use node CPU as effective limit (container is unbounded)
_mongo_node_type = mg.get("placement", {}).get("instanceType", "")
_mongo_node_cpu = next(
    (n["cpuPerNode"] for n in infra.get("nodes", [])
     if n.get("instanceType") == _mongo_node_type), None
) if _mongo_node_type else None
mongo_cpu_lim = mg.get("cpuLimitCores") or _mongo_node_cpu or (mongo_cpu_req * 4) or 1
mongo_mem_req = mg.get("memoryRequestGi") or 0
mongo_mem_lim = mg.get("memoryLimitGi") or 0
mongo_pvc_gi = mg["pvcSizeGi"]
mongo_disk_sku = mg.get("diskSku", "unknown")
mongo_iops_limit = mg.get("diskIopsLimit") or 500  # StandardSSD_LRS default: 500 IOPS
mongo_image = mg.get("image", "unknown")
mongo_node_pool = mg.get("placement", {}).get("nodePool", "unknown")
mongo_node_instance = mg.get("placement", {}).get("instanceType", "unknown")

# Redis
rd = infra["redis"]
redis_master_replicas = rd.get("masterReplicas", 1)
redis_replica_replicas = rd.get("replicaReplicas", 0)
redis_total_replicas = rd.get("totalReplicas", redis_master_replicas + redis_replica_replicas)
redis_master = rd.get("master", {})
redis_master_cpu_req = redis_master.get("cpuRequestCores", rd.get("masterCpuLimitCores", 0))
redis_master_cpu_lim = redis_master.get("cpuLimitCores", rd.get("masterCpuLimitCores", 0))
redis_master_mem_req = redis_master.get("memoryRequestGi", rd.get("masterMemoryLimitGi", 0))
redis_master_mem_lim = redis_master.get("memoryLimitGi", rd.get("masterMemoryLimitGi", 0))
redis_replica_spec = rd.get("replica", {})
redis_rep_cpu_req = redis_replica_spec.get("cpuRequestCores", rd.get("replicaCpuLimitCores", 0))
redis_rep_cpu_lim = redis_replica_spec.get("cpuLimitCores", rd.get("replicaCpuLimitCores", 0))
redis_rep_mem_req = redis_replica_spec.get("memoryRequestGi", rd.get("replicaMemoryLimitGi", 0))
redis_rep_mem_lim = redis_replica_spec.get("memoryLimitGi", rd.get("replicaMemoryLimitGi", 0))
redis_node_pool = rd.get("placement", {}).get("nodePool", "unknown")
redis_node_instance = rd.get("placement", {}).get("instanceType", "unknown")

# Nodes
nodes_list = infra.get("nodes", [])

# Measurement-derived values (only if measurement present)
if meas:
    per_pod_steady_val = meas["perPodCapacity"]["maxSafeMsgPerSec"]
    per_pod_peak_val = per_pod_steady_val  # fallback when we don't store peak separately
    p95_at_knee = meas["perPodCapacity"].get("p95AtMaxMs", args.p95)
    wi = meas.get("workloadIntrinsic", {})
    mongo_cpu_per_msg = wi.get("mongoCpuMilliPerMsg")
    redis_cpu_per_msg = wi.get("redisCpuMilliPerMsg")
    mongo_iops_per_msg = wi.get("mongoIopsPerMsg", scenario["perMessageDatastoreCost"]["mongoWrites"])
    meas_captured = meas["capturedAt"][:10]
    meas_id = meas["measurementId"]
    meas_val_status = meas.get("validationStatus", "unknown")
    meas_infra_hash = meas.get("configKey", {}).get("infraSnapshotHash", "unknown")
    max_measured_pods = max(p["pods"] for p in meas["multiPodScaling"])
    multi_pod_scaling = meas["multiPodScaling"]
    # Measurement's runtime cpuLimit — used as denominator for runtime tier
    # re-scaling heuristic. Convert "2000m" / "2" / numeric to cores.
    _mrc = meas.get("configKey", {}).get("cpuLimit", "2")
    try:
        if isinstance(_mrc, str) and _mrc.endswith("m"):
            meas_runtime_cpu_lim = float(_mrc[:-1]) / 1000
        else:
            meas_runtime_cpu_lim = float(_mrc)
    except Exception:
        meas_runtime_cpu_lim = 2.0  # conservative fallback
else:
    per_pod_steady_val = None
    per_pod_peak_val = None
    p95_at_knee = None
    mongo_cpu_per_msg = None
    redis_cpu_per_msg = None
    mongo_iops_per_msg = scenario["perMessageDatastoreCost"]["mongoWrites"]
    meas_captured = "NONE"
    meas_id = "NONE"
    meas_val_status = "missing"
    meas_infra_hash = "NONE"
    max_measured_pods = 1
    multi_pod_scaling = []
    meas_runtime_cpu_lim = 2.0

# ============================================================================
# Styling
# ============================================================================

YELLOW = PatternFill("solid", fgColor="FFF2CC")
GREEN = PatternFill("solid", fgColor="C6EFCE")
LIGHT_GREEN = PatternFill("solid", fgColor="E2EFDA")
GRAY = PatternFill("solid", fgColor="E7E6E6")
AMBER = PatternFill("solid", fgColor="FFEB9C")
RED = PatternFill("solid", fgColor="FFC7CE")
BLUE_HEADER = PatternFill("solid", fgColor="4472C4")
LIGHT_BLUE = PatternFill("solid", fgColor="D9E1F2")

BOLD = Font(bold=True)
HEADER_FONT = Font(bold=True, color="FFFFFF")
ITALIC = Font(italic=True, color="595959")
MONO = Font(name="Menlo", size=8)
BIG_NUM = Font(bold=True, size=14)

thin = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def write_headers(ws, row, values):
    for i, v in enumerate(values, start=1):
        c = ws.cell(row=row, column=i, value=v)
        c.fill = BLUE_HEADER
        c.font = HEADER_FONT
        c.border = BORDER
        c.alignment = Alignment(horizontal="center", wrap_text=True)


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def put(ws, row, col, value, *, fill=None, font=None, number_format=None, border=True, align=None):
    c = ws.cell(row=row, column=col, value=value)
    if fill: c.fill = fill
    if font: c.font = font
    if number_format: c.number_format = number_format
    if border: c.border = BORDER
    if align: c.alignment = align
    return c


# ============================================================================
# Build workbook
# ============================================================================

wb = Workbook()
wb.remove(wb.active)


def add_name(name, attr_text):
    wb.defined_names[name] = DefinedName(name, attr_text=attr_text)


# ----------------------------------------------------------------------------
# Tab: Controls (editable)
# ----------------------------------------------------------------------------
ws = wb.create_sheet("Controls")
ws["A1"] = "Sizing & Cost — Controls"
ws["A1"].font = Font(bold=True, size=14)
ws["A2"] = f"Scenario: {args.scenario}  |  Generated: {today}"
ws["A2"].font = ITALIC

write_headers(ws, 4, ["Input", "Value", "Unit", "Notes"])

put(ws, 5, 1, "Target msg/s", font=BOLD)
put(ws, 5, 2, args.target, fill=YELLOW)
put(ws, 5, 3, "msg/s")
put(ws, 5, 4, "Primary sizing input", font=ITALIC)
add_name("target_msgps", "Controls!$B$5")

put(ws, 6, 1, "p95 target", font=BOLD)
put(ws, 6, 2, args.p95, fill=YELLOW)
put(ws, 6, 3, "ms")
put(ws, 6, 4, "Latency SLO", font=ITALIC)

put(ws, 7, 1, "Utilization target", font=BOLD)
put(ws, 7, 2, args.util, fill=YELLOW, number_format="0.00")
put(ws, 7, 3, "ratio")
put(ws, 7, 4, "0.70 = 70%; F-1.1 uses this", font=ITALIC)
add_name("util_target", "Controls!$B$7")

put(ws, 8, 1, "Report date", font=BOLD)
put(ws, 8, 2, today, fill=GRAY)
put(ws, 8, 4, "Live age checks on Provenance/Warnings use this", font=ITALIC)
add_name("today_date", "Controls!$B$8")

# ─── Capacity Planning Inputs ───────────────────────────────────────────────
put(ws, 10, 1, "CAPACITY PLANNING INPUTS", font=BOLD, fill=LIGHT_BLUE)

put(ws, 11, 1, "Peak:Avg ratio (burst factor)", font=BOLD)
put(ws, 11, 2, 3.0, fill=YELLOW, number_format="0.0")
put(ws, 11, 3, "×")
put(ws, 11, 4, "Industry 2-5×. Chat platforms typically 3×. Peak provisioning denominator.", font=ITALIC)
add_name("burst_factor", "Controls!$B$11")

put(ws, 12, 1, "Availability headroom (N+K)", font=BOLD)
put(ws, 12, 2, 1, fill=YELLOW, number_format="0")
put(ws, 12, 3, "pods")
put(ws, 12, 4, "Extra pods beyond calculated. N+1 = survive single pod failure.", font=ITALIC)
add_name("availability_headroom_pods", "Controls!$B$12")

put(ws, 13, 1, "HPA scale-up lag", font=BOLD)
put(ws, 13, 2, 60, fill=YELLOW, number_format="0")
put(ws, 13, 3, "seconds")
put(ws, 13, 4, "Time for K8s to schedule+start new pod. Need spare capacity during.", font=ITALIC)
add_name("hpa_scaleup_lag_sec", "Controls!$B$13")

put(ws, 14, 1, "Active hours per day", font=BOLD)
put(ws, 14, 2, 12, fill=YELLOW, number_format="0")
put(ws, 14, 3, "hours")
put(ws, 14, 4, "Business hours with traffic. Affects storage + LLM cost. 24 = always-on.", font=ITALIC)
add_name("active_hours_per_day", "Controls!$B$14")

put(ws, 15, 1, "Active days per week", font=BOLD)
put(ws, 15, 2, 5, fill=YELLOW, number_format="0")
put(ws, 15, 3, "days")
put(ws, 15, 4, "Weekdays. Affects monthly message volume. 7 = no weekend drop.", font=ITALIC)
add_name("active_days_per_week", "Controls!$B$15")

put(ws, 16, 1, "Growth rate (monthly)", font=BOLD)
put(ws, 16, 2, 0.10, fill=YELLOW, number_format="0%")
put(ws, 16, 3, "% per month")
put(ws, 16, 4, "Compound monthly growth for projections. 10% = ~3.1× per year.", font=ITALIC)
add_name("growth_rate_monthly", "Controls!$B$16")

put(ws, 17, 1, "Avg messages per conversation", font=BOLD)
put(ws, 17, 2, msgs_per_session, fill=YELLOW, number_format="0")
put(ws, 17, 3, "msgs")
put(ws, 17, 4, "From turnStructure (creates+followups). For cost-per-conversation.", font=ITALIC)
add_name("msgs_per_conversation", "Controls!$B$17")

put(ws, 18, 1, "Node operational overhead", font=BOLD)
put(ws, 18, 2, 0.12, fill=YELLOW, number_format="0%")
put(ws, 18, 3, "ratio")
put(ws, 18, 4, "kubelet/daemonsets eat 10-15% of node CPU/memory. Subtract from available.", font=ITALIC)
add_name("node_overhead_pct", "Controls!$B$18")

# ─── Tier-selection dropdowns (LIVE CALCULATOR) ──────────────────────────────
# Each dropdown is a data-validation cell. Changing it makes baseline costs
# recompute via the INDEX/MATCH formulas on BaselineSource.
# Default values are pre-filled to the auto-recommended tier.
if baseline_tiers_catalog is not None:
    # Compute auto-recommendations so defaults reflect sensible choices
    mongo_tier_auto_name = mongo_tier_name if mongo_tier else "small"
    redis_tier_auto_name = redis_tier_name if redis_tier else "small"
    # Runtime auto-pick: smallest tier whose per-pod ceiling covers
    # target / expected_pods. Without a good pod estimate yet, default to
    # matching the measurement's tier shape (dev-compressed for current
    # measurement at cpuLimit=2).
    runtime_tiers_catalog = baseline_tiers_catalog.get("runtime_tiers", {})
    runtime_tier_auto_name = "dev-compressed"  # matches measurement 7342076
    if runtime_tiers_catalog and meas is not None:
        # try to match by cpuLimitCores from measurement.configKey
        meas_cpu_lim = meas.get("configKey", {}).get("cpuLimit", "")
        try:
            mcl = int(str(meas_cpu_lim).replace("m", "")) / (1000 if "m" in str(meas_cpu_lim) else 1)
            for tn, ts in runtime_tiers_catalog.items():
                if ts["cpuLimitCores"] == mcl:
                    runtime_tier_auto_name = tn
                    break
        except Exception:
            pass

    put(ws, 20, 1, "TIER SELECTION (live calculator)", font=BOLD, fill=LIGHT_BLUE)

    put(ws, 21, 1, "Runtime tier", font=BOLD)
    put(ws, 21, 2, runtime_tier_auto_name, fill=YELLOW)
    put(ws, 21, 3, "tier", font=ITALIC)
    put(ws, 21, 4,
        "Dropdown → changes per-pod capacity + baseline runtime cost",
        font=ITALIC)
    add_name("runtime_tier_selected", "Controls!$B$21")
    if runtime_tier_table_loaded:
        rt_names = ",".join(runtime_tiers_catalog.keys())
        dv_rt = DataValidation(type="list", formula1=f'"{rt_names}"', allow_blank=False)
        dv_rt.error = "Pick one of the runtime tiers listed on BaselineSource."
        dv_rt.errorTitle = "Invalid tier"
        dv_rt.prompt = "Select a runtime tier"
        ws.add_data_validation(dv_rt)
        dv_rt.add("B21")

    put(ws, 22, 1, "MongoDB tier", font=BOLD)
    put(ws, 22, 2, mongo_tier_auto_name, fill=YELLOW)
    put(ws, 22, 3, "tier", font=ITALIC)
    put(ws, 22, 4,
        "Dropdown → changes baseline Mongo replica specs + PVC floor",
        font=ITALIC)
    add_name("mongo_tier_selected", "Controls!$B$22")
    if mongo_tier_table_loaded:
        m_names = ",".join(baseline_tiers_catalog["mongodb_tiers"].keys())
        dv_m = DataValidation(type="list", formula1=f'"{m_names}"', allow_blank=False)
        dv_m.error = "Pick one of the Mongo tiers listed on BaselineSource."
        dv_m.errorTitle = "Invalid tier"
        dv_m.prompt = "Select a Mongo tier"
        ws.add_data_validation(dv_m)
        dv_m.add("B22")

    put(ws, 23, 1, "Redis tier", font=BOLD)
    put(ws, 23, 2, redis_tier_auto_name, fill=YELLOW)
    put(ws, 23, 3, "tier", font=ITALIC)
    put(ws, 23, 4,
        "Dropdown → changes Redis master+replica specs",
        font=ITALIC)
    add_name("redis_tier_selected", "Controls!$B$23")
    if redis_tier_table_loaded:
        r_names = ",".join(baseline_tiers_catalog["redis_tiers"].keys())
        dv_r = DataValidation(type="list", formula1=f'"{r_names}"', allow_blank=False)
        dv_r.error = "Pick one of the Redis tiers listed on BaselineSource."
        dv_r.errorTitle = "Invalid tier"
        dv_r.prompt = "Select a Redis tier"
        ws.add_data_validation(dv_r)
        dv_r.add("B23")

    # Auto-recommendation hints (what would the auto-picker choose?)
    put(ws, 25, 1, "AUTO-RECOMMENDED TIERS (based on target msg/s)",
        font=BOLD, fill=LIGHT_BLUE)
    put(ws, 26, 1, "Runtime auto", font=ITALIC)
    put(ws, 26, 2, runtime_tier_auto_name, fill=GRAY, font=ITALIC)
    put(ws, 26, 4, "Matches measurement tier. Override at B21.", font=ITALIC)
    put(ws, 27, 1, "Mongo auto", font=ITALIC)
    put(ws, 27, 2,
        "=INDEX(mongo_tier_names_range, MATCH(TRUE, mongo_tier_ceiling_col>=target_msgps, 0))",
        fill=GRAY, font=ITALIC)
    put(ws, 27, 4, "Smallest tier covering target. Override at B22.", font=ITALIC)
    put(ws, 28, 1, "Redis auto", font=ITALIC)
    put(ws, 28, 2,
        "=INDEX(redis_tier_names_range, MATCH(TRUE, redis_tier_ceiling_col>=target_msgps, 0))",
        fill=GRAY, font=ITALIC)
    put(ws, 28, 4, "Smallest tier covering target. Override at B23.", font=ITALIC)

    # Mode indicator — SNAPSHOT vs WHAT-IF
    put(ws, 30, 1, "MODE", font=BOLD, fill=LIGHT_BLUE)
    put(ws, 31, 1, "Snapshot tiers (defaults)", font=ITALIC)
    put(ws, 31, 2,
        f'"{runtime_tier_auto_name} / {mongo_tier_auto_name} / {redis_tier_auto_name}"',
        fill=GRAY, font=ITALIC)
    put(ws, 32, 1, "Current mode", font=BOLD)
    put(ws, 32, 2,
        '=IF(AND(runtime_tier_selected="' + runtime_tier_auto_name + '", '
        'mongo_tier_selected="' + mongo_tier_auto_name + '", '
        'redis_tier_selected="' + redis_tier_auto_name + '"), '
        '"SNAPSHOT (at defaults)", "WHAT-IF (edited from defaults)")',
        fill=GREEN, font=BOLD)
    add_name("calc_mode", "Controls!$B$32")

set_col_widths(ws, [28, 20, 10, 60])

# ----------------------------------------------------------------------------
# Tab: RateCard (editable with Original column)
# ----------------------------------------------------------------------------
ws = wb.create_sheet("RateCard")
ws["A1"] = "Rate Card — Unit Prices"
ws["A1"].font = Font(bold=True, size=14)
ws["A2"] = f"Source: {args.rate_card}  |  Captured: {rate.get('capturedAt', 'unknown')[:10]}"
ws["A2"].font = ITALIC
ws["A3"] = "Yellow = editable; gray 'Original' column retains seed value for reference."
ws["A3"].font = ITALIC

write_headers(ws, 5, ["Category", "Key", "Value (editable)", "Original", "Unit", "Notes"])

rate_rows = [
    ("compute", "cpuCorePerHour", rate["compute"]["cpuCorePerHour"], "$/core-hr", "Canonical compute rate"),
    ("compute", "memoryGBPerHour", rate["compute"]["memoryGBPerHour"], "$/GB-hr", "Canonical memory rate"),
    ("compute", "nodePoolMinCores", rate["compute"]["nodePoolMinCores"], "cores", "Node-pool floor"),
    ("compute", "nodePoolMinMemoryGB", rate["compute"].get("nodePoolMinMemoryGB", 48), "GB", ""),
    ("mongodb", "managedServiceUpliftMultiplier", rate["mongodb"]["managedServiceUpliftMultiplier"], "x", "1.0 = self-hosted"),
    ("mongodb", "storageGBPerMonth", rate["mongodb"]["storageGBPerMonth"], "$/GB-mo", ""),
    ("mongodb", "iopsPerMonth", rate["mongodb"]["iopsPerMonth"], "$/IOPS-mo", ""),
    ("mongodb", "replicaBaseFeePerMonth", rate["mongodb"]["replicaBaseFeePerMonth"], "$/replica-mo", ""),
    ("mongodb", "retentionDays", rate["mongodb"]["retentionDays"], "days", "Affects storage sizing"),
    ("mongodb", "avgDocSizeKB", rate["mongodb"]["avgDocSizeKB"], "KB", ""),
    ("redis", "managedServiceUpliftMultiplier", rate["redis"]["managedServiceUpliftMultiplier"], "x", ""),
    ("redis", "replicaBaseFeePerMonth", rate["redis"]["replicaBaseFeePerMonth"], "$/replica-mo", ""),
    ("redis", "avgSessionSizeKB", rate["redis"]["avgSessionSizeKB"], "KB", "For redis_memory_gb"),
    ("ingress", "ingressBaseFeePerMonth", rate["ingressBaseFeePerMonth"], "$/mo", ""),
    ("egress", "egressGBPerMonth", rate["egressGBPerMonth"], "$/GB-mo", ""),
    ("llm.mock", "inputPerM", rate["llm"]["mock"]["inputPerM"], "$/M tokens", "0 for mock"),
    ("llm.mock", "outputPerM", rate["llm"]["mock"]["outputPerM"], "$/M tokens", "0 for mock"),
]
rate_key_to_row = {}
for i, (cat, key, val, unit, note) in enumerate(rate_rows):
    r = 6 + i
    put(ws, r, 1, cat)
    put(ws, r, 2, key, font=BOLD)
    put(ws, r, 3, val, fill=YELLOW, number_format="0.####")
    put(ws, r, 4, val, fill=GRAY, number_format="0.####", font=ITALIC)
    put(ws, r, 5, unit)
    put(ws, r, 6, note, font=ITALIC)
    rate_key_to_row[key] = r

# Named cells — values live in column C
add_name("cpu_per_hr", f"RateCard!$C${rate_key_to_row['cpuCorePerHour']}")
add_name("mem_per_hr", f"RateCard!$C${rate_key_to_row['memoryGBPerHour']}")
add_name("node_min_cores", f"RateCard!$C${rate_key_to_row['nodePoolMinCores']}")
add_name("node_min_mem", f"RateCard!$C${rate_key_to_row['nodePoolMinMemoryGB']}")
# Uplifts: first managedServiceUpliftMultiplier = mongo (row 10), second = redis (row 16)
mongo_uplift_row = [i for i, (c, k, *_) in enumerate(rate_rows) if c == "mongodb" and k == "managedServiceUpliftMultiplier"][0] + 6
redis_uplift_row = [i for i, (c, k, *_) in enumerate(rate_rows) if c == "redis" and k == "managedServiceUpliftMultiplier"][0] + 6
add_name("mongo_uplift", f"RateCard!$C${mongo_uplift_row}")
add_name("redis_uplift", f"RateCard!$C${redis_uplift_row}")
add_name("mongo_storage_per_gb", f"RateCard!$C${rate_key_to_row['storageGBPerMonth']}")
add_name("mongo_iops_per_mo", f"RateCard!$C${rate_key_to_row['iopsPerMonth']}")
mongo_replica_fee_row = [i for i, (c, k, *_) in enumerate(rate_rows) if c == "mongodb" and k == "replicaBaseFeePerMonth"][0] + 6
redis_replica_fee_row = [i for i, (c, k, *_) in enumerate(rate_rows) if c == "redis" and k == "replicaBaseFeePerMonth"][0] + 6
add_name("mongo_replica_fee", f"RateCard!$C${mongo_replica_fee_row}")
add_name("redis_replica_fee", f"RateCard!$C${redis_replica_fee_row}")
add_name("retention_days", f"RateCard!$C${rate_key_to_row['retentionDays']}")
add_name("avg_doc_kb", f"RateCard!$C${rate_key_to_row['avgDocSizeKB']}")
add_name("session_kb", f"RateCard!$C${rate_key_to_row['avgSessionSizeKB']}")
add_name("ingress_fee", f"RateCard!$C${rate_key_to_row['ingressBaseFeePerMonth']}")
add_name("egress_per_gb", f"RateCard!$C${rate_key_to_row['egressGBPerMonth']}")
add_name("llm_input_per_m", f"RateCard!$C${rate_key_to_row['inputPerM']}")
add_name("llm_output_per_m", f"RateCard!$C${rate_key_to_row['outputPerM']}")

set_col_widths(ws, [14, 36, 14, 12, 14, 40])

# ----------------------------------------------------------------------------
# Tab: Scenario (editable)
# ----------------------------------------------------------------------------
ws = wb.create_sheet("Scenario")
ws["A1"] = "Scenario Definition"
ws["A1"].font = Font(bold=True, size=14)
ws["A2"] = f"Source: {args.scenario_file}"
ws["A2"].font = ITALIC

write_headers(ws, 4, ["Key", "Value", "Unit", "Notes"])

scn_rows = [
    ("scenarioId", scenario["scenarioId"], "", ""),
    ("llmProfile.mode", scenario["llmProfile"]["mode"], "", ""),
    ("llmProfile.model", scenario["llmProfile"]["model"], "", ""),
    ("llmProfile.delayMs", scenario["llmProfile"]["delayMs"], "ms", ""),
    ("llmProfile.tokensInPerTurn", scenario["llmProfile"]["tokensInPerTurn"], "tokens", "0 for mock"),
    ("llmProfile.tokensOutPerTurn", scenario["llmProfile"]["tokensOutPerTurn"], "tokens", "0 for mock"),
    ("requestShape.interMessageDelayMs", scenario["requestShape"]["interMessageDelayMs"], "ms", ""),
    ("requestShape.payloadBytesIn", scenario["requestShape"]["payloadBytesIn"], "bytes", ""),
    ("requestShape.payloadBytesOut", scenario["requestShape"]["payloadBytesOut"], "bytes", "Used in egress"),
    ("requestShape.msgsPerSession", msgs_per_session, "msgs", "creates + followups; for Little's Law session sizing"),
    ("perMessageDatastoreCost.mongoReads", scenario["perMessageDatastoreCost"]["mongoReads"], "per msg", "Empirical from profiler"),
    ("perMessageDatastoreCost.mongoWrites", scenario["perMessageDatastoreCost"]["mongoWrites"], "per msg", "Affects storage + IOPS"),
    ("perMessageDatastoreCost.mongoWriteDocSizeKB", scenario["perMessageDatastoreCost"].get("mongoWriteDocSizeKB", 4), "KB", "Avg doc size for storage calc"),
    ("perMessageDatastoreCost.redisOps", scenario["perMessageDatastoreCost"]["redisOps"], "per msg", "reads + writes + scripts combined"),
    ("sessionPattern.sessionTTLMin", scenario["sessionPattern"]["sessionTTLMin"], "min", "Redis memory sizing"),
    ("sessionPattern.maxInMemorySessionsPerPod", scenario["sessionPattern"]["maxInMemorySessionsPerPod"], "per pod", "Redis cache cap"),
    ("tenantProfile.tier", scenario["tenantProfile"]["tier"], "", ""),
]
scn_key_to_row = {}
for i, (key, val, unit, note) in enumerate(scn_rows):
    r = 5 + i
    put(ws, r, 1, key, font=BOLD)
    put(ws, r, 2, val, fill=YELLOW)
    put(ws, r, 3, unit)
    put(ws, r, 4, note, font=ITALIC)
    scn_key_to_row[key] = r

add_name("llm_tokens_in", f"Scenario!$B${scn_key_to_row['llmProfile.tokensInPerTurn']}")
add_name("llm_tokens_out", f"Scenario!$B${scn_key_to_row['llmProfile.tokensOutPerTurn']}")
add_name("payload_bytes_out", f"Scenario!$B${scn_key_to_row['requestShape.payloadBytesOut']}")
add_name("msgs_per_session", f"Scenario!$B${scn_key_to_row['requestShape.msgsPerSession']}")
add_name("mongo_writes_per_msg", f"Scenario!$B${scn_key_to_row['perMessageDatastoreCost.mongoWrites']}")
add_name("mongo_write_doc_kb", f"Scenario!$B${scn_key_to_row['perMessageDatastoreCost.mongoWriteDocSizeKB']}")
add_name("session_ttl_min", f"Scenario!$B${scn_key_to_row['sessionPattern.sessionTTLMin']}")
add_name("session_cap_per_pod", f"Scenario!$B${scn_key_to_row['sessionPattern.maxInMemorySessionsPerPod']}")

set_col_widths(ws, [44, 20, 10, 40])

# ----------------------------------------------------------------------------
# Tab: Infra — EXPANDED: nodes + requests+limits + placement
# ----------------------------------------------------------------------------
ws = wb.create_sheet("Infra")
ws["A1"] = "Infra Snapshot (live kubectl)"
ws["A1"].font = Font(bold=True, size=14)
ws["A2"] = f"Source: {args.infra}  |  Captured: {infra_captured}  |  Cluster: {infra.get('cluster', 'unknown')}  |  Hash: {infra_hash[:23]}…"
ws["A2"].font = ITALIC

# --- Node inventory ---
put(ws, 4, 1, "NODE POOLS", font=Font(bold=True, size=12), fill=LIGHT_BLUE)
write_headers(ws, 5, ["Pool", "Instance type", "Count", "CPU per node", "Memory per node (GiB)", "OS", "Kubelet"])
for i, n in enumerate(nodes_list):
    r = 6 + i
    put(ws, r, 1, n.get("pool", "?"), font=BOLD)
    put(ws, r, 2, n.get("instanceType", "?"), fill=GRAY)
    put(ws, r, 3, n.get("count", 0), fill=GRAY)
    put(ws, r, 4, n.get("cpuPerNode", "?"), fill=GRAY)
    put(ws, r, 5, n.get("memoryPerNodeGi", 0), fill=GRAY)
    put(ws, r, 6, n.get("os", "?"), font=ITALIC)
    put(ws, r, 7, n.get("kubeletVersion", "?"), font=ITALIC)

node_rows_end = 6 + len(nodes_list)

# --- Runtime ---
rt_header_row = node_rows_end + 2
put(ws, rt_header_row, 1, "RUNTIME POD", font=Font(bold=True, size=12), fill=LIGHT_BLUE)
write_headers(ws, rt_header_row + 1, ["Key", "Value", "Unit", "Notes"])
rt_block_rows = [
    ("hpaMinReplicas", runtime_hpa_min, "pods", "HA minimum — baseline always-on pod count"),
    ("hpaMaxReplicas", runtime_hpa_max, "pods", "HPA ceiling"),
    ("cpuRequestCores", runtime_cpu_req, "cores", "HPA scaling trigger uses this"),
    ("cpuLimitCores", runtime_cpu_lim, "cores", "CFS hard ceiling"),
    ("memoryRequestGi", runtime_mem_req, "GiB", ""),
    ("memoryLimitGi", runtime_mem_lim, "GiB", "OOM kill threshold"),
    ("image", runtime_image, "", "Container image tag"),
    ("placement.nodePool", runtime_node_pool, "", "AKS node pool"),
    ("placement.instanceType", runtime_node_instance, "", "VM SKU"),
]
for i, (key, val, unit, note) in enumerate(rt_block_rows):
    r = rt_header_row + 2 + i
    put(ws, r, 1, key, font=BOLD)
    put(ws, r, 2, val, fill=YELLOW)
    put(ws, r, 3, unit)
    put(ws, r, 4, note, font=ITALIC)
    # register named cells for numeric ones
    if key == "hpaMinReplicas":
        add_name("runtime_hpa_min", f"Infra!$B${r}")
    elif key == "cpuLimitCores":
        add_name("runtime_cpu_limit", f"Infra!$B${r}")
    elif key == "memoryLimitGi":
        add_name("runtime_mem_limit", f"Infra!$B${r}")
    elif key == "cpuRequestCores":
        add_name("runtime_cpu_request", f"Infra!$B${r}")
    elif key == "memoryRequestGi":
        add_name("runtime_mem_request", f"Infra!$B${r}")

# --- MongoDB ---
mongo_header_row = rt_header_row + 2 + len(rt_block_rows) + 1
put(ws, mongo_header_row, 1, "MONGODB STATEFULSET", font=Font(bold=True, size=12), fill=LIGHT_BLUE)
write_headers(ws, mongo_header_row + 1, ["Key", "Value", "Unit", "Notes"])
mongo_block_rows = [
    ("replicas", mongo_replicas, "pods", "Replica set quorum"),
    ("cpuRequestCores", mongo_cpu_req, "cores", ""),
    ("cpuLimitCores", mongo_cpu_lim, "cores", ""),
    ("memoryRequestGi", mongo_mem_req, "GiB", ""),
    ("memoryLimitGi", mongo_mem_lim, "GiB", ""),
    ("pvcSizeGi", mongo_pvc_gi, "GiB", "Per-replica PVC"),
    ("diskSku", mongo_disk_sku, "", "StorageClass name"),
    ("diskIopsLimit", mongo_iops_limit, "IOPS", "Derived from SKU + PVC size"),
    ("image", mongo_image, "", ""),
    ("placement.nodePool", mongo_node_pool, "", "AKS node pool"),
    ("placement.instanceType", mongo_node_instance, "", "VM SKU"),
]
for i, (key, val, unit, note) in enumerate(mongo_block_rows):
    r = mongo_header_row + 2 + i
    put(ws, r, 1, key, font=BOLD)
    put(ws, r, 2, val, fill=YELLOW)
    put(ws, r, 3, unit)
    put(ws, r, 4, note, font=ITALIC)
    if key == "replicas": add_name("mongo_replicas", f"Infra!$B${r}")
    elif key == "cpuLimitCores": add_name("mongo_cpu_limit", f"Infra!$B${r}")
    elif key == "memoryLimitGi": add_name("mongo_mem_limit", f"Infra!$B${r}")
    elif key == "pvcSizeGi": add_name("mongo_pvc_gib", f"Infra!$B${r}")
    elif key == "diskIopsLimit": add_name("mongo_iops_limit", f"Infra!$B${r}")

# --- Redis ---
redis_header_row = mongo_header_row + 2 + len(mongo_block_rows) + 1
put(ws, redis_header_row, 1, "REDIS STATEFULSETS (master + replicas)", font=Font(bold=True, size=12), fill=LIGHT_BLUE)
write_headers(ws, redis_header_row + 1, ["Key", "Value", "Unit", "Notes"])
redis_block_rows = [
    ("topology", "master + replicas", "", ""),
    ("masterReplicas", redis_master_replicas, "pods", ""),
    ("replicaReplicas", redis_replica_replicas, "pods", ""),
    ("totalReplicas", redis_total_replicas, "pods", "Master + Replicas"),
    ("master.cpuRequestCores", redis_master_cpu_req, "cores", ""),
    ("master.cpuLimitCores", redis_master_cpu_lim, "cores", ""),
    ("master.memoryRequestGi", redis_master_mem_req, "GiB", ""),
    ("master.memoryLimitGi", redis_master_mem_lim, "GiB", ""),
    ("replica.cpuRequestCores", redis_rep_cpu_req, "cores", ""),
    ("replica.cpuLimitCores", redis_rep_cpu_lim, "cores", ""),
    ("replica.memoryRequestGi", redis_rep_mem_req, "GiB", ""),
    ("replica.memoryLimitGi", redis_rep_mem_lim, "GiB", ""),
    ("placement.nodePool", redis_node_pool, "", ""),
    ("placement.instanceType", redis_node_instance, "", ""),
]
for i, (key, val, unit, note) in enumerate(redis_block_rows):
    r = redis_header_row + 2 + i
    put(ws, r, 1, key, font=BOLD)
    put(ws, r, 2, val, fill=YELLOW)
    put(ws, r, 3, unit)
    put(ws, r, 4, note, font=ITALIC)
    if key == "totalReplicas": add_name("redis_replicas", f"Infra!$B${r}")
    elif key == "masterReplicas": add_name("redis_master_replicas", f"Infra!$B${r}")
    elif key == "replicaReplicas": add_name("redis_replica_replicas", f"Infra!$B${r}")
    elif key == "master.cpuLimitCores":
        add_name("redis_cpu_limit", f"Infra!$B${r}")  # legacy: master cpuLimit
        add_name("redis_master_cpu_limit", f"Infra!$B${r}")
    elif key == "master.memoryLimitGi":
        add_name("redis_mem_limit", f"Infra!$B${r}")  # legacy: master memLimit
        add_name("redis_master_mem_limit", f"Infra!$B${r}")
    elif key == "replica.cpuLimitCores":
        add_name("redis_replica_cpu_limit", f"Infra!$B${r}")
    elif key == "replica.memoryLimitGi":
        add_name("redis_replica_mem_limit", f"Infra!$B${r}")

# --- Required Infrastructure (formula-driven, updates with target msg/s) ---
req_header_row = redis_header_row + 2 + len(redis_block_rows) + 2
put(ws, req_header_row, 1, "REQUIRED INFRASTRUCTURE (auto-calculated from target msg/s)",
    font=Font(bold=True, size=12), fill=PatternFill("solid", fgColor="E2EFDA"))
put(ws, req_header_row + 1, 1, "← Changes when you edit Controls → Target msg/s",
    font=ITALIC)
write_headers(ws, req_header_row + 2, ["Metric", "Required", "Current", "Headroom", "Unit", "Status"])

req_rows = [
    ("Runtime pods",
     "=pods_required", f"={runtime_hpa_min}",
     f"=({runtime_hpa_max}-pods_required)", "pods",
     f'=IF(pods_required<={runtime_hpa_max},"OK","SCALE HPA")'),
    ("PRODUCTION pods (peak+HA)",
     "=pods_production", f"={runtime_hpa_max}",
     f"=({runtime_hpa_max}-pods_production)", "pods",
     f'=IF(pods_production<={runtime_hpa_max},"OK","SCALE HPA MAX")'),
    ("Runtime CPU per pod",
     f"={runtime_cpu_req}", f"={runtime_cpu_lim}",
     f"={runtime_cpu_lim}-{runtime_cpu_req}", "cores",
     f'=IF({runtime_cpu_req}<={runtime_cpu_lim},"OK","OVER LIMIT")'),
    ("Mongo CPU total (fleet)",
     "=mongo_cores_req", f"={mongo_cpu_lim * mongo_replicas}",
     f"={mongo_cpu_lim * mongo_replicas}-mongo_cores_req", "cores",
     f'=IF(mongo_cores_req<={mongo_cpu_lim * mongo_replicas},"OK","SCALE MONGO CPU")'),
    ("Mongo IOPS",
     "=mongo_iops_req", f"={mongo_iops_limit}",
     f"={mongo_iops_limit}-mongo_iops_req", "IOPS",
     f'=IF(mongo_iops_req<={mongo_iops_limit},"OK","UPGRADE DISK")'),
    ("Mongo storage (replicated)",
     "=mongo_storage_gb", f"={mongo_pvc_gi * mongo_replicas}",
     f"={mongo_pvc_gi * mongo_replicas}-mongo_storage_gb", "GB",
     f'=IF(mongo_storage_gb<={mongo_pvc_gi * mongo_replicas},"OK","EXPAND PVC")'),
    ("Redis CPU (fleet)",
     "=redis_cores_req",
     f"={redis_master_cpu_lim * redis_master_replicas + redis_rep_cpu_lim * redis_replica_replicas}",
     f"={redis_master_cpu_lim * redis_master_replicas + redis_rep_cpu_lim * redis_replica_replicas}-redis_cores_req",
     "cores",
     f'=IF(redis_cores_req<={redis_master_cpu_lim * redis_master_replicas + redis_rep_cpu_lim * redis_replica_replicas},"OK","SCALE REDIS")'),
    ("Redis memory (sessions)",
     "=redis_mem_gb",
     f"={redis_master_mem_lim * redis_master_replicas + redis_rep_mem_lim * redis_replica_replicas}",
     f"={redis_master_mem_lim * redis_master_replicas + redis_rep_mem_lim * redis_replica_replicas}-redis_mem_gb",
     "GB",
     f'=IF(redis_mem_gb<={redis_master_mem_lim * redis_master_replicas + redis_rep_mem_lim * redis_replica_replicas},"OK","SCALE REDIS MEM")'),
    ("Monthly cost",
     "=total_cost", "—", "—", "$/mo", "—"),
]

for i, (metric, required, current, headroom, unit, status) in enumerate(req_rows):
    r = req_header_row + 3 + i
    put(ws, r, 1, metric, font=BOLD)
    put(ws, r, 2, required, fill=PatternFill("solid", fgColor="E2EFDA"))
    put(ws, r, 3, current, fill=GRAY)
    put(ws, r, 4, headroom)
    put(ws, r, 5, unit)
    put(ws, r, 6, status)

set_col_widths(ws, [35, 28, 20, 20, 12, 20])

# ----------------------------------------------------------------------------
# Tab: BaselineSource — FULL tier catalog (lookup table) + live resolved specs
# ----------------------------------------------------------------------------
# This tab is a LOOKUP TABLE of all tier options. Baseline formulas use XLOOKUP
# to resolve the selected tier (dropdowns live on Controls) into concrete
# cpu/mem/pvc specs. This makes the xlsx a LIVE calculator — change a dropdown
# and baseline cost updates immediately.
ws = wb.create_sheet("BaselineSource")
ws["A1"] = "Baseline Tier Catalog — lookup table for live tier selection"
ws["A1"].font = Font(bold=True, size=14)
ws["A2"] = ("All tier specs listed below. Controls!B10-B12 dropdowns pick which tier to use; "
            "XLOOKUP formulas resolve specs live. Edit ceiling values to adjust auto-recommendation.")
ws["A2"].font = ITALIC

# Flags (mongo_tier_table_loaded, redis_tier_table_loaded, runtime_tier_table_loaded)
# are already set near line 140 based on baseline_tiers_catalog presence.
if baseline_tiers_catalog is None:
    ws["A4"] = "baseline-tiers.json MISSING — falling back to live infra specs for baseline."
    ws["A4"].font = Font(bold=True, color="FF0000")

# ─── Runtime tier table (new) ────────────────────────────────────────────────
runtime_table_start = 4
if runtime_tier_table_loaded:
    put(ws, runtime_table_start, 1, "RUNTIME TIERS", font=BOLD, fill=LIGHT_BLUE)
    write_headers(ws, runtime_table_start + 1, [
        "Tier name", "cpuReq (cores)", "cpuLim (cores)", "memReq (Gi)", "memLim (Gi)",
        "msg/s per pod (ceiling)", "Description"
    ])
    runtime_tiers = baseline_tiers_catalog["runtime_tiers"]
    runtime_tier_rows = {}  # tier_name -> row number
    for i, (name, spec) in enumerate(runtime_tiers.items()):
        r = runtime_table_start + 2 + i
        put(ws, r, 1, name, font=BOLD, fill=YELLOW)
        put(ws, r, 2, spec["cpuRequestCores"], fill=YELLOW)
        put(ws, r, 3, spec["cpuLimitCores"], fill=YELLOW)
        put(ws, r, 4, spec["memoryRequestGi"], fill=YELLOW)
        put(ws, r, 5, spec["memoryLimitGi"], fill=YELLOW)
        put(ws, r, 6, spec["msgPerSecPerPodCeiling"], fill=YELLOW)
        put(ws, r, 7, spec["description"], font=ITALIC)
        runtime_tier_rows[name] = r
    runtime_table_last_row = runtime_table_start + 1 + len(runtime_tiers)
    runtime_table_range_names = f"$A${runtime_table_start + 2}:$A${runtime_table_last_row}"
    runtime_table_range_all = f"$A${runtime_table_start + 2}:$G${runtime_table_last_row}"
    # Named ranges for XLOOKUP lookup_array and return_array
    add_name("runtime_tier_names_range", f"BaselineSource!{runtime_table_range_names}")
    add_name("runtime_tier_cpureq_col", f"BaselineSource!$B${runtime_table_start + 2}:$B${runtime_table_last_row}")
    add_name("runtime_tier_cpulim_col", f"BaselineSource!$C${runtime_table_start + 2}:$C${runtime_table_last_row}")
    add_name("runtime_tier_memreq_col", f"BaselineSource!$D${runtime_table_start + 2}:$D${runtime_table_last_row}")
    add_name("runtime_tier_memlim_col", f"BaselineSource!$E${runtime_table_start + 2}:$E${runtime_table_last_row}")
    add_name("runtime_tier_ceiling_col", f"BaselineSource!$F${runtime_table_start + 2}:$F${runtime_table_last_row}")
else:
    runtime_table_last_row = runtime_table_start

# ─── Mongo tier table ────────────────────────────────────────────────────────
mongo_table_start = runtime_table_last_row + 2
if mongo_tier_table_loaded:
    put(ws, mongo_table_start, 1, "MONGODB TIERS", font=BOLD, fill=LIGHT_BLUE)
    write_headers(ws, mongo_table_start + 1, [
        "Tier name", "Replicas", "cpuLim (cores)", "memLim (Gi)", "PVC (Gi)",
        "Disk SKU", "IOPS", "msg/s ceiling"
    ])
    mongo_tiers = baseline_tiers_catalog["mongodb_tiers"]
    mongo_tier_rows = {}
    for i, (name, spec) in enumerate(mongo_tiers.items()):
        r = mongo_table_start + 2 + i
        put(ws, r, 1, name, font=BOLD, fill=YELLOW)
        put(ws, r, 2, spec["replicas"], fill=YELLOW)
        put(ws, r, 3, spec["cpuLimitCores"], fill=YELLOW)
        put(ws, r, 4, spec["memoryLimitGi"], fill=YELLOW)
        put(ws, r, 5, spec["pvcSizeGi"], fill=YELLOW)
        put(ws, r, 6, spec["diskSku"], font=ITALIC)
        put(ws, r, 7, spec.get("iopsProvisioned", 500), fill=YELLOW)
        put(ws, r, 8, spec["msgPerSecCeiling"], fill=YELLOW)
        mongo_tier_rows[name] = r
    mongo_table_last_row = mongo_table_start + 1 + len(mongo_tiers)
    add_name("mongo_tier_names_range", f"BaselineSource!$A${mongo_table_start + 2}:$A${mongo_table_last_row}")
    add_name("mongo_tier_replicas_col", f"BaselineSource!$B${mongo_table_start + 2}:$B${mongo_table_last_row}")
    add_name("mongo_tier_cpulim_col", f"BaselineSource!$C${mongo_table_start + 2}:$C${mongo_table_last_row}")
    add_name("mongo_tier_memlim_col", f"BaselineSource!$D${mongo_table_start + 2}:$D${mongo_table_last_row}")
    add_name("mongo_tier_pvc_col", f"BaselineSource!$E${mongo_table_start + 2}:$E${mongo_table_last_row}")
    add_name("mongo_tier_iops_col", f"BaselineSource!$G${mongo_table_start + 2}:$G${mongo_table_last_row}")
    add_name("mongo_tier_ceiling_col", f"BaselineSource!$H${mongo_table_start + 2}:$H${mongo_table_last_row}")
else:
    mongo_table_last_row = mongo_table_start

# ─── Redis tier table ────────────────────────────────────────────────────────
redis_table_start = mongo_table_last_row + 2
if redis_tier_table_loaded:
    put(ws, redis_table_start, 1, "REDIS TIERS", font=BOLD, fill=LIGHT_BLUE)
    write_headers(ws, redis_table_start + 1, [
        "Tier name", "Master replicas", "Replica replicas",
        "Master cpuLim", "Master memLim", "Replica cpuLim", "Replica memLim",
        "msg/s ceiling"
    ])
    redis_tiers = baseline_tiers_catalog["redis_tiers"]
    redis_tier_rows = {}
    for i, (name, spec) in enumerate(redis_tiers.items()):
        r = redis_table_start + 2 + i
        put(ws, r, 1, name, font=BOLD, fill=YELLOW)
        put(ws, r, 2, spec["masterReplicas"], fill=YELLOW)
        put(ws, r, 3, spec["replicaReplicas"], fill=YELLOW)
        put(ws, r, 4, spec["master"]["cpuLimitCores"], fill=YELLOW)
        put(ws, r, 5, spec["master"]["memoryLimitGi"], fill=YELLOW)
        put(ws, r, 6, spec["replica"]["cpuLimitCores"], fill=YELLOW)
        put(ws, r, 7, spec["replica"]["memoryLimitGi"], fill=YELLOW)
        put(ws, r, 8, spec["msgPerSecCeiling"], fill=YELLOW)
        redis_tier_rows[name] = r
    redis_table_last_row = redis_table_start + 1 + len(redis_tiers)
    add_name("redis_tier_names_range", f"BaselineSource!$A${redis_table_start + 2}:$A${redis_table_last_row}")
    add_name("redis_tier_master_replicas_col", f"BaselineSource!$B${redis_table_start + 2}:$B${redis_table_last_row}")
    add_name("redis_tier_replica_replicas_col", f"BaselineSource!$C${redis_table_start + 2}:$C${redis_table_last_row}")
    add_name("redis_tier_master_cpulim_col", f"BaselineSource!$D${redis_table_start + 2}:$D${redis_table_last_row}")
    add_name("redis_tier_master_memlim_col", f"BaselineSource!$E${redis_table_start + 2}:$E${redis_table_last_row}")
    add_name("redis_tier_replica_cpulim_col", f"BaselineSource!$F${redis_table_start + 2}:$F${redis_table_last_row}")
    add_name("redis_tier_replica_memlim_col", f"BaselineSource!$G${redis_table_start + 2}:$G${redis_table_last_row}")
    add_name("redis_tier_ceiling_col", f"BaselineSource!$H${redis_table_start + 2}:$H${redis_table_last_row}")
else:
    redis_table_last_row = redis_table_start

# ─── Resolved specs (live formulas using selected tier) ──────────────────────
# These named cells resolve the user's tier dropdown selection into concrete
# spec values. Baseline formulas reference THESE names.
# Uses INDEX/MATCH (broader formulas-lib compat than XLOOKUP).
resolved_start = redis_table_last_row + 3
put(ws, resolved_start, 1, "RESOLVED TIER SPECS (live from dropdowns)",
    font=BOLD, fill=LIGHT_BLUE)
write_headers(ws, resolved_start + 1, ["Resolved name", "Value", "Unit", "Tier source"])

resolved_rows = []
if runtime_tier_table_loaded:
    resolved_rows.extend([
        ("baseline_runtime_cpu_req",
            "=INDEX(runtime_tier_cpureq_col, MATCH(runtime_tier_selected, runtime_tier_names_range, 0))",
            "cores", "runtime_tier_cpureq[runtime_tier_selected]"),
        ("baseline_runtime_cpu_limit",
            "=INDEX(runtime_tier_cpulim_col, MATCH(runtime_tier_selected, runtime_tier_names_range, 0))",
            "cores", "runtime_tier_cpulim[runtime_tier_selected]"),
        ("baseline_runtime_mem_req",
            "=INDEX(runtime_tier_memreq_col, MATCH(runtime_tier_selected, runtime_tier_names_range, 0))",
            "Gi", "runtime_tier_memreq[runtime_tier_selected]"),
        ("baseline_runtime_mem_limit",
            "=INDEX(runtime_tier_memlim_col, MATCH(runtime_tier_selected, runtime_tier_names_range, 0))",
            "Gi", "runtime_tier_memlim[runtime_tier_selected]"),
        ("baseline_runtime_ceiling",
            "=INDEX(runtime_tier_ceiling_col, MATCH(runtime_tier_selected, runtime_tier_names_range, 0))",
            "msg/s/pod", "tier's per-pod ceiling"),
    ])
if mongo_tier_table_loaded:
    resolved_rows.extend([
        ("baseline_mongo_replicas",
            "=INDEX(mongo_tier_replicas_col, MATCH(mongo_tier_selected, mongo_tier_names_range, 0))",
            "pods", ""),
        ("baseline_mongo_cpu_limit",
            "=INDEX(mongo_tier_cpulim_col, MATCH(mongo_tier_selected, mongo_tier_names_range, 0))",
            "cores", "per pod"),
        ("baseline_mongo_mem_limit",
            "=INDEX(mongo_tier_memlim_col, MATCH(mongo_tier_selected, mongo_tier_names_range, 0))",
            "Gi", "per pod"),
        ("baseline_mongo_pvc_gib",
            "=INDEX(mongo_tier_pvc_col, MATCH(mongo_tier_selected, mongo_tier_names_range, 0))",
            "Gi", "per pod"),
        ("baseline_mongo_iops_limit",
            "=INDEX(mongo_tier_iops_col, MATCH(mongo_tier_selected, mongo_tier_names_range, 0))",
            "IOPS", "provisioned per volume"),
        ("baseline_mongo_ceiling",
            "=INDEX(mongo_tier_ceiling_col, MATCH(mongo_tier_selected, mongo_tier_names_range, 0))",
            "msg/s", ""),
    ])
if redis_tier_table_loaded:
    resolved_rows.extend([
        ("baseline_redis_master_replicas",
            "=INDEX(redis_tier_master_replicas_col, MATCH(redis_tier_selected, redis_tier_names_range, 0))",
            "pods", ""),
        ("baseline_redis_replica_replicas",
            "=INDEX(redis_tier_replica_replicas_col, MATCH(redis_tier_selected, redis_tier_names_range, 0))",
            "pods", ""),
        ("baseline_redis_master_cpu_limit",
            "=INDEX(redis_tier_master_cpulim_col, MATCH(redis_tier_selected, redis_tier_names_range, 0))",
            "cores", "master"),
        ("baseline_redis_master_mem_limit",
            "=INDEX(redis_tier_master_memlim_col, MATCH(redis_tier_selected, redis_tier_names_range, 0))",
            "Gi", "master"),
        ("baseline_redis_replica_cpu_limit",
            "=INDEX(redis_tier_replica_cpulim_col, MATCH(redis_tier_selected, redis_tier_names_range, 0))",
            "cores", "each replica"),
        ("baseline_redis_replica_mem_limit",
            "=INDEX(redis_tier_replica_memlim_col, MATCH(redis_tier_selected, redis_tier_names_range, 0))",
            "Gi", "each replica"),
        ("baseline_redis_total_replicas",
            "=baseline_redis_master_replicas + baseline_redis_replica_replicas",
            "pods", "master + replicas"),
        ("baseline_redis_ceiling",
            "=INDEX(redis_tier_ceiling_col, MATCH(redis_tier_selected, redis_tier_names_range, 0))",
            "msg/s", ""),
    ])

for i, (nm, formula, unit, note) in enumerate(resolved_rows):
    r = resolved_start + 2 + i
    put(ws, r, 1, nm, font=BOLD)
    put(ws, r, 2, formula, fill=GREEN, number_format="0.##")
    put(ws, r, 3, unit)
    put(ws, r, 4, note, font=ITALIC)
    add_name(nm, f"BaselineSource!$B${r}")

set_col_widths(ws, [34, 16, 14, 20, 16, 16, 16, 50])

# ----------------------------------------------------------------------------
# Tab: Measurement — workloadIntrinsic ratios + multiPodScaling
# ----------------------------------------------------------------------------
ws = wb.create_sheet("Measurement")
if meas:
    ws["A1"] = f"Measurement — {meas_id}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Captured: {meas_captured}  |  Status: {meas_val_status}  |  Infra hash at capture: {meas_infra_hash[:23]}…"
    ws["A2"].font = ITALIC

    write_headers(ws, 4, ["Field", "Value", "Unit", "Source"])

    meas_rows = [
        ("perPodCapacity.maxSafeMsgPerSec", per_pod_steady_val, "msg/s/pod", "Saturation knee, hold-window avg"),
        ("perPodCapacity.p95AtMaxMs", p95_at_knee, "ms", ""),
        ("workloadIntrinsic.mongoCpuMilliPerMsg", mongo_cpu_per_msg, "milli/msg-s", "Workload-intrinsic — infra-independent"),
        ("workloadIntrinsic.redisCpuMilliPerMsg", redis_cpu_per_msg, "milli/msg-s", "Workload-intrinsic"),
        ("workloadIntrinsic.mongoIopsPerMsg", mongo_iops_per_msg, "IOPS/msg-s", "Usually = scenario.perMessageDatastoreCost.mongoWrites"),
        ("scalingFactor (within measured)", 1.0, "ratio", "Linear within multiPodScaling[] range"),
        ("maxMeasuredPods", max_measured_pods, "pods", "Max pods in multiPodScaling[]"),
        ("measurementCapturedAt", meas_captured, "date", "For live age check"),
    ]
    meas_key_to_row = {}
    for i, (key, val, unit, src) in enumerate(meas_rows):
        r = 5 + i
        put(ws, r, 1, key, font=BOLD)
        put(ws, r, 2, val if val is not None else "NULL", fill=YELLOW, number_format="0.####")
        put(ws, r, 3, unit)
        put(ws, r, 4, src, font=ITALIC)
        meas_key_to_row[key] = r

    add_name("per_pod_steady", f"Measurement!$B${meas_key_to_row['perPodCapacity.maxSafeMsgPerSec']}")
    add_name("mongo_cpu_per_msg", f"Measurement!$B${meas_key_to_row['workloadIntrinsic.mongoCpuMilliPerMsg']}")
    add_name("redis_cpu_per_msg", f"Measurement!$B${meas_key_to_row['workloadIntrinsic.redisCpuMilliPerMsg']}")
    add_name("mongo_iops_per_msg", f"Measurement!$B${meas_key_to_row['workloadIntrinsic.mongoIopsPerMsg']}")
    add_name("scaling_factor", f"Measurement!$B${meas_key_to_row['scalingFactor (within measured)']}")
    add_name("max_measured_pods", f"Measurement!$B${meas_key_to_row['maxMeasuredPods']}")
    add_name("meas_captured_at", f"Measurement!$B${meas_key_to_row['measurementCapturedAt']}")

    # multiPodScaling raw
    scaling_header_row = 5 + len(meas_rows) + 1
    put(ws, scaling_header_row, 1, "Raw multiPodScaling[] (from measurement JSON)", font=BOLD, fill=LIGHT_BLUE)
    write_headers(ws, scaling_header_row + 1, ["Pods", "Total msg/s", "Scaling factor", "Shared bottleneck"])
    for i, p in enumerate(sorted(multi_pod_scaling, key=lambda x: x["pods"])):
        r = scaling_header_row + 2 + i
        put(ws, r, 1, p["pods"], font=BOLD)
        put(ws, r, 2, p["totalMsgPerSec"], number_format="0.##")
        put(ws, r, 3, p.get("scalingFactor", 1.0), number_format="0.###")
        put(ws, r, 4, p.get("sharedBottleneck") or "—", font=ITALIC)
else:
    ws["A1"] = "Measurement — NONE (baseline-only mode)"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "No saturation measurement present for this scenario. Sizing calculations are restricted to baseline."
    ws["A2"].font = ITALIC
    ws["A3"] = "Run /saturation-finder and import with benchmarks/scripts/import-saturation.sh to populate this tab."
    ws["A3"].font = ITALIC
    # Provide placeholder named cells so Scaling formulas don't crash — they'll show #N/A
    put(ws, 5, 1, "per_pod_steady", font=BOLD)
    put(ws, 5, 2, "=NA()", fill=GRAY)
    add_name("per_pod_steady", "Measurement!$B$5")
    put(ws, 6, 1, "mongo_cpu_per_msg", font=BOLD)
    put(ws, 6, 2, "=NA()", fill=GRAY)
    add_name("mongo_cpu_per_msg", "Measurement!$B$6")
    put(ws, 7, 1, "redis_cpu_per_msg", font=BOLD)
    put(ws, 7, 2, "=NA()", fill=GRAY)
    add_name("redis_cpu_per_msg", "Measurement!$B$7")
    put(ws, 8, 1, "mongo_iops_per_msg", font=BOLD)
    put(ws, 8, 2, scenario["perMessageDatastoreCost"]["mongoWrites"], fill=GRAY)
    add_name("mongo_iops_per_msg", "Measurement!$B$8")
    put(ws, 9, 1, "scaling_factor", font=BOLD)
    put(ws, 9, 2, 1.0, fill=GRAY)
    add_name("scaling_factor", "Measurement!$B$9")
    put(ws, 10, 1, "max_measured_pods", font=BOLD)
    put(ws, 10, 2, 1, fill=GRAY)
    add_name("max_measured_pods", "Measurement!$B$10")
    put(ws, 11, 1, "meas_captured_at", font=BOLD)
    put(ws, 11, 2, "2000-01-01", fill=GRAY)
    add_name("meas_captured_at", "Measurement!$B$11")

set_col_widths(ws, [42, 18, 14, 50])

# ----------------------------------------------------------------------------
# Tab: TestEvidence — saturation run per-step scorecard
# ----------------------------------------------------------------------------
# Resolve saturation report path:
#   1. --sat-report CLI override
#   2. measurement.saturationReportPath (relative to CWD)
#   3. None -> tab shows "no report" placeholder
sat_report_path = None
if args.sat_report:
    sat_report_path = args.sat_report
elif meas and meas.get("saturationReportPath"):
    sat_report_path = meas["saturationReportPath"]

ws = wb.create_sheet("TestEvidence")
ws["A1"] = "Test Evidence — saturation run scorecard"
ws["A1"].font = Font(bold=True, size=14)

if sat_report_path is None:
    ws["A2"] = "No saturation report path available (measurement absent or report path not set)."
    ws["A2"].font = ITALIC
    ws["A3"] = "Run /saturation-finder, import via benchmarks/scripts/import-saturation.sh, then regenerate this xlsx with --sat-report."
    ws["A3"].font = ITALIC
else:
    parsed = parse_saturation_report(sat_report_path)
    hdr = parsed["header"]
    steps = parsed["steps"]

    # Run header block
    ws["A2"] = f"Report: {sat_report_path}"
    ws["A2"].font = ITALIC
    info_row = 4
    write_headers(ws, info_row, ["Field", "Value"])
    info_rows = [
        ("Run ID", hdr.get("runId", "unknown")),
        ("Run date", hdr.get("runDate", "unknown")),
        ("Scenario", meas.get("scenarioId", "—") if meas else "—"),
        ("LLM mode", (meas or {}).get("configKey", {}).get("llmMode", "—")),
        ("Turn structure", (meas or {}).get("configKey", {}).get("turnStructure", "—")),
        ("CPU limit", (meas or {}).get("configKey", {}).get("cpuLimit", "—")),
        ("Memory limit", (meas or {}).get("configKey", {}).get("memoryLimit", "—")),
        ("Pods pinned", str((meas or {}).get("multiPodScaling", [{}])[-1].get("pods", "—"))),
        ("Max safe msg/s/pod", str((meas or {}).get("perPodCapacity", {}).get("maxSafeMsgPerSec", "—"))),
        ("p95 at max (ms)", str((meas or {}).get("perPodCapacity", {}).get("p95AtMaxMs", "—"))),
        ("Primary bottleneck", (meas or {}).get("perPodCapacity", {}).get("primaryBottleneck", "—")),
        ("Measurement captured at", (meas or {}).get("capturedAt", "—")),
    ]
    for i, (k, v) in enumerate(info_rows):
        r = info_row + 1 + i
        put(ws, r, 1, k, font=BOLD)
        put(ws, r, 2, v, font=ITALIC)

    # Per-step scorecard
    scorecard_row = info_row + 2 + len(info_rows)
    put(ws, scorecard_row, 1, "PER-STEP SCORECARD (from saturation report)", font=BOLD, fill=LIGHT_BLUE)

    if not steps:
        put(ws, scorecard_row + 1, 1,
            f"Could not parse scorecard table ({hdr.get('parseError','unknown')}).",
            font=ITALIC, fill=AMBER)
        put(ws, scorecard_row + 2, 1,
            "See the full report at the path above. Reports vary in table shape; the xlsx parses the first table that contains 'VUs' and 'p95' columns.",
            font=ITALIC)
    else:
        # Use the report's own column headers so we preserve its shape exactly
        cols = hdr.get("columns", list(steps[0].keys()))
        write_headers(ws, scorecard_row + 1, cols)
        for i, step in enumerate(steps):
            r = scorecard_row + 2 + i
            for j, col in enumerate(cols, start=1):
                val = step.get(col, "")
                cell = put(ws, r, j, val)
                low = val.lower() if isinstance(val, str) else ""
                if "🔴" in low or "red" in low or "stop" in low or "critical" in low:
                    cell.fill = RED
                elif "🟡" in low or "amber" in low or "warn" in low or "degrad" in low:
                    cell.fill = AMBER
                elif "🟢" in low or "green" in low or "proceed" in low or "healthy" in low or "optimal" in low:
                    cell.fill = LIGHT_GREEN

    # Saturation summary (component RAG) — derived from measurement if available
    rag_row = scorecard_row + 3 + max(1, len(steps))
    put(ws, rag_row, 1, "COMPONENT SATURATION SUMMARY", font=BOLD, fill=LIGHT_BLUE)
    write_headers(ws, rag_row + 1, ["Component", "Status", "Signal"])
    primary = (meas or {}).get("perPodCapacity", {}).get("primaryBottleneck", "")
    components = [
        ("Runtime CPU", "runtime" in primary.lower() or "cpu" in primary.lower()),
        ("Event loop / GC", "loop" in primary.lower() or "gc" in primary.lower()),
        ("MongoDB", "mongo" in primary.lower()),
        ("Redis", "redis" in primary.lower()),
        ("Storage IOPS", "iops" in primary.lower() or "disk" in primary.lower() or "storage" in primary.lower()),
    ]
    for i, (name, is_binding) in enumerate(components):
        r = rag_row + 2 + i
        put(ws, r, 1, name, font=BOLD)
        status = "RED (binding)" if is_binding else "GREEN"
        put(ws, r, 2, status, fill=(RED if is_binding else LIGHT_GREEN))
        put(ws, r, 3, primary if is_binding else "not the limiting layer at max", font=ITALIC)

    # Linkage footer
    link_row = rag_row + 3 + len(components)
    put(ws, link_row, 1, f"Full poll-level history in {sat_report_path}", font=ITALIC)
    put(ws, link_row + 1, 1,
        "Winning step's Msg/s and p95 became perPodCapacity.maxSafeMsgPerSec / p95AtMaxMs on the Measurement tab.",
        font=ITALIC)

set_col_widths(ws, [16, 14, 14, 12, 14, 12, 14, 12, 14, 14, 16, 22])

# ----------------------------------------------------------------------------
# Tab: Provenance — live age formulas
# ----------------------------------------------------------------------------
ws = wb.create_sheet("Provenance")
ws["A1"] = "Provenance — data source freshness"
ws["A1"].font = Font(bold=True, size=14)
ws["A2"] = "Age columns are live formulas; status recomputes daily."
ws["A2"].font = ITALIC

write_headers(ws, 4, ["Source", "Value", "Age (days)", "Status", "Notes"])

prov_rows = [
    ("infra-snapshot.json capturedAt", infra_captured,
        f'=DAYS(today_date, DATEVALUE("{infra_captured}"))',
        '=IF(C5>1, IF(C5>7, "STALE-RED", "STALE-YELLOW"), "FRESH")',
        "Re-run refresh-infra-snapshot.sh if stale"),
    ("infra canonical hash", infra_hash[:30] + "…", "", "", "SHA-256 of live kubectl state"),
    ("rate-card.json capturedAt", rate.get("capturedAt", "unknown")[:10],
        f'=DAYS(today_date, DATEVALUE("{rate.get("capturedAt", "2000-01-01")[:10]}"))',
        '=IF(C7>365, "EXPIRED", IF(C7>180, "STALE", "FRESH"))',
        "Re-run refresh-rate-card.sh if stale"),
    ("measurement capturedAt", meas_captured,
        f'=IF(meas_captured_at="2000-01-01", "N/A", DAYS(today_date, DATEVALUE(meas_captured_at)))',
        '=IF(NOT(ISNUMBER(C8)), "MISSING", IF(C8>90, "STALE-RED", IF(C8>30, "STALE-YELLOW", "FRESH")))',
        "Re-run /saturation-finder if stale"),
    ("measurement validationStatus", meas_val_status, "", "", "Flip to 'accepted' after review"),
    ("scenario file", args.scenario_file, "", "", ""),
    ("report generated", today, 0, "FRESH", ""),
]
for i, (label, value, age, status, note) in enumerate(prov_rows):
    r = 5 + i
    put(ws, r, 1, label, font=BOLD)
    put(ws, r, 2, str(value))
    if age != "":
        put(ws, r, 3, age, number_format="0")
    if status != "":
        put(ws, r, 4, status)
    put(ws, r, 5, note, font=ITALIC)

# Conditional formatting
status_range = f"D5:D{5 + len(prov_rows) - 1}"
ws.conditional_formatting.add(status_range, FormulaRule(formula=[f'ISNUMBER(SEARCH("RED", D5))'], fill=RED))
ws.conditional_formatting.add(status_range, FormulaRule(formula=[f'ISNUMBER(SEARCH("EXPIRED", D5))'], fill=RED))
ws.conditional_formatting.add(status_range, FormulaRule(formula=[f'ISNUMBER(SEARCH("MISSING", D5))'], fill=RED))
ws.conditional_formatting.add(status_range, FormulaRule(formula=[f'ISNUMBER(SEARCH("YELLOW", D5))'], fill=AMBER))
ws.conditional_formatting.add(status_range, FormulaRule(formula=[f'ISNUMBER(SEARCH("STALE", D5))'], fill=AMBER))
ws.conditional_formatting.add(status_range, FormulaRule(formula=[f'EXACT(D5,"FRESH")'], fill=GREEN))

set_col_widths(ws, [32, 44, 12, 18, 48])

# ----------------------------------------------------------------------------
# Tab: Baseline (F-7)
# ----------------------------------------------------------------------------
ws = wb.create_sheet("Baseline")
ws["A1"] = "Production-Minimum Baseline (F-7)"
ws["A1"].font = Font(bold=True, size=14)
ws["A2"] = "Always-on cost floor. Formulas reference Infra + RateCard cells."
ws["A2"].font = ITALIC

write_headers(ws, 4, ["Line", "Monthly USD", "Formula ID"])

# Baseline uses tier-selected HA shape (calculation-only). If baseline-tiers.json
# was not present, fall back to live infra names (prior behavior, suboptimal).
if mongo_tier is not None:
    mongo_compute_only = (
        "(baseline_mongo_replicas*baseline_mongo_cpu_limit*730*cpu_per_hr "
        "+ baseline_mongo_replicas*baseline_mongo_mem_limit*730*mem_per_hr)")
    mongo_pvc_expr = "baseline_mongo_replicas*baseline_mongo_pvc_gib"
    mongo_replicas_expr = "baseline_mongo_replicas"
else:
    mongo_compute_only = (
        "(mongo_replicas*mongo_cpu_limit*730*cpu_per_hr "
        "+ mongo_replicas*mongo_mem_limit*730*mem_per_hr)")
    mongo_pvc_expr = "mongo_replicas*mongo_pvc_gib"
    mongo_replicas_expr = "mongo_replicas"

if redis_tier is not None:
    redis_compute_only = ("("
        "(baseline_redis_master_replicas*baseline_redis_master_cpu_limit "
        "+ baseline_redis_replica_replicas*baseline_redis_replica_cpu_limit)"
        "*730*cpu_per_hr + "
        "(baseline_redis_master_replicas*baseline_redis_master_mem_limit "
        "+ baseline_redis_replica_replicas*baseline_redis_replica_mem_limit)"
        "*730*mem_per_hr)")
    redis_replicas_expr = "baseline_redis_total_replicas"
else:
    redis_compute_only = ("("
        "(redis_master_replicas*redis_master_cpu_limit + redis_replica_replicas*redis_replica_cpu_limit)"
        "*730*cpu_per_hr + "
        "(redis_master_replicas*redis_master_mem_limit + redis_replica_replicas*redis_replica_mem_limit)"
        "*730*mem_per_hr)")
    redis_replicas_expr = "redis_replicas"

put(ws, 5, 1, "Runtime baseline (HPA min pods)", font=BOLD)
# Baseline = HPA min replicas × per-pod cost (the always-on minimum)
if baseline_tiers_catalog and "runtime_tiers" in baseline_tiers_catalog:
    runtime_baseline_formula = ("=runtime_hpa_min*(baseline_runtime_cpu_limit*730*cpu_per_hr "
                                "+ baseline_runtime_mem_limit*730*mem_per_hr)")
else:
    runtime_baseline_formula = ("=runtime_hpa_min*(runtime_cpu_limit*730*cpu_per_hr "
                                "+ runtime_mem_limit*730*mem_per_hr)")
put(ws, 5, 2, runtime_baseline_formula,
    fill=GREEN, number_format='"$"#,##0.00')
put(ws, 5, 3, "F-7.3", font=ITALIC)
add_name("baseline_runtime", "Baseline!$B$5")

put(ws, 6, 1, "MongoDB baseline", font=BOLD)
put(ws, 6, 2, f"={mongo_compute_only}*mongo_uplift + {mongo_pvc_expr}*mongo_storage_per_gb + {mongo_replicas_expr}*mongo_replica_fee",
    fill=GREEN, number_format='"$"#,##0.00')
put(ws, 6, 3, "F-7.6", font=ITALIC)
add_name("baseline_mongo_full", "Baseline!$B$6")

put(ws, 7, 1, "  └─ Mongo compute-only (for F-8.2)", font=ITALIC)
put(ws, 7, 2, f"={mongo_compute_only}*mongo_uplift",
    fill=GRAY, number_format='"$"#,##0.00')
put(ws, 7, 3, "F-7.6.a", font=ITALIC)
add_name("baseline_mongo_compute", "Baseline!$B$7")

put(ws, 8, 1, "Redis baseline", font=BOLD)
put(ws, 8, 2, f"={redis_compute_only}*redis_uplift + {redis_replicas_expr}*redis_replica_fee",
    fill=GREEN, number_format='"$"#,##0.00')
put(ws, 8, 3, "F-7.9", font=ITALIC)
add_name("baseline_redis_full", "Baseline!$B$8")

put(ws, 9, 1, "  └─ Redis compute-only (for F-8.3)", font=ITALIC)
put(ws, 9, 2, f"={redis_compute_only}*redis_uplift",
    fill=GRAY, number_format='"$"#,##0.00')
put(ws, 9, 3, "F-7.9.a", font=ITALIC)
add_name("baseline_redis_compute", "Baseline!$B$9")

put(ws, 10, 1, "Ingress baseline", font=BOLD)
put(ws, 10, 2, "=ingress_fee", fill=GREEN, number_format='"$"#,##0.00')
put(ws, 10, 3, "F-7.10", font=ITALIC)

node_full = "(node_min_cores*730*cpu_per_hr + node_min_mem*730*mem_per_hr)"
mongo_selfhosted = f"IF(mongo_uplift=1, {mongo_compute_only}, 0)"
redis_selfhosted = f"IF(redis_uplift=1, {redis_compute_only}, 0)"
put(ws, 11, 1, "Node floor", font=BOLD)
put(ws, 11, 2, f"=MAX(0, {node_full} - (baseline_runtime + {mongo_selfhosted} + {redis_selfhosted}))",
    fill=GREEN, number_format='"$"#,##0.00')
put(ws, 11, 3, "F-7.11", font=ITALIC)

put(ws, 12, 1, "BASELINE TOTAL", font=BOLD)
put(ws, 12, 2, "=baseline_runtime + baseline_mongo_full + baseline_redis_full + B10 + B11",
    fill=GREEN, font=BOLD, number_format='"$"#,##0.00')
put(ws, 12, 3, "F-7.12", font=ITALIC)
add_name("baseline_total", "Baseline!$B$12")

set_col_widths(ws, [44, 16, 12])

# ----------------------------------------------------------------------------
# Tab: Scaling (F-1..F-8) using named cells — compact version
# ----------------------------------------------------------------------------
ws = wb.create_sheet("Scaling")
ws["A1"] = "Scaling Recommendation (F-1..F-8)"
ws["A1"].font = Font(bold=True, size=14)
ws["A2"] = "All formulas use named cells. Edit inputs on any tab → everything recomputes."
ws["A2"].font = ITALIC

write_headers(ws, 4, ["Metric", "Live Value", "Formula ID", "Unit"])

scaling_rows = [
    # F-1.1 per-pod effective, with runtime tier re-scaling heuristic:
    # if user picks a different runtime tier than the measurement, scale per-pod
    # capacity by (selected_cpu_limit / measurement_cpu_limit). Linear, LOW confidence.
    ("Per-pod effective (F-1.1)",
        (f"=per_pod_steady*(util_target/0.85)*(baseline_runtime_cpu_limit/{meas_runtime_cpu_lim})"
         if (baseline_tiers_catalog and "runtime_tiers" in baseline_tiers_catalog and meas is not None)
         else "=per_pod_steady*(util_target/0.85)"),
        "per_pod_eff", "0.00", "F-1.1", "msg/s/pod"),
    ("Runtime tier CPU scaling factor",
        (f"=baseline_runtime_cpu_limit/{meas_runtime_cpu_lim}"
         if (baseline_tiers_catalog and "runtime_tiers" in baseline_tiers_catalog and meas is not None)
         else "=1"),
        "runtime_cpu_scale", "0.00", "F-1.1.a",
        "ratio; =1 means same-as-measurement. Re-scaling is LOW confidence."),
    # F-1.0: WITHIN-MEASURED SHORTCUT — if target <= measured fleet ceiling,
    # use measured pod count (no derating). Else fall through to F-1.3 extrapolation.
    # Ceiling scales with runtime_cpu_scale if user picked a different runtime tier.
    ("Measured fleet ceiling (F-1.0)",
        "=per_pod_steady*max_measured_pods*runtime_cpu_scale",
        "measured_fleet_ceiling", "0.00", "F-1.0", "msg/s"),
    ("Pods required (F-1.0 or F-1.3)",
        "=IF(target_msgps<=measured_fleet_ceiling, max_measured_pods, CEILING(target_msgps/per_pod_eff, 1))",
        "pods_required", "0", "F-1.0/F-1.3", "pods"),
    ("Pods required — extrapolation? (F-1.0)",
        '=IF(target_msgps<=measured_fleet_ceiling, "MEASURED-DIRECT (HIGH)", "EXTRAPOLATION (see §9)")',
        None, None, "F-1.0", ""),
    ("Effective total msg/s (F-1.4)",
        "=IF(target_msgps<=measured_fleet_ceiling, target_msgps, pods_required*per_pod_eff)",
        "eff_total", "0.00", "F-1.4", "msg/s"),
    ("Runtime total cost (F-2.5)",
        ("=pods_required*baseline_runtime_cpu_limit*730*cpu_per_hr "
         "+ pods_required*baseline_runtime_mem_limit*730*mem_per_hr")
        if baseline_tiers_catalog and "runtime_tiers" in baseline_tiers_catalog
        else "=pods_required*runtime_cpu_limit*730*cpu_per_hr + pods_required*runtime_mem_limit*730*mem_per_hr",
        "runtime_total_cost", '"$"#,##0.00', "F-2.5", "$/mo"),
    ("Mongo CPU milli required (F-3.2)", "=target_msgps*mongo_cpu_per_msg/0.70", "mongo_cpu_milli", "0.00", "F-3.2", "milli"),
    ("Mongo cores required", "=mongo_cpu_milli/1000", "mongo_cores_req", "0.00", "F-3.2", "cores"),
    ("Mongo IOPS required (F-3.4)", "=target_msgps*mongo_iops_per_msg/0.70", "mongo_iops_req", "0.0", "F-3.4", "IOPS"),
    # F-3.5 Storage: writes/msg × msg/s × sec/day × retention × doc_size_KB → GB
    # Industry standard: size for uncompressed data (WiredTiger gives 2-4× compression
    # as bonus headroom, but never rely on it for provisioning).
    # Includes: ×replicas (each replica stores full copy) × 1.3 (index overhead 20-40%).
    ("Mongo raw storage GB (F-3.5)",
        "=mongo_writes_per_msg*target_msgps*86400*retention_days*mongo_write_doc_kb/(1024*1024)",
        "mongo_raw_storage_gb", "0.00", "F-3.5", "GB (uncompressed, single-copy)"),
    ("Mongo storage GB (replicated + indexes)",
        ("=mongo_raw_storage_gb*baseline_mongo_replicas*1.3"
         if baseline_tiers_catalog is not None
         else "=mongo_raw_storage_gb*mongo_replicas*1.3"),
        "mongo_storage_gb", "0.00", "F-3.5.a",
        "raw × replicas × 1.3 index overhead"),
    ("Mongo memory required GB (F-3.7)",
        ("=mongo_cores_req*(baseline_mongo_mem_limit/baseline_mongo_cpu_limit)"
         if baseline_tiers_catalog is not None
         else "=mongo_cores_req*(mongo_mem_limit/mongo_cpu_limit)"),
        "mongo_mem_req", "0.00", "F-3.7", "GB"),
    ("Mongo compute cost (F-3.8)", "=mongo_cores_req*730*cpu_per_hr*mongo_uplift + mongo_mem_req*730*mem_per_hr*mongo_uplift",
        "mongo_compute_cost", '"$"#,##0.00', "F-3.8", "$/mo"),
    # F-3.9 storage cost: use MAX(tier PVC × replicas, workload storage gb) × $/GB.
    # Rationale: tier PVC is the minimum you'd provision (baseline already paid for it);
    # workload may need more (beyond-tier storage). Real cost is the larger of the two.
    # This eliminates the "baseline pays for 150Gi tier + scaling adds 2,965Gi extra" double-count.
    ("Mongo effective storage GB",
        (f"=MAX(baseline_mongo_replicas*baseline_mongo_pvc_gib, mongo_storage_gb)"
         if baseline_tiers_catalog is not None
         else "=mongo_storage_gb"),
        "mongo_effective_storage_gb", "0.00", "F-3.5.b",
        "max(tier PVC fleet, workload-required GB)"),
    ("Mongo total cost (F-3.9)",
        "=mongo_compute_cost + mongo_effective_storage_gb*mongo_storage_per_gb + mongo_iops_req*mongo_iops_per_mo",
        "mongo_total_cost", '"$"#,##0.00', "F-3.9", "$/mo"),
    ("Redis CPU milli required (F-4.2)", "=target_msgps*redis_cpu_per_msg/0.50", "redis_cpu_milli", "0.00", "F-4.2", "milli"),
    ("Redis cores required", "=redis_cpu_milli/1000", "redis_cores_req", "0.00", "F-4.2", "cores"),
    # F-4.3 Little's Law: L = λ × W
    # λ = session arrival rate = msg_rate / msgs_per_session
    # W = session lifetime = session_ttl_min × 60
    # L = concurrent sessions = (target_msgps / msgs_per_session) × session_ttl_min × 60
    ("Session arrival rate (λ)", "=target_msgps/msgs_per_session", "session_arrival_rate", "0.00", "F-4.3.a", "sessions/s"),
    ("Concurrent active sessions (F-4.3)", "=session_arrival_rate*session_ttl_min*60", "conc_sessions", "0", "F-4.3", "sessions"),
    ("Fleet session cache budget (F-4.4)", "=session_cap_per_pod*pods_required", "fleet_cache_budget", "0", "F-4.4", "sessions"),
    ("Redis memory GB (F-4.4.b)", "=MIN(conc_sessions, fleet_cache_budget)*session_kb/(1024*1024)",
        "redis_mem_gb", "0.00", "F-4.4.b", "GB"),
    ("Redis compute cost (F-4.7)", "=redis_cores_req*730*cpu_per_hr*redis_uplift + redis_mem_gb*730*mem_per_hr*redis_uplift",
        "redis_compute_cost", '"$"#,##0.00', "F-4.7", "$/mo"),
    ("Redis total cost (F-4.8)", "=redis_compute_cost", "redis_total_cost", '"$"#,##0.00', "F-4.8", "$/mo"),
    ("LLM cost (F-5.3)",
        "=(target_msgps*86400*30*llm_tokens_in/1000000)*llm_input_per_m + (target_msgps*86400*30*llm_tokens_out/1000000)*llm_output_per_m",
        "llm_cost", '"$"#,##0.00', "F-5.3", "$/mo"),
    ("Egress GB/mo (F-6.1)", "=target_msgps*86400*30*payload_bytes_out/(1024*1024*1024)", "egress_gb", "0.00", "F-6.1", "GB"),
    ("Egress cost (F-6.2)", "=egress_gb*egress_per_gb", "egress_cost", '"$"#,##0.00', "F-6.2", "$/mo"),
    ("Δ runtime (F-8.1)", "=runtime_total_cost - baseline_runtime", "delta_runtime", '"$"#,##0.00', "F-8.1", "$/mo"),
    ("Δ mongo (F-8.2)", "=mongo_total_cost - baseline_mongo_full", "delta_mongo", '"$"#,##0.00', "F-8.2", "$/mo"),
    ("Δ redis (F-8.3)", "=redis_total_cost - baseline_redis_full", "delta_redis", '"$"#,##0.00', "F-8.3", "$/mo"),
    ("Scaling cost (F-8.4)", "=MAX(0,delta_runtime)+MAX(0,delta_mongo)+MAX(0,delta_redis)+llm_cost+egress_cost",
        "scaling_cost", '"$"#,##0.00', "F-8.4", "$/mo"),
    ("TOTAL monthly cost (F-8.5)", "=baseline_total + scaling_cost", "total_cost", '"$"#,##0.00', "F-8.5", "$/mo"),
    ("Cost per million messages (F-8.6)", "=total_cost / (target_msgps*86400*30/1000000)",
        "cost_per_m", '"$"#,##0.00', "F-8.6", "$/M msg"),
    # ─── CAPACITY PLANNING: Peak Provisioning + Availability ────────────────
    ("", None, None, None, None, None),  # spacer
    ("── PRODUCTION SIZING (burst + HA) ──", None, None, None, None, None),
    # Peak pods: capacity to handle burst_factor × avg without SLA breach
    ("Peak msg/s (burst)", "=target_msgps*burst_factor",
        "peak_msgps", "0.00", "F-9.1", "msg/s"),
    ("Pods for peak (F-9.2)",
        "=CEILING(peak_msgps/per_pod_eff, 1)",
        "pods_for_peak", "0", "F-9.2", "pods"),
    # N+K availability: survive K pod failures at average load
    ("Pods with N+K headroom (F-9.3)",
        "=pods_required + availability_headroom_pods",
        "pods_with_headroom", "0", "F-9.3", "pods"),
    # Production recommendation: MAX(peak pods, headroom pods, HPA min)
    ("PRODUCTION PODS (recommended)",
        "=MAX(pods_for_peak, pods_with_headroom, runtime_hpa_min)",
        "pods_production", "0", "F-9.4", "pods"),
    ("Production runtime cost",
        ("=pods_production*baseline_runtime_cpu_limit*730*cpu_per_hr "
         "+ pods_production*baseline_runtime_mem_limit*730*mem_per_hr")
        if baseline_tiers_catalog and "runtime_tiers" in baseline_tiers_catalog
        else "=pods_production*runtime_cpu_limit*730*cpu_per_hr + pods_production*runtime_mem_limit*730*mem_per_hr",
        "prod_runtime_cost", '"$"#,##0.00', "F-9.5", "$/mo"),
    # ─── ACTIVE HOURS: volume correction ────────────────────────────────────
    ("", None, None, None, None, None),  # spacer
    ("── ACTIVE HOURS VOLUME ──", None, None, None, None, None),
    # Monthly active seconds: active_hours × active_days × 4.33 weeks × 3600
    ("Monthly active seconds",
        "=active_hours_per_day*active_days_per_week*4.33*3600",
        "monthly_active_sec", "0", "F-10.1", "seconds"),
    # Full utilization would be 86400*30 = 2,592,000 seconds
    ("Duty cycle ratio",
        "=monthly_active_sec/(86400*30)",
        "duty_cycle", "0.00%", "F-10.2", "ratio (vs 24/7)"),
    ("Effective monthly messages",
        "=target_msgps*monthly_active_sec",
        "eff_monthly_msgs", "0", "F-10.3", "messages"),
    # LLM cost corrected for active hours only
    ("LLM cost (active hours)",
        "=(eff_monthly_msgs*llm_tokens_in/1000000)*llm_input_per_m + (eff_monthly_msgs*llm_tokens_out/1000000)*llm_output_per_m",
        "llm_cost_active", '"$"#,##0.00', "F-10.4", "$/mo"),
    # Storage corrected for active hours
    ("Mongo storage (active hours, replicated)",
        ("=mongo_writes_per_msg*target_msgps*monthly_active_sec*retention_days/30*mongo_write_doc_kb/(1024*1024)*baseline_mongo_replicas*1.3"
         if baseline_tiers_catalog is not None
         else "=mongo_writes_per_msg*target_msgps*monthly_active_sec*retention_days/30*mongo_write_doc_kb/(1024*1024)*mongo_replicas*1.3"),
        "mongo_storage_active", "0.00", "F-10.5", "GB"),
    # ─── UNIT ECONOMICS ─────────────────────────────────────────────────────
    ("", None, None, None, None, None),  # spacer
    ("── UNIT ECONOMICS ──", None, None, None, None, None),
    ("Conversations per month",
        "=eff_monthly_msgs/msgs_per_conversation",
        "conversations_per_mo", "0", "F-11.1", "conversations"),
    ("Cost per conversation",
        "=IF(conversations_per_mo>0, total_cost/conversations_per_mo, 0)",
        "cost_per_conversation", '"$"#,##0.0000', "F-11.2", "$/conv"),
    ("Cost per active-user-hour (at 3 conv/hr)",
        "=cost_per_conversation*3",
        "cost_per_user_hr", '"$"#,##0.0000', "F-11.3", "$/user-hr"),
    # ─── PRODUCTION TOTAL (full picture) ────────────────────────────────────
    ("", None, None, None, None, None),  # spacer
    ("── PRODUCTION TOTAL ──", None, None, None, None, None),
    ("PROD monthly (peak-sized + active-hours LLM)",
        "=baseline_total + MAX(0, prod_runtime_cost - baseline_runtime) + MAX(0,delta_mongo) + MAX(0,delta_redis) + llm_cost_active + egress_cost",
        "prod_total_cost", '"$"#,##0.00', "F-12.1", "$/mo"),
    ("PROD cost per M messages",
        "=IF(eff_monthly_msgs>0, prod_total_cost/(eff_monthly_msgs/1000000), 0)",
        "prod_cost_per_m", '"$"#,##0.00', "F-12.2", "$/M msg"),
]
for i, (label, formula, nm, nfmt, fid, unit) in enumerate(scaling_rows):
    r = 5 + i
    if formula is None:
        # Section header or spacer row
        if label:
            put(ws, r, 1, label, font=Font(bold=True, size=11), fill=LIGHT_BLUE)
        continue
    put(ws, r, 1, label, font=BOLD)
    put(ws, r, 2, formula, fill=GREEN, number_format=nfmt)
    if fid is not None:
        put(ws, r, 3, fid, font=ITALIC)
    if unit is not None:
        put(ws, r, 4, unit, font=ITALIC)
    if nm is not None:
        add_name(nm, f"Scaling!$B${r}")

total_row = 5 + len(scaling_rows) - 2
for col in range(1, 5):
    ws.cell(row=total_row, column=col).font = BOLD

set_col_widths(ws, [40, 16, 10, 14])

# ----------------------------------------------------------------------------
# Tab: Bottleneck
# ----------------------------------------------------------------------------
ws = wb.create_sheet("Bottleneck")
ws["A1"] = "Bottleneck Forecast — per-layer utilization"
ws["A1"].font = Font(bold=True, size=14)
ws["A2"] = "RED ≥ 100%, AMBER 80-99%, GREEN < 80%"
ws["A2"].font = ITALIC

write_headers(ws, 4, ["Layer", "Required", "Available", "Util @ target", "Util @ 2× target", "Evidence"])

# Bottleneck compares workload requirements vs SIZED (tier-selected) availability.
# When baseline_tiers_catalog is loaded, use baseline_* names (= what you'd provision
# for the target); otherwise fall back to live infra names (mongo_* etc).
if baseline_tiers_catalog is not None:
    _mongo_cpu_avail = "baseline_mongo_replicas*baseline_mongo_cpu_limit"
    _mongo_iops_avail = "baseline_mongo_iops_limit"
    _mongo_pvc_avail = "baseline_mongo_replicas*baseline_mongo_pvc_gib"
    _redis_cpu_avail = ("baseline_redis_master_replicas*baseline_redis_master_cpu_limit "
                        "+ baseline_redis_replica_replicas*baseline_redis_replica_cpu_limit")
    _redis_mem_avail = ("baseline_redis_master_replicas*baseline_redis_master_mem_limit "
                        "+ baseline_redis_replica_replicas*baseline_redis_replica_mem_limit")
else:
    _mongo_cpu_avail = "mongo_replicas*mongo_cpu_limit"
    _mongo_iops_avail = "mongo_iops_limit"
    _mongo_pvc_avail = "mongo_replicas*mongo_pvc_gib"
    _redis_cpu_avail = ("redis_master_replicas*redis_master_cpu_limit "
                        "+ redis_replica_replicas*redis_replica_cpu_limit")
    _redis_mem_avail = ("redis_master_replicas*redis_master_mem_limit "
                        "+ redis_replica_replicas*redis_replica_mem_limit")

btl_rows = [
    ("Pods vs measured range", "=pods_required", "=max_measured_pods", "=B5/C5",
        "=CEILING(target_msgps*2/per_pod_eff, 1) / C5",
        "pods_required vs maxMeasuredPods"),
    ("Mongo CPU", "=mongo_cores_req", f"={_mongo_cpu_avail}", "=B6/C6", "=B6*2/C6", "F-3.2"),
    ("Mongo IOPS", "=mongo_iops_req", f"={_mongo_iops_avail}", "=B7/C7", "=B7*2/C7", "F-3.4"),
    ("Mongo storage", "=mongo_storage_gb", f"={_mongo_pvc_avail}", "=B8/C8", "=B8*2/C8", "F-3.5"),
    ("Redis CPU", "=redis_cores_req", f"={_redis_cpu_avail}", "=B9/C9", "=B9*2/C9", "F-4.2"),
    ("Redis memory", "=redis_mem_gb", f"={_redis_mem_avail}", "=B10/C10", "=B10*2/C10", "F-4.4.b"),
]
for i, (label, req, avail, util, util2, note) in enumerate(btl_rows):
    r = 5 + i
    put(ws, r, 1, label, font=BOLD)
    put(ws, r, 2, req, number_format="0.00")
    put(ws, r, 3, avail, number_format="0.00")
    put(ws, r, 4, util, number_format="0.0%")
    put(ws, r, 5, util2, number_format="0.0%")
    put(ws, r, 6, note, font=ITALIC)

for col_letter in ("D", "E"):
    rng = f"{col_letter}5:{col_letter}10"
    ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThanOrEqual", formula=["1"], fill=RED))
    ws.conditional_formatting.add(rng, CellIsRule(operator="between", formula=["0.8", "0.999"], fill=AMBER))
    ws.conditional_formatting.add(rng, CellIsRule(operator="lessThan", formula=["0.8"], fill=GREEN))

put(ws, 12, 1, "Binding layer at target", font=BOLD, fill=LIGHT_BLUE)
put(ws, 12, 2, '=INDEX(A5:A10, MATCH(MAX(D5:D10), D5:D10, 0))', fill=GREEN, font=BOLD)
put(ws, 12, 4, "=MAX(D5:D10)", fill=GREEN, number_format="0.0%", font=BOLD)
add_name("binding_layer", "Bottleneck!$B$12")
add_name("binding_util_pct", "Bottleneck!$D$12")

put(ws, 13, 1, "Binding layer at 2× target", font=BOLD, fill=LIGHT_BLUE)
put(ws, 13, 2, '=INDEX(A5:A10, MATCH(MAX(E5:E10), E5:E10, 0))', fill=AMBER, font=BOLD)
put(ws, 13, 4, "=MAX(E5:E10)", fill=AMBER, number_format="0.0%", font=BOLD)

set_col_widths(ws, [28, 14, 16, 16, 18, 42])

# ----------------------------------------------------------------------------
# Tab: WhatIf
# ----------------------------------------------------------------------------
ws = wb.create_sheet("WhatIf")
ws["A1"] = "What-If — target msg/s sweep"; ws["A1"].font = Font(bold=True, size=14)
ws["A2"] = "Each column reruns F-1..F-8 at target_msgps × multiplier."
ws["A2"].font = ITALIC

write_headers(ws, 4, ["Multiplier", "0.25×", "0.5×", "1×", "2×", "5×"])

put(ws, 5, 1, "Target msg/s", font=BOLD)
for col, mult in enumerate([0.25, 0.5, 1.0, 2.0, 5.0], start=2):
    put(ws, 5, col, f"=target_msgps*{mult}", fill=GRAY, number_format="0.##")

put(ws, 6, 1, "Pods required", font=BOLD)
for col in range(2, 7):
    put(ws, 6, col, f"=CEILING({get_column_letter(col)}5/per_pod_eff, 1)", fill=GREEN, number_format="0")

put(ws, 7, 1, "Runtime cost", font=BOLD)
_runtime_cpu_name = ("baseline_runtime_cpu_limit"
                     if baseline_tiers_catalog and "runtime_tiers" in baseline_tiers_catalog
                     else "runtime_cpu_limit")
_runtime_mem_name = ("baseline_runtime_mem_limit"
                     if baseline_tiers_catalog and "runtime_tiers" in baseline_tiers_catalog
                     else "runtime_mem_limit")
for col in range(2, 7):
    pods = f"{get_column_letter(col)}6"
    put(ws, 7, col,
        f"={pods}*{_runtime_cpu_name}*730*cpu_per_hr + {pods}*{_runtime_mem_name}*730*mem_per_hr",
        fill=GREEN, number_format='"$"#,##0.00')

put(ws, 8, 1, "Mongo total", font=BOLD)
_mm_ratio = ("(baseline_mongo_mem_limit/baseline_mongo_cpu_limit)"
             if baseline_tiers_catalog is not None
             else "(mongo_mem_limit/mongo_cpu_limit)")
# Storage uses MAX(tier PVC fleet, workload raw GB × replicas × 1.3 index overhead)
_mongo_pvc_floor = ("baseline_mongo_replicas*baseline_mongo_pvc_gib"
                    if baseline_tiers_catalog is not None
                    else "mongo_replicas*mongo_pvc_gib")
_mongo_replicas_for_storage = ("baseline_mongo_replicas"
                               if baseline_tiers_catalog is not None
                               else "mongo_replicas")
for col in range(2, 7):
    t = f"{get_column_letter(col)}5"
    put(ws, 8, col,
        f"=({t}*mongo_cpu_per_msg/0.70/1000)*730*cpu_per_hr*mongo_uplift"
        f"+({t}*mongo_cpu_per_msg/0.70/1000)*{_mm_ratio}*730*mem_per_hr*mongo_uplift"
        f"+MAX({_mongo_pvc_floor}, mongo_writes_per_msg*{t}*86400*retention_days*mongo_write_doc_kb/(1024*1024)*{_mongo_replicas_for_storage}*1.3)*mongo_storage_per_gb"
        f"+({t}*mongo_iops_per_msg/0.70)*mongo_iops_per_mo",
        fill=GREEN, number_format='"$"#,##0.00')

put(ws, 9, 1, "Redis total", font=BOLD)
# Fix: use Little's Law — session_rate = msg_rate / msgs_per_session
for col in range(2, 7):
    t = f"{get_column_letter(col)}5"; pods = f"{get_column_letter(col)}6"
    put(ws, 9, col,
        f"=({t}*redis_cpu_per_msg/0.50/1000)*730*cpu_per_hr*redis_uplift"
        f"+MIN({t}/msgs_per_session*session_ttl_min*60, session_cap_per_pod*{pods})*session_kb/(1024*1024)*730*mem_per_hr*redis_uplift",
        fill=GREEN, number_format='"$"#,##0.00')

put(ws, 10, 1, "LLM cost", font=BOLD)
for col in range(2, 7):
    t = f"{get_column_letter(col)}5"
    put(ws, 10, col,
        f"=({t}*86400*30*llm_tokens_in/1000000)*llm_input_per_m + ({t}*86400*30*llm_tokens_out/1000000)*llm_output_per_m",
        fill=GREEN, number_format='"$"#,##0.00')

put(ws, 11, 1, "Egress cost", font=BOLD)
for col in range(2, 7):
    t = f"{get_column_letter(col)}5"
    put(ws, 11, col, f"=({t}*86400*30*payload_bytes_out/(1024*1024*1024))*egress_per_gb",
        fill=GREEN, number_format='"$"#,##0.00')

put(ws, 12, 1, "TOTAL monthly", font=BOLD)
for col in range(2, 7):
    cl = get_column_letter(col)
    put(ws, 12, col,
        f"=baseline_total + MAX(0, {cl}7-baseline_runtime) + MAX(0, {cl}8-baseline_mongo_full) + MAX(0, {cl}9-baseline_redis_full) + {cl}10 + {cl}11",
        fill=GREEN, number_format='"$"#,##0.00', font=BOLD)

put(ws, 13, 1, "Cost / M messages", font=BOLD)
for col in range(2, 7):
    cl = get_column_letter(col)
    put(ws, 13, col, f"={cl}12/({cl}5*86400*30/1000000)", fill=GREEN, number_format='"$"#,##0.00')

put(ws, 14, 1, "Within measured range?", font=BOLD)
for col in range(2, 7):
    cl = get_column_letter(col)
    put(ws, 14, col, f'=IF({cl}6 <= max_measured_pods, "✓", "⚠ exceeds measured")')

set_col_widths(ws, [28, 14, 14, 14, 14, 14, 14])

# ----------------------------------------------------------------------------
# Tab: Projections — Growth forecasting at 6/12/18 month horizons
# ----------------------------------------------------------------------------
ws = wb.create_sheet("Projections")
ws["A1"] = "Growth Projections — Capacity & Cost Forecast"
ws["A1"].font = Font(bold=True, size=14)
ws["A2"] = "Compound monthly growth applied to target msg/s. Infrastructure scales accordingly."
ws["A2"].font = ITALIC
ws["A3"] = "Edit growth_rate_monthly on Controls tab to adjust. Current default: 10%/mo = ~3.1×/year."
ws["A3"].font = ITALIC

write_headers(ws, 5, ["Horizon", "Today", "+3 months", "+6 months", "+12 months", "+18 months", "+24 months"])

horizons = [0, 3, 6, 12, 18, 24]
horizon_cols = list(range(2, 8))

# Row 6: Target msg/s at each horizon
put(ws, 6, 1, "Target msg/s", font=BOLD)
for col, m in zip(horizon_cols, horizons):
    put(ws, 6, col, f"=target_msgps*(1+growth_rate_monthly)^{m}",
        fill=GREEN, number_format="0.0")

# Row 7: Peak msg/s (with burst)
put(ws, 7, 1, "Peak msg/s (×burst)", font=BOLD)
for col, m in zip(horizon_cols, horizons):
    put(ws, 7, col, f"={get_column_letter(col)}6*burst_factor",
        fill=GRAY, number_format="0.0")

# Row 8: Pods for avg load
put(ws, 8, 1, "Pods (avg load)", font=BOLD)
for col in horizon_cols:
    put(ws, 8, col, f"=CEILING({get_column_letter(col)}6/per_pod_eff, 1)",
        fill=GREEN, number_format="0")

# Row 9: Pods for peak + headroom (production)
put(ws, 9, 1, "PRODUCTION PODS (peak+N+K)", font=BOLD)
for col in horizon_cols:
    cl = get_column_letter(col)
    put(ws, 9, col,
        f"=MAX(CEILING({cl}7/per_pod_eff, 1) + availability_headroom_pods, runtime_hpa_min)",
        fill=GREEN, font=BOLD, number_format="0")

# Row 10: Within measured range?
put(ws, 10, 1, "Within measured?", font=BOLD)
for col in horizon_cols:
    cl = get_column_letter(col)
    put(ws, 10, col,
        f'=IF({cl}8<=max_measured_pods, "✓ measured", "⚠ extrapolated")',
        font=ITALIC)

# Row 11: Monthly messages (active hours)
put(ws, 11, 1, "Monthly messages", font=BOLD)
for col in horizon_cols:
    cl = get_column_letter(col)
    put(ws, 11, col, f"={cl}6*monthly_active_sec",
        fill=GRAY, number_format="#,##0")

# Row 12: Runtime cost
put(ws, 12, 1, "Runtime cost", font=BOLD)
_rc = "baseline_runtime_cpu_limit" if baseline_tiers_catalog and "runtime_tiers" in baseline_tiers_catalog else "runtime_cpu_limit"
_rm = "baseline_runtime_mem_limit" if baseline_tiers_catalog and "runtime_tiers" in baseline_tiers_catalog else "runtime_mem_limit"
for col in horizon_cols:
    cl = get_column_letter(col)
    put(ws, 12, col,
        f"={cl}9*{_rc}*730*cpu_per_hr + {cl}9*{_rm}*730*mem_per_hr",
        fill=GREEN, number_format='"$"#,##0')

# Row 13: LLM cost (active hours)
put(ws, 13, 1, "LLM cost (active hrs)", font=BOLD)
for col in horizon_cols:
    cl = get_column_letter(col)
    put(ws, 13, col,
        f"=({cl}11*llm_tokens_in/1000000)*llm_input_per_m + ({cl}11*llm_tokens_out/1000000)*llm_output_per_m",
        fill=GREEN, number_format='"$"#,##0')

# Row 14: Mongo storage at horizon (cumulative from today)
put(ws, 14, 1, "Cumulative Mongo storage (GB)", font=BOLD)
_mr = "baseline_mongo_replicas" if baseline_tiers_catalog is not None else "mongo_replicas"
for col, m in zip(horizon_cols, horizons):
    cl = get_column_letter(col)
    # Integral of compound growth over m months: sum of geometric series × daily_gb
    # Approximation: avg_rate × days = target × ((1+g)^m + 1)/2 × 30*m × sec/day
    put(ws, 14, col,
        f"=mongo_writes_per_msg*target_msgps*((1+growth_rate_monthly)^{m}+1)/2*86400*30*MAX(1,{m})*mongo_write_doc_kb/(1024*1024)*{_mr}*1.3",
        fill=GRAY, number_format="#,##0")

# Row 15: Total estimated monthly cost
put(ws, 15, 1, "TOTAL monthly cost", font=BOLD)
for col in horizon_cols:
    cl = get_column_letter(col)
    put(ws, 15, col,
        f"=baseline_total + MAX(0, {cl}12 - baseline_runtime) + MAX(0, delta_mongo) + MAX(0, delta_redis) + {cl}13 + egress_cost",
        fill=GREEN, font=BOLD, number_format='"$"#,##0')

# Row 16: Cost per conversation
put(ws, 16, 1, "Cost per conversation", font=BOLD)
for col in horizon_cols:
    cl = get_column_letter(col)
    put(ws, 16, col,
        f"=IF({cl}11>0, {cl}15/({cl}11/msgs_per_conversation), 0)",
        fill=GREEN, number_format='"$"#,##0.0000')

# Row 17: Capacity confidence
put(ws, 17, 1, "Confidence level", font=BOLD)
for col in horizon_cols:
    cl = get_column_letter(col)
    put(ws, 17, col,
        f'=IF({cl}8<=max_measured_pods, "HIGH (measured)", IF({cl}8<=max_measured_pods*2, "MEDIUM (1 step extrapolation)", "LOW (far extrapolation)"))',
        font=ITALIC)

# Conditional formatting on confidence
for col in horizon_cols:
    cl = get_column_letter(col)
    ws.conditional_formatting.add(
        f"{cl}17", FormulaRule(formula=[f'ISNUMBER(SEARCH("HIGH", {cl}17))'], fill=GREEN))
    ws.conditional_formatting.add(
        f"{cl}17", FormulaRule(formula=[f'ISNUMBER(SEARCH("MEDIUM", {cl}17))'], fill=AMBER))
    ws.conditional_formatting.add(
        f"{cl}17", FormulaRule(formula=[f'ISNUMBER(SEARCH("LOW", {cl}17))'], fill=RED))

set_col_widths(ws, [30, 14, 14, 14, 14, 14, 14])

# ----------------------------------------------------------------------------
# Tab: Warnings
# ----------------------------------------------------------------------------
ws = wb.create_sheet("Warnings")
ws["A1"] = "Warnings — live flags"; ws["A1"].font = Font(bold=True, size=14)
ws["A2"] = "All status cells are LIVE formulas; recompute on any edit."
ws["A2"].font = ITALIC

write_headers(ws, 4, ["Check", "Status", "Value", "Threshold", "Remediation"])

_tier_pvc_expr = ("baseline_mongo_replicas*baseline_mongo_pvc_gib"
                  if baseline_tiers_catalog is not None
                  else "mongo_replicas*mongo_pvc_gib")
_tier_iops_expr = ("baseline_mongo_iops_limit"
                   if baseline_tiers_catalog is not None
                   else "mongo_iops_limit")

warn_rows = [
    ("Mongo storage fits selected tier PVC",
        f'=IF(mongo_storage_gb > {_tier_pvc_expr}, "RED: " & ROUND(mongo_storage_gb/({_tier_pvc_expr}),1) & "x over tier PVC", "OK")',
        '=mongo_storage_gb & " GB"', f'={_tier_pvc_expr} & " GB"', "Pick bigger Mongo tier on Controls!B12 OR reduce retentionDays"),
    # Tier catalog max — can ANY tier cover the workload? If not, retention is too high.
    ("Mongo storage fits MAX tier PVC (catalog limit)",
        # xlarge = 500 Gi per pod × 3 replicas = 1500 Gi fleet (hardcoded from catalog)
        '=IF(mongo_storage_gb > 1500, "RED: exceeds largest tier (1500 GB). Reduce retentionDays OR use custom oversized PVC.", "OK (fits xlarge)")',
        '=mongo_storage_gb & " GB"', '"1500 GB (xlarge)"',
        "Reduce retentionDays; current 90d is unusual — chat platforms typically use 7-30d"),
    ("Mongo IOPS within disk limit",
        f'=IF(mongo_iops_req > {_tier_iops_expr}, "RED: " & ROUND(mongo_iops_req/{_tier_iops_expr},1) & "x over", IF(mongo_iops_req > {_tier_iops_expr}*0.8, "AMBER", "OK"))',
        '=mongo_iops_req & " IOPS"', f'={_tier_iops_expr} & " IOPS"', "Pick bigger Mongo tier (higher IOPS)"),
    ("Pods within measured range",
        '=IF(pods_required > max_measured_pods, "AMBER: " & pods_required & " > max measured " & max_measured_pods, "OK")',
        '=pods_required', '=max_measured_pods', "Run /saturation-finder at higher pod count"),
    ("Capacity ≥ target",
        # F-1.0 path: when within measured range, eff_total == target by construction → OK
        # F-1.3 path: check pods_required * per_pod_eff * scaling_factor >= target
        '=IF(target_msgps<=measured_fleet_ceiling, "OK (measured-direct)", IF(pods_required*per_pod_eff*scaling_factor >= target_msgps, "OK", "RED: capacity < target"))',
        '=pods_required*per_pod_eff*scaling_factor', '=target_msgps', "Increase pods"),
    ("Infra snapshot age ≤ 1 day",
        f'=IF(DAYS(today_date, DATEVALUE("{infra_captured}"))>7, "STALE-RED: >7d", IF(DAYS(today_date, DATEVALUE("{infra_captured}"))>1, "STALE-YELLOW: >1d", "OK"))',
        f'=DAYS(today_date, DATEVALUE("{infra_captured}")) & " days"', "1 day",
        "Run ./benchmarks/scripts/refresh-infra-snapshot.sh"),
    ("Measurement age ≤ 30 days",
        '=IF(NOT(ISNUMBER(DAYS(today_date, DATEVALUE(meas_captured_at)))), "MISSING", IF(DAYS(today_date, DATEVALUE(meas_captured_at))>90, "STALE-RED: >90d", IF(DAYS(today_date, DATEVALUE(meas_captured_at))>30, "STALE-YELLOW: >30d", "OK")))',
        '=IFERROR(DAYS(today_date, DATEVALUE(meas_captured_at)) & " days", "N/A")', "30 days",
        "Re-run /saturation-finder"),
    ("Rate card age ≤ 180 days",
        f'=IF(DAYS(today_date, DATEVALUE("{rate.get("capturedAt", "2000-01-01")[:10]}"))>365, "EXPIRED", IF(DAYS(today_date, DATEVALUE("{rate.get("capturedAt", "2000-01-01")[:10]}"))>180, "STALE", "OK"))',
        f'=DAYS(today_date, DATEVALUE("{rate.get("capturedAt", "2000-01-01")[:10]}")) & " days"', "180 days",
        "Run refresh-rate-card.sh"),
    ("Binding layer within 80% utilization",
        '=IF(binding_util_pct >= 1, "RED: binding layer over 100%", IF(binding_util_pct >= 0.8, "AMBER", "OK"))',
        '=binding_layer', "80% util",
        "Address binding layer (Bottleneck tab)"),
    ("LLM model covered in rate card",
        f'="{"OK" if scenario["llmProfile"]["model"] in rate["llm"] else "RED: " + scenario["llmProfile"]["model"] + " not in rate card"}"',
        f'="{scenario["llmProfile"]["model"]}"', "", "Add model to RateCard.llm"),
    ("Redis cache budget covers sessions",
        '=IF(conc_sessions > fleet_cache_budget, "AMBER: active sessions exceed cache", "OK")',
        '=conc_sessions', '=fleet_cache_budget', "Grow maxInMemorySessionsPerPod OR shorter TTL"),
    # Tier-ceiling warnings — catch user picking a tier too small for target
    ("Mongo tier covers target msg/s",
        '=IF(ISNUMBER(baseline_mongo_ceiling), IF(baseline_mongo_ceiling >= target_msgps, "OK", "RED: tier ceiling " & baseline_mongo_ceiling & " < target " & target_msgps), "N/A")',
        '=baseline_mongo_ceiling', '=target_msgps', "Pick bigger Mongo tier on Controls!B12"),
    ("Redis tier covers target msg/s",
        '=IF(ISNUMBER(baseline_redis_ceiling), IF(baseline_redis_ceiling >= target_msgps, "OK", "RED: tier ceiling " & baseline_redis_ceiling & " < target " & target_msgps), "N/A")',
        '=baseline_redis_ceiling', '=target_msgps', "Pick bigger Redis tier on Controls!B13"),
    ("Runtime tier CPU scaling confidence",
        '=IF(runtime_cpu_scale=1, "OK (same as measurement)", "LOW-CONFIDENCE: per-pod capacity re-scaled by " & ROUND(runtime_cpu_scale,2) & "x — needs re-measurement")',
        '=runtime_cpu_scale', '=1', "Re-run /saturation-finder at the selected runtime tier"),
    ("Calculator mode (SNAPSHOT vs WHAT-IF)",
        '=IF(ISTEXT(calc_mode), calc_mode, "N/A")',
        '=calc_mode', '', "Edits to dropdowns switch mode to WHAT-IF"),
]
for i, (check, status_formula, value, threshold, remedy) in enumerate(warn_rows):
    r = 5 + i
    put(ws, r, 1, check, font=BOLD)
    put(ws, r, 2, status_formula)
    put(ws, r, 3, value)
    put(ws, r, 4, threshold)
    put(ws, r, 5, remedy, font=ITALIC)

status_range = f"B5:B{5 + len(warn_rows) - 1}"
for pattern, fill in [('"RED"', RED), ('"EXPIRED"', RED), ('"STALE-RED"', RED), ('"MISSING"', RED),
                      ('"AMBER"', AMBER), ('"STALE-YELLOW"', AMBER), ('"STALE"', AMBER),
                      ('"OK"', GREEN)]:
    ws.conditional_formatting.add(status_range, FormulaRule(formula=[f'ISNUMBER(SEARCH({pattern}, B5))'], fill=fill))

set_col_widths(ws, [42, 36, 22, 22, 52])

# ----------------------------------------------------------------------------
# Tab: Summary (insert as first)
# ----------------------------------------------------------------------------
ws = wb.create_sheet("Cost Summary", 0)
ws["A1"] = "Sizing & Cost — Summary (Live Calculator)"; ws["A1"].font = Font(bold=True, size=16)
ws["A2"] = (f"Scenario: {args.scenario} | Generated: {today} | "
            "🟢 GREEN tabs = editable inputs | 🔵 BLUE tabs = auto-calculated | ⬜ GRAY tabs = hidden reference data")
ws["A2"].font = ITALIC

# Mode banner
if baseline_tiers_catalog is not None:
    put(ws, 3, 1, "Mode:", font=BOLD)
    put(ws, 3, 2, "=calc_mode", fill=GREEN, font=BOLD)
    put(ws, 3, 4, "SNAPSHOT = at defaults (auditable). WHAT-IF = any tier dropdown edited.", font=ITALIC)

put(ws, 4, 1, "Target msg/s", font=BOLD)
put(ws, 4, 2, "=target_msgps", fill=GRAY)

# Live tier labels (pull from dropdowns)
if baseline_tiers_catalog is not None:
    put(ws, 4, 4, ('="Runtime: " & runtime_tier_selected & " | Mongo: " & '
                   'mongo_tier_selected & " | Redis: " & redis_tier_selected'),
        font=ITALIC)

put(ws, 5, 1, "Pods (avg load)", font=BOLD)
put(ws, 5, 2, "=pods_required", fill=GRAY)
put(ws, 5, 3, "", font=BOLD)
put(ws, 5, 4, "PRODUCTION PODS (peak+HA)", font=BOLD)

put(ws, 6, 1, "PRODUCTION PODS", font=Font(bold=True, size=12))
put(ws, 6, 2, "=pods_production", fill=GREEN, font=BIG_NUM)
put(ws, 6, 4, '="burst " & burst_factor & "× + N+" & availability_headroom_pods', font=ITALIC)

put(ws, 7, 1, "PROD monthly cost", font=Font(bold=True, size=12))
put(ws, 7, 2, "=prod_total_cost", fill=GREEN, font=BIG_NUM, number_format='"$"#,##0.00')

put(ws, 8, 1, "Cost per conversation", font=BOLD)
put(ws, 8, 2, "=cost_per_conversation", fill=GREEN, number_format='"$"#,##0.0000')
put(ws, 8, 4, '="(" & TEXT(conversations_per_mo, "#,##0") & " conv/mo)"', font=ITALIC)

# Cost breakdown
put(ws, 9, 1, "COST BREAKDOWN", font=Font(bold=True, size=12), fill=LIGHT_BLUE)
write_headers(ws, 10, ["Component", "Monthly $", "% of total", "Notes"])
brk = [
    ("Baseline (always-on)", "=baseline_total", "=baseline_total/total_cost", "Infra floor"),
    ("Runtime scaling", "=MAX(0, delta_runtime)", "=B12/total_cost", "Extra pods beyond 1"),
    ("Mongo extra", "=MAX(0, delta_mongo)", "=B13/total_cost", "Above Mongo baseline"),
    ("Redis extra", "=MAX(0, delta_redis)", "=B14/total_cost", ""),
    ("LLM", "=llm_cost", "=llm_cost/total_cost", "0 for mock"),
    ("Egress", "=egress_cost", "=egress_cost/total_cost", ""),
]
for i, (label, val, pct, note) in enumerate(brk):
    r = 11 + i
    put(ws, r, 1, label, font=BOLD)
    put(ws, r, 2, val, fill=GRAY, number_format='"$"#,##0.00')
    put(ws, r, 3, pct, fill=GRAY, number_format="0.0%")
    put(ws, r, 4, note, font=ITALIC)
put(ws, 17, 1, "TOTAL (sum check)", font=BOLD, fill=LIGHT_BLUE)
put(ws, 17, 2, "=SUM(B11:B16)", fill=GREEN, font=BOLD, number_format='"$"#,##0.00')
put(ws, 17, 3, "=SUM(C11:C16)", fill=GREEN, font=BOLD, number_format="0.0%")

# Binding layer
put(ws, 19, 1, "BINDING LAYER", font=Font(bold=True, size=12), fill=LIGHT_BLUE)
put(ws, 20, 1, "Which layer saturates first?", font=BOLD)
put(ws, 20, 2, "=binding_layer", fill=AMBER, font=BOLD)
put(ws, 21, 1, "Utilization at current target", font=BOLD)
put(ws, 21, 2, "=binding_util_pct", fill=AMBER, number_format="0.0%", font=BOLD)

# Warnings
put(ws, 23, 1, "WARNINGS ACTIVE", font=Font(bold=True, size=12), fill=LIGHT_BLUE)
put(ws, 24, 1, "RED count", font=BOLD)
_warn_end = 5 + len(warn_rows) - 1
put(ws, 24, 2, f'=COUNTIF(Warnings!B5:B{_warn_end}, "RED*") + COUNTIF(Warnings!B5:B{_warn_end}, "EXPIRED*") + COUNTIF(Warnings!B5:B{_warn_end}, "STALE-RED*") + COUNTIF(Warnings!B5:B{_warn_end}, "MISSING*")',
    fill=RED, font=BOLD)
put(ws, 25, 1, "AMBER count", font=BOLD)
put(ws, 25, 2, f'=COUNTIF(Warnings!B5:B{_warn_end}, "AMBER*") + COUNTIF(Warnings!B5:B{_warn_end}, "STALE-YELLOW*") + COUNTIF(Warnings!B5:B{_warn_end}, "STALE*")',
    fill=AMBER, font=BOLD)

# Sanity
put(ws, 27, 1, "SANITY (baseline + scaling = total)", font=Font(bold=True, size=12), fill=LIGHT_BLUE)
put(ws, 28, 1, "Difference", font=BOLD)
put(ws, 28, 2, "=ABS(baseline_total + scaling_cost - total_cost)", fill=GRAY, number_format='"$"#,##0.0000')
put(ws, 29, 1, "Match?", font=BOLD)
put(ws, 29, 2, '=IF(B28<0.01, "✓ PASS", "✗ FAIL")', fill=GREEN, font=BOLD)

set_col_widths(ws, [36, 22, 14, 40])

# ============================================================================
# Tab colors — visual grouping for usability
# GREEN  = editable inputs (change these to explore sizing)
# BLUE   = live calculated outputs (auto-update from inputs)
# GRAY   = reference/evidence snapshots (read-only context)
# ============================================================================
TAB_GREEN = "70AD47"   # editable inputs
TAB_BLUE = "4472C4"    # live calculated outputs
TAB_GRAY = "A5A5A5"    # static reference/evidence

tab_colors = {
    "Cost Summary": TAB_BLUE,     # output: headline results
    "Controls": TAB_GREEN,        # INPUT: target msg/s, tiers, planning params
    "RateCard": TAB_GREEN,        # INPUT: unit prices (editable)
    "Scenario": TAB_GREEN,        # INPUT: LLM profile, session pattern
    "Infra": TAB_BLUE,            # output: current + required (auto-calc section)
    "BaselineSource": TAB_GRAY,   # reference: tier catalog (lookup table)
    "Measurement": TAB_GRAY,      # reference: saturation test results
    "TestEvidence": TAB_GRAY,     # reference: per-step scorecard snapshot
    "Provenance": TAB_GRAY,       # reference: data source freshness
    "Baseline": TAB_BLUE,         # output: always-on cost floor
    "Scaling": TAB_BLUE,          # output: sizing formulas F-1..F-8
    "Bottleneck": TAB_BLUE,       # output: which layer saturates first
    "WhatIf": TAB_BLUE,           # output: target sweep scenarios
    "Projections": TAB_BLUE,      # output: growth projections
    "Warnings": TAB_BLUE,         # output: live contradiction flags
}
for sheet_name, color in tab_colors.items():
    if sheet_name in wb.sheetnames:
        wb[sheet_name].sheet_properties.tabColor = color

# Hide reference/evidence tabs — users don't need to see these directly.
# They contain named cells that formulas reference. Unhide to inspect raw data.
hidden_tabs = ["BaselineSource", "Measurement", "TestEvidence", "Provenance", "Baseline"]
for sheet_name in hidden_tabs:
    if sheet_name in wb.sheetnames:
        wb[sheet_name].sheet_state = "hidden"

# Reorder tabs: Inputs first, then Outputs, then hidden (already hidden)
desired_order = [
    "Cost Summary",   # BLUE — executive summary
    "Controls",       # GREEN — primary inputs
    "Infra",          # BLUE — current vs required
    "Scaling",        # BLUE — detailed formulas
    "Bottleneck",     # BLUE — saturation forecast
    "WhatIf",         # BLUE — target sweep
    "Projections",    # BLUE — growth projections
    "Warnings",       # BLUE — contradiction flags
    "RateCard",       # GREEN — unit prices
    "Scenario",       # GREEN — workload shape
    # Hidden tabs last (not visible anyway)
    "BaselineSource",
    "Measurement",
    "TestEvidence",
    "Provenance",
    "Baseline",
]
# Only reorder tabs that exist
existing = [s for s in desired_order if s in wb.sheetnames]
# Add any tabs not in desired_order at the end
for s in wb.sheetnames:
    if s not in existing:
        existing.append(s)
wb._sheets = [wb[s] for s in existing]

# ============================================================================
# Save (formulas-only first pass)
# ============================================================================

out_path = Path(args.out)
out_path.parent.mkdir(parents=True, exist_ok=True)
wb.save(out_path)

# ============================================================================
# Cache formula results into the workbook so values display immediately
# regardless of viewer — Excel on Mac without auto-recalc, Google Sheets on
# first open, BI tools that read cached values directly, data-only openpyxl
# loads, etc.
#
# Strategy:
#   1. Evaluate every formula with the `formulas` library → get scalar values.
#   2. Post-process the xlsx ZIP directly. For each <c> element that has an
#      <f> formula child, inject or replace its <v> cached-value child with
#      the evaluated result. This is the authoritative xlsx caching mechanism
#      — <f> holds the formula, <v> holds the last computed value.
#   3. Also set <workbook.calcPr fullCalcOnLoad="1"/> so any viewer that
#      *does* recalculate picks fresh values on open.
# ============================================================================
try:
    import formulas as fx
    import zipfile
    import shutil
    from io import BytesIO
    from xml.etree import ElementTree as ET

    print(f"[cache] evaluating {out_path.name} formulas …")
    xl = fx.ExcelModel().loads(str(out_path)).finish()
    sol = xl.calculate()

    def _unwrap(v):
        """formulas returns numpy arrays / nested lists — unwrap to scalar."""
        if hasattr(v, "tolist"):
            v = v.tolist()
        while isinstance(v, list) and len(v) == 1:
            v = v[0]
        if isinstance(v, list):
            return v[0] if v else None
        return v

    def _format_value(v):
        """Convert Python value to xlsx-cell string and xlsx type.
        Handles numpy scalars, bool, int, float, str. Returns (None, None)
        for unrepresentable or error values.
        """
        if v is None:
            return None, None
        # formulas returns 'Error' class instances for #N/A, #REF!, etc.
        if isinstance(v, Exception):
            return None, None
        tname = type(v).__name__
        if tname in ("XlErrors", "XlError"):
            return None, None
        # numpy scalars → unwrap via .item()
        if hasattr(v, "item") and not isinstance(v, str):
            try:
                v = v.item()
            except Exception:
                pass
        # Booleans (check before int — bool is a subclass)
        if isinstance(v, bool):
            return ("1" if v else "0"), "b"
        # Numbers
        if isinstance(v, (int, float)):
            import math
            if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
                return None, None
            if isinstance(v, float):
                # Use repr() for max precision, strip trailing zeros later
                # via the spreadsheet's number format. Excel tolerates scientific.
                return repr(v), None
            return str(int(v)), None  # numeric, no t= attr
        # Strings — inline string (t="str")
        return str(v), "str"

    # Build a map: {sheet_xml_filename: {coord: (value_str, type)}}
    # We need to know each sheet's xml filename inside xl/worksheets/
    # openpyxl uses sheet1.xml, sheet2.xml, ... in order of wb.worksheets
    sheet_xml_map = {}
    for i, sheet_name in enumerate(wb.sheetnames, start=1):
        sheet_upper = sheet_name.upper()
        xml_name = f"xl/worksheets/sheet{i}.xml"
        cell_values = {}
        for key, info in sol.items():
            # key format: "'[filename.xlsx]SHEETNAME'!A1"
            if not key.startswith(f"'[{out_path.name}]{sheet_upper}'!"):
                continue
            coord = key.split("!", 1)[1]
            try:
                v = _unwrap(info.value)
                val_str, val_type = _format_value(v)
                if val_str is not None:
                    cell_values[coord] = (val_str, val_type)
            except Exception:
                pass  # skip cells that can't be evaluated
        sheet_xml_map[xml_name] = cell_values

    # Post-process the xlsx as a zip file
    NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    ET.register_namespace("", NS)

    tmp_path = out_path.with_suffix(".xlsx.tmp")
    total_patched = 0
    with zipfile.ZipFile(out_path, "r") as zin, \
         zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.namelist():
            data = zin.read(item)
            # Force Excel to recalculate all formulas on open
            if item == "xl/workbook.xml":
                wb_tree = ET.parse(BytesIO(data))
                wb_root = wb_tree.getroot()
                calc_pr = wb_root.find(f"{{{NS}}}calcPr")
                if calc_pr is None:
                    calc_pr = ET.SubElement(wb_root, f"{{{NS}}}calcPr")
                calc_pr.set("calcId", "0")
                calc_pr.set("fullCalcOnLoad", "1")
                calc_pr.set("forceFullCalc", "1")
                buf = BytesIO()
                wb_tree.write(buf, xml_declaration=True, encoding="UTF-8")
                data = buf.getvalue()
            if item in sheet_xml_map and sheet_xml_map[item]:
                # Patch this sheet's XML
                tree = ET.parse(BytesIO(data))
                root = tree.getroot()
                cell_values = sheet_xml_map[item]
                patched_in_sheet = 0
                for c in root.iter(f"{{{NS}}}c"):
                    ref = c.get("r")
                    if ref not in cell_values:
                        continue
                    # Only patch if the cell has a formula child
                    f_elem = c.find(f"{{{NS}}}f")
                    if f_elem is None:
                        continue
                    val_str, val_type = cell_values[ref]
                    # Find or create <v> child
                    v_elem = c.find(f"{{{NS}}}v")
                    if v_elem is None:
                        v_elem = ET.SubElement(c, f"{{{NS}}}v")
                    v_elem.text = val_str
                    # Set cell type attr if the result is string/bool
                    if val_type:
                        c.set("t", val_type)
                    elif "t" in c.attrib:
                        # numeric result: clear any existing t= (e.g. t="str" from prior save)
                        del c.attrib["t"]
                    patched_in_sheet += 1
                # Serialize the patched XML back
                buf = BytesIO()
                tree.write(buf, xml_declaration=True, encoding="UTF-8")
                data = buf.getvalue()
                total_patched += patched_in_sheet
            zout.writestr(item, data)

    shutil.move(tmp_path, out_path)
    print(f"[cache] patched {total_patched} formula cells with cached values")

    # IMPORTANT: we do NOT re-open the xlsx with openpyxl here, because
    # openpyxl's save() strips the <v> cached values we just wrote (it only
    # knows how to serialize data_type='f' as <f> without <v>). The xlsx
    # already has the correct <calcPr> from openpyxl's first save; setting
    # fullCalcOnLoad is a nice-to-have but not worth losing cached values.
    #
    # If we need fullCalcOnLoad later, patch xl/workbook.xml directly in the
    # same zip pass above — not via openpyxl round-trip.

except ImportError as e:
    print(f"[cache] required library missing ({e}); skipping caching. "
          f"Install: pip install formulas")
except Exception as e:
    import traceback
    print(f"[cache] caching step failed ({e}); xlsx will still compute on open.")
    traceback.print_exc()

print(f"wrote {out_path}")
print(f"Tabs ({len(wb.sheetnames)}):")
for i, name in enumerate(wb.sheetnames, start=1):
    print(f"  {i:2d}. {name}")
print()
if meas is None:
    print("NOTE: No measurement provided — Scaling/Bottleneck/WhatIf tabs will show #N/A.")
    print("      Run /saturation-finder + import to populate.")
print("Upload to Drive → right-click → Open with Sheets → share with team.")
