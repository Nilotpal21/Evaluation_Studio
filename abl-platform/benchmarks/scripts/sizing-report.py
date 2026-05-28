#!/usr/bin/env python3
"""
sizing-report.py — Clean 2-tab sizing & cost spreadsheet.

Tab 1: Calculator — infra overview, VM types, editable inputs, pod sizing, cost breakdown, bottleneck check
Tab 2: Benchmarks — saturation test evidence, per-step scorecard, provenance

Usage:
  python3 benchmarks/scripts/sizing-report.py \
    --env qa \
    --measurement benchmarks/config/sizing/measurements/7456464-2026-05-05.json \
    --out benchmarks/docs/sizing-report.xlsx
"""

import argparse
import json
import subprocess
import sys
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

# ============================================================================
# CLI
# ============================================================================
parser = argparse.ArgumentParser()
parser.add_argument("--env", default=None, help="Environment (dev/qa/staging) — auto-refreshes infra from kubectl")
parser.add_argument("--scenario", default="chat-agent-mock")
parser.add_argument("--measurement", required=True, help="Path to saturation-measurement JSON")
parser.add_argument("--sat-report", default=None, help="Path to saturation-run markdown report")
parser.add_argument("--rate-card", default="benchmarks/config/sizing/rate-card.json")
parser.add_argument("--infra", default="benchmarks/config/sizing/infra-snapshot.json")
parser.add_argument("--scenario-file", default=None)
parser.add_argument("--out", required=True)
args = parser.parse_args()

if args.scenario_file is None:
    args.scenario_file = f"benchmarks/config/sizing/scenarios/{args.scenario}.json"


def load_json(path):
    return json.loads(Path(path).read_text())


# Auto-refresh infra from kubectl
if args.env:
    refresh_script = Path(__file__).parent / "refresh-infra-snapshot.sh"
    if refresh_script.exists():
        print(f"[infra] Refreshing from kubectl (env={args.env})...")
        result = subprocess.run(
            ["bash", str(refresh_script)],
            env={**__import__("os").environ, "ENV": args.env},
            capture_output=True, text=True,
            cwd=str(Path(__file__).parent.parent.parent)
        )
        if result.returncode == 0:
            for line in result.stdout.strip().split("\n"):
                if line.startswith("[refresh]"):
                    print(f"  {line}")
        else:
            print(f"  WARNING: refresh failed: {result.stderr.strip()[:200]}")

# Load inputs
rate = load_json(args.rate_card)
infra = load_json(args.infra)
scenario = load_json(args.scenario_file)
meas = load_json(args.measurement)

today = datetime.now(timezone.utc).strftime("%Y-%m-%d")

# ============================================================================
# Extract values
# ============================================================================
# Measurement
per_pod = meas["perPodCapacity"]["maxSafeMsgPerSec"]
p95_at_max = meas["perPodCapacity"].get("p95AtMaxMs", 1500)
bottleneck = meas["perPodCapacity"].get("primaryBottleneck", "unknown")
max_pods = max(p["pods"] for p in meas["multiPodScaling"])
fleet_max = max(p["totalMsgPerSec"] for p in meas["multiPodScaling"])
wi = meas.get("workloadIntrinsic", {})
mongo_cpu_per_msg = wi.get("mongoCpuMilliPerMsg", 0)
redis_cpu_per_msg = wi.get("redisCpuMilliPerMsg", 0)
mongo_iops_per_msg = wi.get("mongoIopsPerMsg", scenario["perMessageDatastoreCost"]["mongoWrites"])

# Infra
rt = infra.get("runtimeDefaults", {})
rt_cpu_req = rt.get("cpuRequestCores") or 1
rt_cpu_lim = rt.get("cpuLimitCores") or 4
rt_mem_req = rt.get("memoryRequestGi") or 1
rt_mem_lim = rt.get("memoryLimitGi") or 2
rt_hpa_min = rt.get("hpaMinReplicas", 2)
rt_hpa_max = rt.get("hpaMaxReplicas", 10)
rt_instance = rt.get("placement", {}).get("instanceType", "?")

mg = infra.get("mongodb", {})
mongo_replicas = mg.get("replicas", 3)
mongo_cpu_lim = mg.get("cpuLimitCores") or 16
mongo_mem_lim = mg.get("memoryLimitGi") or 64
mongo_pvc_gi = mg.get("pvcSizeGi", 256)
mongo_iops_lim = mg.get("diskIopsLimit") or 500
mongo_disk = mg.get("diskSku", "?")

rd = infra.get("redis", {})
redis_master_cpu = rd.get("master", {}).get("cpuLimitCores") or 8
redis_master_mem = rd.get("master", {}).get("memoryLimitGi") or 16
redis_rep_cpu = rd.get("replica", {}).get("cpuLimitCores") or 4
redis_reps = rd.get("replicaReplicas", 3)
redis_total_cpu = redis_master_cpu + redis_rep_cpu * redis_reps

# Rate card
cpu_hr = rate["compute"]["cpuCorePerHour"]
mem_hr = rate["compute"]["memoryGBPerHour"]
retention = rate["mongodb"]["retentionDays"]
doc_kb = rate["mongodb"]["avgDocSizeKB"]
mongo_storage_rate = rate["mongodb"]["storageGBPerMonth"]
session_kb = rate["redis"]["avgSessionSizeKB"]

# Scenario
turn = scenario["requestShape"].get("turnStructure", {})
msgs_per_session = turn.get("creates", 1) + turn.get("followups", 4)
session_ttl = scenario["sessionPattern"]["sessionTTLMin"]
mongo_writes = scenario["perMessageDatastoreCost"]["mongoWrites"]
redis_ops = scenario["perMessageDatastoreCost"]["redisOps"]
llm_mode = scenario["llmProfile"]["mode"]
llm_model = scenario["llmProfile"].get("model", "mock")
tokens_in = scenario["llmProfile"].get("tokensInPerTurn", 0)
tokens_out = scenario["llmProfile"].get("tokensOutPerTurn", 0)
llm_input_rate = rate["llm"].get(llm_model, {}).get("inputPerM", 0)
llm_output_rate = rate["llm"].get(llm_model, {}).get("outputPerM", 0)

# k6 summary
k6s = meas.get("k6Summary", {})

# Coroot
coroot = meas.get("corootMetrics", {})

# Node capacity (how many runtime pods fit per user-pool node)
# Use CPU LIMIT for packing — pods burst to limit under load, so can't overpack
nodes = infra.get("nodes", [])
node_cpu = next((n.get("cpuPerNode", 8) for n in nodes if "user" in n.get("pool", "").lower()), 8)
node_mem = next((n.get("memoryPerNodeGi", 32) for n in nodes if "user" in n.get("pool", "").lower()), 32)
pods_per_node = max(1, min(int(node_cpu // rt_cpu_lim), int(node_mem // rt_mem_lim)))

# ============================================================================
# Styles
# ============================================================================
BOLD = Font(bold=True)
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
INPUT_FILL = PatternFill("solid", fgColor="E2EFDA")  # light green
OUTPUT_FILL = PatternFill("solid", fgColor="DDEBF7")  # light blue
WARN_FILL = PatternFill("solid", fgColor="FFC7CE")  # light red
GRAY_FILL = PatternFill("solid", fgColor="F2F2F2")
SECTION_FILL = PatternFill("solid", fgColor="D9E2F3")
MONEY_FMT = '"$"#,##0'
NUM_FMT = "#,##0.00"
UNLOCKED = Protection(locked=False)
INPUT_BORDER = Border(
    left=Side("medium", "4472C4"), right=Side("medium", "4472C4"),
    top=Side("medium", "4472C4"), bottom=Side("medium", "4472C4")
)


def write_header(ws, row, cols):
    for i, col in enumerate(cols, 1):
        c = ws.cell(row, i, col)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", wrap_text=True)


def section(ws, row, title):
    ws.cell(row, 1, title).font = Font(bold=True, size=12)
    ws.cell(row, 1).fill = SECTION_FILL
    return row + 1


def put(ws, row, col, val, fmt=None, fill=None, font=None):
    c = ws.cell(row, col, val)
    if fmt:
        c.number_format = fmt
    if fill:
        c.fill = fill
    if font:
        c.font = font
    return c


def add_name(wb, name, ref):
    from openpyxl.workbook.defined_name import DefinedName
    dn = DefinedName(name, attr_text=ref)
    wb.defined_names.add(dn)


# ============================================================================
# TAB 1: Calculator
# ============================================================================
wb = Workbook()
ws = wb.active
ws.title = "Calculator"
ws.sheet_properties.tabColor = "4472C4"

r = 1
put(ws, r, 1, "Infrastructure Sizing & Cost Calculator", font=Font(bold=True, size=16)); r += 1
put(ws, r, 1, f"Scenario: {args.scenario} | Generated: {today} | Edit GREEN cells → everything recalculates",
    font=Font(italic=True)); r += 2

# --- CURRENT INFRASTRUCTURE (compact reference) ---
r = section(ws, r, "CURRENT CLUSTER (reference)")
put(ws, r, 1, "Runtime", font=BOLD)
put(ws, r, 2, f"{rt_cpu_req}c/{rt_cpu_lim}c CPU, {rt_mem_req}/{rt_mem_lim}Gi, HPA {rt_hpa_min}-{rt_hpa_max}, {rt_instance}")
r += 1
put(ws, r, 1, "MongoDB", font=BOLD)
mongo_cpu_display = mg.get("cpuLimitCores")
mongo_cpu_str = f"{mongo_cpu_display}c" if mongo_cpu_display else "no limit"
put(ws, r, 2, f"{mongo_replicas}× replicas, {mongo_cpu_str} CPU limit, {mongo_mem_lim}Gi, {mongo_pvc_gi}Gi disk ({mongo_disk})")
r += 1
put(ws, r, 1, "Redis", font=BOLD)
put(ws, r, 2, f"master {redis_master_cpu}c/{redis_master_mem}Gi + {redis_reps}× replica {redis_rep_cpu}c")
r += 2

# --- INPUTS ---
r = section(ws, r, "SIZING INPUTS (edit green cells)")
write_header(ws, r, ["Parameter", "Value", "Unit", "Notes"]); r += 1

input_start = r
# (label, value, unit, note, editable)
inputs = [
    ("Target msg/s", fleet_max, "msg/s", "Primary sizing driver — change this!", True),
    ("p95 target", 1300, "ms", "Latency SLO", False),
    ("Utilization target", 0.80, "ratio", "0.80 = 80% headroom", True),
    ("Burst factor (peak:avg)", 3, "×", "Peak traffic / average traffic", True),
    ("Availability headroom (N+K)", 1, "pods", "Extra pods for rolling deploy", True),
    ("Growth rate (monthly)", 0.10, "% per month", "For projections", True),
]
target_row = r  # row where target_msgps lives
input_cells = []
for label, val, unit, note, editable in inputs:
    put(ws, r, 1, label, font=BOLD)
    if editable:
        c = put(ws, r, 2, val, fill=INPUT_FILL, font=Font(bold=True, size=12))
        c.border = INPUT_BORDER
        input_cells.append(c)
    else:
        put(ws, r, 2, val, fill=GRAY_FILL, font=Font(bold=True, size=12))
    put(ws, r, 3, unit)
    put(ws, r, 4, note, font=Font(italic=True))
    r += 1

# Named cells for formulas
add_name(wb, "target_msgps", f"'Calculator'!$B${target_row}")
add_name(wb, "p95_target", f"'Calculator'!$B${target_row+1}")
add_name(wb, "util_target", f"'Calculator'!$B${target_row+2}")
add_name(wb, "burst_factor", f"'Calculator'!$B${target_row+3}")
add_name(wb, "avail_headroom", f"'Calculator'!$B${target_row+4}")
add_name(wb, "growth_rate", f"'Calculator'!$B${target_row+5}")

r += 1

# --- POD SIZING ---
r = section(ws, r, "POD SIZING (auto-calculated)")
write_header(ws, r, ["Metric", "Value", "Unit"]); r += 1

put(ws, r, 1, "Per-pod capacity (measured)", font=BOLD)
put(ws, r, 2, per_pod, fill=OUTPUT_FILL)
put(ws, r, 3, "msg/s/pod")
add_name(wb, "per_pod", f"'Calculator'!$B${r}")
per_pod_row = r; r += 1

put(ws, r, 1, "Pods required (avg load)", font=BOLD)
put(ws, r, 2, f"=CEILING(target_msgps/(per_pod*util_target), 1)", fill=OUTPUT_FILL)
put(ws, r, 3, "pods")
add_name(wb, "pods_required", f"'Calculator'!$B${r}")
r += 1

put(ws, r, 1, "Pods for peak (burst)", font=BOLD)
put(ws, r, 2, f"=CEILING(target_msgps*burst_factor/(per_pod*util_target), 1)", fill=OUTPUT_FILL)
put(ws, r, 3, "pods")
add_name(wb, "pods_peak", f"'Calculator'!$B${r}")
r += 1

put(ws, r, 1, "PRODUCTION PODS (recommended)", font=Font(bold=True, size=12))
put(ws, r, 2, f"=pods_peak + avail_headroom", fill=OUTPUT_FILL,
    font=Font(bold=True, size=12))
put(ws, r, 3, "pods")
add_name(wb, "prod_pods", f"'Calculator'!$B${r}")
r += 1

put(ws, r, 1, "Runtime nodes required", font=BOLD)
put(ws, r, 2, f"=CEILING(prod_pods/{pods_per_node}, 1)", fill=OUTPUT_FILL)
put(ws, r, 3, "nodes")
add_name(wb, "runtime_nodes", f"'Calculator'!$B${r}")
r += 1

put(ws, r, 1, "Fleet throughput (actual)", font=BOLD)
put(ws, r, 2, f"=prod_pods*per_pod", fill=OUTPUT_FILL)
put(ws, r, 3, "msg/s")
r += 2

# --- REQUIRED INFRASTRUCTURE (dynamic) ---
r = section(ws, r, "REQUIRED INFRASTRUCTURE (auto-calculated)")
write_header(ws, r, ["Component", "Count", "VM Type", "Per-node spec", "How calculated"]); r += 1

# Gather node pool info
user_pool = next((n for n in nodes if "user" in n.get("pool", "").lower()), {})
db_pool = next((n for n in nodes if "database" in n.get("pool", "").lower() or "db" in n.get("pool", "").lower()), {})
gpu_pool = next((n for n in nodes if "gpu" in n.get("pool", "").lower()), {})
sys_pool = next((n for n in nodes if "system" in n.get("pool", "").lower()), {})

user_vm = user_pool.get("instanceType", rt_instance)
user_cpu_per = user_pool.get("cpuPerNode", node_cpu)
user_mem_per = user_pool.get("memoryPerNodeGi", node_mem)

db_vm = db_pool.get("instanceType", "Standard_D16s_v5")
db_cpu_per = db_pool.get("cpuPerNode", 16)
db_mem_per = db_pool.get("memoryPerNodeGi", 64)

# Runtime/user nodes: driven by pod count
put(ws, r, 1, "Runtime / User nodes", font=Font(bold=True, size=11))
put(ws, r, 2, f"=CEILING(prod_pods/{pods_per_node}, 1)", fill=OUTPUT_FILL, font=Font(bold=True, size=11))
put(ws, r, 3, user_vm)
put(ws, r, 4, f"{user_cpu_per} vCPU / {user_mem_per} GiB")
put(ws, r, 5, f"= prod_pods / {pods_per_node} pods/node ({rt_cpu_lim}c limit/pod on {node_cpu}c node)")
add_name(wb, "user_nodes", f"'Calculator'!$B${r}")
r += 1

# Database nodes: based on mongo replica count (fixed unless mongo scales)
mongo_pods_per_db_node = max(1, int(db_cpu_per // (mongo_cpu_lim if mongo_cpu_lim < db_cpu_per else db_cpu_per)))
put(ws, r, 1, "Database nodes", font=Font(bold=True, size=11))
db_node_count = max(db_pool.get("count", 2), mongo_replicas)  # at least enough for mongo replicas
put(ws, r, 2, db_node_count, fill=GRAY_FILL, font=Font(bold=True, size=11))
put(ws, r, 3, db_vm)
put(ws, r, 4, f"{db_cpu_per} vCPU / {db_mem_per} GiB")
put(ws, r, 5, f"Fixed: {mongo_replicas} Mongo + Redis (shared)")
r += 1

# GPU nodes (fixed, for ML workloads)
if gpu_pool:
    put(ws, r, 1, "GPU nodes", font=Font(bold=True, size=11))
    put(ws, r, 2, gpu_pool.get("count", 0), fill=GRAY_FILL, font=Font(bold=True, size=11))
    put(ws, r, 3, gpu_pool.get("instanceType", "?"))
    put(ws, r, 4, f"{gpu_pool.get('cpuPerNode', '?')} vCPU / {gpu_pool.get('memoryPerNodeGi', '?')} GiB")
    put(ws, r, 5, "Fixed: ML/embedding workloads")
    r += 1

# System nodes (fixed)
if sys_pool:
    put(ws, r, 1, "System nodes", font=Font(bold=True, size=11))
    put(ws, r, 2, sys_pool.get("count", 2), fill=GRAY_FILL, font=Font(bold=True, size=11))
    put(ws, r, 3, sys_pool.get("instanceType", "?"))
    put(ws, r, 4, f"{sys_pool.get('cpuPerNode', '?')} vCPU / {sys_pool.get('memoryPerNodeGi', '?')} GiB")
    put(ws, r, 5, "Fixed: k8s system, monitoring")
    r += 1

# Total nodes
put(ws, r, 1, "TOTAL NODES", font=Font(bold=True, size=12))
fixed_nodes = db_node_count + gpu_pool.get("count", 0) + sys_pool.get("count", 0)
put(ws, r, 2, f"=user_nodes+{fixed_nodes}", fill=OUTPUT_FILL, font=Font(bold=True, size=12))
put(ws, r, 3, "", font=BOLD)
put(ws, r, 4, "")
put(ws, r, 5, f"= user_nodes (dynamic) + {fixed_nodes} (fixed db+gpu+sys)")
add_name(wb, "total_nodes", f"'Calculator'!$B${r}")
r += 2

# --- COST BREAKDOWN ---
r = section(ws, r, "MONTHLY COST BREAKDOWN")
write_header(ws, r, ["Component", "Monthly $"]); r += 1

# Runtime cost
rt_cost_formula = f"=prod_pods*{rt_cpu_lim}*730*{cpu_hr} + prod_pods*{rt_mem_lim}*730*{mem_hr}"
put(ws, r, 1, "Runtime compute", font=BOLD)
put(ws, r, 2, rt_cost_formula, fill=OUTPUT_FILL, fmt=MONEY_FMT)
add_name(wb, "runtime_cost", f"'Calculator'!$B${r}")
r += 1

# Mongo cost
mongo_compute = mongo_cpu_lim * mongo_replicas * 730 * cpu_hr + mongo_mem_lim * mongo_replicas * 730 * mem_hr
mongo_storage_formula = f"=target_msgps*{mongo_writes}*86400*{retention}*{doc_kb}/(1024*1024)*{mongo_replicas}*1.3*{mongo_storage_rate}"
put(ws, r, 1, "MongoDB (compute)", font=BOLD)
put(ws, r, 2, mongo_compute, fill=GRAY_FILL, fmt=MONEY_FMT)
r += 1

put(ws, r, 1, "MongoDB (storage)", font=BOLD)
put(ws, r, 2, mongo_storage_formula, fill=OUTPUT_FILL, fmt=MONEY_FMT)
add_name(wb, "mongo_storage_cost", f"'Calculator'!$B${r}")
r += 1

# Redis cost
redis_compute = (redis_master_cpu * 730 * cpu_hr + redis_master_mem * 730 * mem_hr +
                 redis_reps * (redis_rep_cpu * 730 * cpu_hr + rd.get("replica", {}).get("memoryLimitGi", 8) * 730 * mem_hr))
put(ws, r, 1, "Redis (compute)", font=BOLD)
put(ws, r, 2, redis_compute, fill=GRAY_FILL, fmt=MONEY_FMT)
r += 1

# LLM cost
if tokens_in > 0:
    llm_formula = f"=(target_msgps*86400*30*{tokens_in}/1000000)*{llm_input_rate} + (target_msgps*86400*30*{tokens_out}/1000000)*{llm_output_rate}"
else:
    llm_formula = 0
put(ws, r, 1, f"LLM ({llm_model})", font=BOLD)
put(ws, r, 2, llm_formula, fill=OUTPUT_FILL if tokens_in > 0 else GRAY_FILL, fmt=MONEY_FMT)
add_name(wb, "llm_cost", f"'Calculator'!$B${r}")
r += 1

# Egress
put(ws, r, 1, "Egress", font=BOLD)
egress_formula = f"=target_msgps*{scenario['requestShape']['payloadBytesOut']}*86400*30/(1024*1024*1024)*{rate['egressGBPerMonth']}"
put(ws, r, 2, egress_formula, fill=OUTPUT_FILL, fmt=MONEY_FMT)
r += 1

# Total
put(ws, r, 1, "TOTAL MONTHLY", font=Font(bold=True, size=13))
total_formula = f"=runtime_cost + {mongo_compute} + mongo_storage_cost + {redis_compute} + llm_cost"
put(ws, r, 2, total_formula, fill=OUTPUT_FILL, font=Font(bold=True, size=13), fmt=MONEY_FMT)
add_name(wb, "total_cost", f"'Calculator'!$B${r}")
r += 2

# --- CAPACITY CHECK ---
r = section(ws, r, "CAPACITY CHECK (at target msg/s)")
write_header(ws, r, ["Layer", "Required", "Utilization", "Status"]); r += 1

cap_check_data = [
    ("Runtime pods", "=pods_required",
     f"=pods_required/{rt_hpa_max}", f'=IF(pods_required<={rt_hpa_max},"OK","SCALE HPA")'),
    ("Mongo CPU (cores)", f"=target_msgps*{mongo_cpu_per_msg}/700",
     f"=target_msgps*{mongo_cpu_per_msg}/700/{mongo_cpu_lim * mongo_replicas}", f'=IF(target_msgps*{mongo_cpu_per_msg}/700/{mongo_cpu_lim * mongo_replicas}<0.8,"OK","SCALE")'),
    ("Mongo IOPS", f"=target_msgps*{mongo_iops_per_msg}/0.7",
     f"=target_msgps*{mongo_iops_per_msg}/0.7/{mongo_iops_lim}", f'=IF(target_msgps*{mongo_iops_per_msg}/0.7/{mongo_iops_lim}<0.8,"OK","UPGRADE DISK")'),
    ("Mongo storage (GB)", f"=target_msgps*{mongo_writes}*86400*{retention}*{doc_kb}/(1024*1024)*{mongo_replicas}*1.3",
     f"=target_msgps*{mongo_writes}*86400*{retention}*{doc_kb}/(1024*1024)*{mongo_replicas}*1.3/{mongo_pvc_gi * mongo_replicas}", f'=IF(target_msgps*{mongo_writes}*86400*{retention}*{doc_kb}/(1024*1024)*{mongo_replicas}*1.3/{mongo_pvc_gi * mongo_replicas}<0.8,"OK","EXPAND PVC")'),
    ("Redis CPU (cores)", f"=target_msgps*{redis_cpu_per_msg}/500",
     f"=target_msgps*{redis_cpu_per_msg}/500/{redis_total_cpu}", f'=IF(target_msgps*{redis_cpu_per_msg}/500/{redis_total_cpu}<0.8,"OK","SCALE REDIS")'),
]
for label, req, util_f, status_f in cap_check_data:
    put(ws, r, 1, label, font=BOLD)
    put(ws, r, 2, req, fill=OUTPUT_FILL, fmt=NUM_FMT)
    put(ws, r, 3, util_f, fill=OUTPUT_FILL, fmt="0%")
    put(ws, r, 4, status_f, fill=OUTPUT_FILL)
    r += 1

r += 1

# --- WHAT-IF ---
r = section(ws, r, "WHAT-IF SCENARIOS (auto-calculated from target)")
write_header(ws, r, ["Target msg/s", "Pods needed", "Runtime Nodes", "Runtime $/mo", "Mongo storage $/mo", "LLM $/mo", "Total $/mo"]); r += 1

multipliers = [0.5, 1, 2, 5, 10]
for mult in multipliers:
    # pods = CEILING(target * burst / (per_pod * utilization)) + headroom
    pods_f = f"=CEILING(target_msgps*{mult}*burst_factor/(per_pod*util_target),1)+avail_headroom"
    nodes_f = f"=CEILING((CEILING(target_msgps*{mult}*burst_factor/(per_pod*util_target),1)+avail_headroom)/{pods_per_node},1)"
    rt_f = f"=(CEILING(target_msgps*{mult}*burst_factor/(per_pod*util_target),1)+avail_headroom)*({rt_cpu_lim}*730*{cpu_hr}+{rt_mem_lim}*730*{mem_hr})"
    mongo_s = f"=target_msgps*{mult}*{mongo_writes}*86400*{retention}*{doc_kb}/(1024*1024)*{mongo_replicas}*1.3*{mongo_storage_rate}"
    llm_f = f"=(target_msgps*{mult}*86400*30*{tokens_in}/1000000)*{llm_input_rate}+(target_msgps*{mult}*86400*30*{tokens_out}/1000000)*{llm_output_rate}" if tokens_in > 0 else "0"

    put(ws, r, 1, f"=target_msgps*{mult}", fill=OUTPUT_FILL, fmt="#,##0")
    put(ws, r, 2, pods_f, fill=OUTPUT_FILL, fmt="#,##0")
    put(ws, r, 3, nodes_f, fill=OUTPUT_FILL, fmt="#,##0")
    put(ws, r, 4, rt_f, fill=OUTPUT_FILL, fmt=MONEY_FMT)
    put(ws, r, 5, mongo_s, fill=OUTPUT_FILL, fmt=MONEY_FMT)
    put(ws, r, 6, llm_f, fill=OUTPUT_FILL, fmt=MONEY_FMT)
    put(ws, r, 7, f"=D{r}+{mongo_compute}+E{r}+{redis_compute}+F{r}", fill=OUTPUT_FILL, fmt=MONEY_FMT)
    r += 1

r += 1

# --- ASSUMPTIONS & NOTES ---
r = section(ws, r, "ASSUMPTIONS & NOTES")
assumptions = [
    "1. Per-pod capacity (13.5 msg/s) is measured under mock LLM with 1s simulated latency.",
    "   With real LLM, throughput may differ based on token count and model latency.",
    f"2. Node packing: {pods_per_node} runtime pods per {rt_instance} node ({node_cpu} vCPU).",
    f"   Each pod has {rt_cpu_lim}c CPU limit — cannot overpack without CPU contention.",
    f"3. Burst factor accounts for peak-to-average traffic ratio (default 3×).",
    "   Size infrastructure for peak, not average, to avoid latency degradation.",
    f"4. MongoDB and Redis compute costs are FIXED baselines (current cluster config).",
    "   Only MongoDB storage scales with throughput (more messages = more documents).",
    f"5. Storage formula: msg/s × {mongo_writes} writes × 86400 sec × {retention} days × {doc_kb}KB × {mongo_replicas} replicas × 1.3 (index overhead).",
    f"6. Cost rates: Azure pay-as-you-go (Central US). CPU: ${cpu_hr}/core/hr, Memory: ${mem_hr}/GB/hr.",
    "   Reserved instances (1yr/3yr) reduce compute costs 30-60%.",
    f"7. Primary bottleneck: {bottleneck}. This limits per-pod throughput before CPU saturates.",
    "8. Scaling is linear up to tested pod count. Beyond that, shared datastores (Mongo, Redis)",
    "   may become the bottleneck — validate with additional saturation tests.",
    f"9. Database nodes are fixed ({mongo_replicas}× MongoDB + Redis on dedicated pool).",
    "   GPU and system nodes do not scale with message throughput.",
    "10. All sizing assumes single-tenant deployment. Multi-tenant adds scheduling overhead.",
]
for line in assumptions:
    put(ws, r, 1, line, font=Font(size=10))
    r += 1

r += 1

# Column widths
for col, w in enumerate([35, 25, 30, 25, 18, 18, 18], 1):
    ws.column_dimensions[get_column_letter(col)].width = w

# Protect sheet — only green INPUT cells are editable
for c in input_cells:
    c.protection = UNLOCKED
ws.protection.sheet = True
ws.protection.password = ""
ws.protection.enable()

# ============================================================================
# TAB 2: Benchmarks
# ============================================================================
ws2 = wb.create_sheet("Benchmarks")
ws2.sheet_properties.tabColor = "70AD47"

r = 1
put(ws2, r, 1, "Benchmark Results & Test Evidence", font=Font(bold=True, size=14)); r += 2

# --- INFRASTRUCTURE UNDER TEST ---
r = section(ws2, r, "INFRASTRUCTURE UNDER TEST")
write_header(ws2, r, ["Component", "Configuration", "Notes"]); r += 1

# ConfigKey from measurement (what was actually tested)
ck = meas.get("configKey", {})
infra_test_data = [
    ("Cluster", infra.get("cluster", "?"), infra.get("namespace", "?")),
    ("Runtime pods", f"{ck.get('pods', max_pods)} pods pinned", f"Node: {ck.get('nodeType', rt_instance)}"),
    ("Runtime CPU", f"{ck.get('cpuRequestCores', rt_cpu_req)}c request / {ck.get('cpuLimitCores', rt_cpu_lim)}c limit", ""),
    ("Runtime Memory", f"{ck.get('memoryRequestGi', rt_mem_req)}Gi request / {ck.get('memoryLimitGi', rt_mem_lim)}Gi limit", ""),
    ("MongoDB", f"{mongo_replicas}× replicas, {mg.get('cpuRequestCores', 0.5)}c req", f"Pool size: {ck.get('mongoPoolSize', 10)}"),
    ("MongoDB disk", f"{mongo_pvc_gi}Gi × {mongo_replicas} ({mongo_disk})", f"IOPS limit: {mongo_iops_lim}"),
    ("Redis", f"master {redis_master_cpu}c/{redis_master_mem}Gi + {redis_reps}× replica {redis_rep_cpu}c", ""),
    ("Disk SKU", ck.get("diskSku", mongo_disk), ""),
]
for label, val, note in infra_test_data:
    put(ws2, r, 1, label, font=BOLD)
    put(ws2, r, 2, val)
    put(ws2, r, 3, note, font=Font(italic=True))
    r += 1

# Node pools used
if nodes:
    r += 1
    put(ws2, r, 1, "Node Pools:", font=Font(bold=True, size=11)); r += 1
    for n in nodes:
        put(ws2, r, 1, f"  {n.get('pool', '?')}", font=BOLD)
        put(ws2, r, 2, f"{n.get('count', '?')}× {n.get('instanceType', '?')} ({n.get('cpuPerNode', '?')} vCPU / {n.get('memoryPerNodeGi', '?')} GiB)")
        r += 1

r += 1

# --- TEST CONFIGURATION ---
r = section(ws2, r, "TEST SCENARIO")
write_header(ws2, r, ["Parameter", "Value", "Details"]); r += 1

config_data = [
    ("Scenario", args.scenario, scenario.get("description", "")),
    ("Endpoint", scenario["requestShape"]["endpoint"], scenario["requestShape"]["transport"]),
    ("Turn structure", f"1 create + {turn.get('followups', 4)} followups", f"{msgs_per_session} msgs/session"),
    ("Inter-message delay", f"{scenario['requestShape'].get('interMessageDelayMs', 1000)} ms", "Think time between msgs"),
    ("LLM mode", f"{llm_mode} ({llm_model})", f"Delay: {scenario['llmProfile'].get('delayMs', 'N/A')}ms" if llm_mode == "mock" else f"Tokens: {tokens_in} in / {tokens_out} out"),
    ("Payload size", f"{scenario['requestShape']['payloadBytesIn']} B in / {scenario['requestShape']['payloadBytesOut']} B out", ""),
    ("Session TTL", f"{session_ttl} min", f"Max in-memory: {scenario['sessionPattern']['maxInMemorySessionsPerPod']:,}/pod"),
    ("Tenant tier", scenario["tenantProfile"]["tier"], f"Rate limit: {scenario['tenantProfile']['rateLimitPerMin']:,}/min"),
    ("Datastore cost/msg", f"{mongo_writes} Mongo writes + {redis_ops} Redis ops", f"{scenario['perMessageDatastoreCost']['mongoReads']} Mongo reads"),
    ("k6 load zone", k6s.get("loadZone", "?"), "Grafana Cloud k6"),
]
for label, val, detail in config_data:
    put(ws2, r, 1, label, font=BOLD)
    put(ws2, r, 2, val)
    put(ws2, r, 3, detail, font=Font(italic=True))
    r += 1

r += 1

# --- k6 RESULTS ---
r = section(ws2, r, "k6 CLOUD RESULTS")
write_header(ws2, r, ["Metric", "Value", "Notes"]); r += 1

k6_data = [
    ("Run ID", meas["k6RunId"], f"Date: {meas['capturedAt'][:10]}"),
    ("Duration", f"{k6s.get('durationSeconds', '?')} sec", f"({k6s.get('durationSeconds', 0)//60} min {k6s.get('durationSeconds', 0)%60}s)"),
    ("Total requests", f"{k6s.get('totalRequests', '?'):,}" if isinstance(k6s.get('totalRequests'), int) else k6s.get('totalRequests', '?'), ""),
    ("Avg RPS (fleet)", k6s.get("avgRps", "?"), f"= {k6s.get('avgRps', 0)/max_pods:.1f} per pod" if max_pods > 0 else ""),
    ("Peak RPS (fleet)", k6s.get("peakRps", "?"), f"= {k6s.get('peakRps', 0)/max_pods:.1f} per pod" if max_pods > 0 else ""),
    ("Overall P95 latency", f"{k6s.get('overallP95Ms', '?')} ms", f"Target: < {p95_at_max} ms"),
    ("HTTP failures", k6s.get("httpFailures", 0), "Zero = clean run"),
    ("Max VUs reached", k6s.get("maxVUs", "?"), f"= {k6s.get('maxVUs', 0)/max_pods:.0f} per pod" if max_pods > 0 else ""),
]
for label, val, note in k6_data:
    put(ws2, r, 1, label, font=BOLD)
    put(ws2, r, 2, val, fill=GRAY_FILL)
    put(ws2, r, 3, note, font=Font(italic=True))
    r += 1

r += 1

# --- SATURATION FINDING ---
r = section(ws2, r, "SATURATION FINDING")
write_header(ws2, r, ["Metric", "Value", "Evidence"]); r += 1

sat_data = [
    ("Per-pod safe capacity", f"{per_pod} msg/s", f"p95 = {p95_at_max} ms at this rate"),
    ("Safe VUs per pod", meas["perPodCapacity"].get("maxSafeVUs", "?"), f"Total safe VUs: {meas['perPodCapacity'].get('maxSafeVUs', 0) * max_pods}"),
    ("Efficiency at max", f"{meas['perPodCapacity'].get('efficiencyAtMaxPct', '?')}%", ""),
    ("Primary bottleneck", bottleneck, ""),
    ("Bottleneck evidence", "", meas["perPodCapacity"].get("bottleneckEvidence", "")),
    ("k6 dashboard evidence", "", meas["perPodCapacity"].get("k6DashboardEvidence", "")),
]
for label, val, evidence in sat_data:
    put(ws2, r, 1, label, font=BOLD)
    put(ws2, r, 2, val, fill=GRAY_FILL)
    put(ws2, r, 3, evidence, font=Font(italic=True))
    r += 1

r += 1

# --- RUNTIME HEALTH (Coroot) ---
r = section(ws2, r, "RUNTIME HEALTH (from Coroot)")
write_header(ws2, r, ["Metric", "Avg", "Peak", "Threshold", "Status"]); r += 1

el_avg = coroot.get("eventLoopLagAvgMs", 0)
el_peak = coroot.get("eventLoopLagPeakMs", 0)
cpu_avg = coroot.get("cpuUsageAvgMilli", 0)
cpu_peak = coroot.get("cpuUsagePeakCores", 0)
throttle_avg = coroot.get("cpuThrottleAvgMs", 0)
throttle_peak = coroot.get("cpuThrottlePeakMs", 0)
delay_avg = coroot.get("cpuDelayAvgMs", 0)
delay_peak = coroot.get("cpuDelayPeakMs", 0)

health_data = [
    ("Event loop lag", f"{el_avg} ms", f"{el_peak} ms", "< 100ms avg", "RED" if el_avg > 100 else "YELLOW" if el_avg > 50 else "GREEN"),
    ("CPU usage", f"{cpu_avg}m", f"{cpu_peak} cores", f"< {rt_cpu_lim} cores", "RED" if cpu_peak > rt_cpu_lim * 0.9 else "GREEN"),
    ("CPU throttle", f"{throttle_avg} ms/s", f"{throttle_peak} ms/s", "< 50 ms/s avg", "RED" if throttle_avg > 50 else "YELLOW" if throttle_avg > 20 else "GREEN"),
    ("CPU delay", f"{delay_avg} ms", f"{delay_peak} ms", "< 50 ms avg", "RED" if delay_avg > 50 else "YELLOW" if delay_avg > 20 else "GREEN"),
]
for label, avg, peak, threshold, status in health_data:
    put(ws2, r, 1, label, font=BOLD)
    put(ws2, r, 2, avg)
    put(ws2, r, 3, peak)
    put(ws2, r, 4, threshold)
    status_fill = INPUT_FILL if status == "GREEN" else (WARN_FILL if status == "RED" else PatternFill("solid", fgColor="FFEB9C"))
    put(ws2, r, 5, status, fill=status_fill, font=BOLD)
    r += 1

r += 1

# --- DATASTORE HEALTH ---
r = section(ws2, r, "DATASTORE HEALTH (at max safe load)")
write_header(ws2, r, ["Metric", "Observed", "Limit / Capacity", "Utilization", "Status"]); r += 1

ds = meas.get("datastoreObserved", {})
mongo_cpu_obs = ds.get("mongoCpuMilli", 0) or 0
mongo_cpu_pct = ds.get("mongoCpuPctOfLimit", 0) or 0
redis_cpu_obs = ds.get("redisCpuMilli", 0) or 0
redis_cpu_pct = ds.get("redisCpuPctOfLimit", 0) or 0

ds_data = [
    ("MongoDB CPU", f"{mongo_cpu_obs}m", f"{mongo_cpu_lim * 1000 * mongo_replicas}m ({mongo_replicas}× {mongo_cpu_lim}c)", f"{mongo_cpu_pct}%",
     "RED" if mongo_cpu_pct > 80 else "YELLOW" if mongo_cpu_pct > 60 else "GREEN"),
    ("MongoDB write latency", f"{ds.get('mongoWriteLatencyMs', 'N/A')} ms", "< 10ms", "", "GREEN"),
    ("MongoDB IOPS", f"{ds.get('mongoIopsWrite', 'N/A')}", f"{mongo_iops_lim} (disk limit)", "",
     "GREEN"),
    ("Redis CPU", f"{redis_cpu_obs}m", f"{redis_total_cpu * 1000}m ({redis_total_cpu}c total)", f"{redis_cpu_pct}%",
     "RED" if redis_cpu_pct > 80 else "YELLOW" if redis_cpu_pct > 60 else "GREEN"),
]
for label, obs, limit, util, status in ds_data:
    put(ws2, r, 1, label, font=BOLD)
    put(ws2, r, 2, obs)
    put(ws2, r, 3, limit)
    put(ws2, r, 4, util)
    status_fill = INPUT_FILL if status == "GREEN" else (WARN_FILL if status == "RED" else PatternFill("solid", fgColor="FFEB9C"))
    put(ws2, r, 5, status, fill=status_fill, font=BOLD)
    r += 1

r += 1

# --- PER-STEP SCORECARD ---
steps = meas.get("perStepSummary", [])
if steps:
    r = section(ws2, r, "PER-STEP SCORECARD")
    write_header(ws2, r, ["Step", "VUs", "VUs/pod", "Avg msg/s", "Peak msg/s", "Avg p95 (ms)", "Peak p95 (ms)", "Avg CPU(m)", "Peak CPU(m)", "Efficiency", "Status"]); r += 1
    for s in steps:
        put(ws2, r, 1, s["step"])
        put(ws2, r, 2, s["vus"])
        put(ws2, r, 3, s["vus"] / max_pods if max_pods > 0 else "?", fmt=NUM_FMT)
        put(ws2, r, 4, s["avgMsgPerSec"], fmt=NUM_FMT)
        put(ws2, r, 5, s["peakMsgPerSec"], fmt=NUM_FMT)
        put(ws2, r, 6, s["avgP95Ms"], fmt="#,##0")
        put(ws2, r, 7, s["peakP95Ms"], fmt="#,##0")
        put(ws2, r, 8, s["avgCpuMilli"], fmt="#,##0")
        put(ws2, r, 9, s["peakCpuMilli"], fmt="#,##0")
        put(ws2, r, 10, s.get("efficiency", ""), font=Font(italic=True))
        status_fill = INPUT_FILL if s["status"] == "GREEN" else (WARN_FILL if s["status"] == "RED" else None)
        put(ws2, r, 11, s["status"], fill=status_fill, font=BOLD)
        r += 1

r += 1

# --- SCALING LINEARITY ---
scaling = meas.get("multiPodScaling", [])
if scaling:
    r = section(ws2, r, "SCALING LINEARITY")
    write_header(ws2, r, ["Pods", "Total msg/s", "Per-pod msg/s", "Scaling Factor", "Shared Bottleneck"]); r += 1
    for s in scaling:
        put(ws2, r, 1, s["pods"])
        put(ws2, r, 2, s["totalMsgPerSec"], fmt=NUM_FMT)
        put(ws2, r, 3, s["totalMsgPerSec"] / s["pods"] if s["pods"] > 0 else 0, fmt=NUM_FMT)
        put(ws2, r, 4, s["scalingFactor"], fmt="0.00")
        put(ws2, r, 5, s.get("sharedBottleneck") or "none")
        r += 1

r += 1

# --- WORKLOAD INTRINSICS ---
r = section(ws2, r, "WORKLOAD INTRINSICS (derived from test)")
write_header(ws2, r, ["Metric", "Value", "Unit", "Derivation"]); r += 1

intrinsic_data = [
    ("Mongo CPU per msg", mongo_cpu_per_msg, "milli", wi.get("source", "")),
    ("Redis CPU per msg", redis_cpu_per_msg, "milli", wi.get("source", "")),
    ("Mongo IOPS per msg", mongo_iops_per_msg, "IOPS", "Direct from scenario config"),
    ("Mongo writes per msg", mongo_writes, "ops", "From scenario perMessageDatastoreCost"),
    ("Redis ops per msg", redis_ops, "ops", "From scenario perMessageDatastoreCost"),
]
for label, val, unit, source in intrinsic_data:
    put(ws2, r, 1, label, font=BOLD)
    put(ws2, r, 2, val, fmt=NUM_FMT)
    put(ws2, r, 3, unit)
    put(ws2, r, 4, source, font=Font(italic=True))
    r += 1

r += 1

# --- VALIDATION & PROVENANCE ---
r = section(ws2, r, "VALIDATION & PROVENANCE")
write_header(ws2, r, ["Field", "Value", "Notes"]); r += 1

prov = [
    ("Validation status", meas.get("validationStatus", "?"), ""),
    ("Validation notes", "", meas.get("validationNotes", "")),
    ("Infra snapshot", infra.get("capturedAt", "?"), "live from kubectl" if args.env else "static file"),
    ("Rate card", rate.get("capturedAt", "?"), f"Region: {rate.get('region', '?')}"),
    ("Scenario version", scenario.get("version", "?"), f"File: {args.scenario_file}"),
    ("Cluster", infra.get("cluster", "?"), infra.get("namespace", "?")),
    ("Measurement file", args.measurement, ""),
    ("Saturation report", meas.get("saturationReportPath", "N/A"), "Full markdown report"),
]
for label, val, note in prov:
    put(ws2, r, 1, label, font=BOLD)
    put(ws2, r, 2, val)
    put(ws2, r, 3, note, font=Font(italic=True))
    r += 1

# Column widths
for col, w in enumerate([28, 25, 30, 22, 15], 1):
    ws2.column_dimensions[get_column_letter(col)].width = w

# ============================================================================
# Save
# ============================================================================
# Force Excel to recalculate on open
wb.calculation.calcMode = "auto"

out_path = Path(args.out)
out_path.parent.mkdir(parents=True, exist_ok=True)
wb.save(out_path)

# Note: forceFullCalc post-processing removed — Numbers on macOS chokes on
# rewritten zip. openpyxl's calcMode="auto" is sufficient for Excel/Numbers.

import os
size = os.path.getsize(out_path)
print(f"\nwrote {out_path} ({size/1024:.1f} KB)")
print(f"Tabs: {[s.title for s in wb.worksheets]}")
print(f"\n  Tab 1: Calculator — edit GREEN cells, includes infra + VM types + cost formulas")
print(f"  Tab 2: Benchmarks — saturation test evidence, per-step scorecard, provenance")
