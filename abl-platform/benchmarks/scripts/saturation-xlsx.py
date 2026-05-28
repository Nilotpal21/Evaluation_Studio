#!/usr/bin/env python3
"""
saturation-xlsx.py — Multi-service sizing & cost XLSX.

Generates the reviewed sizing workbook from a built-in service catalog.
No benchmark result files, reports, kubectl access, or cluster context are required.

Usage:
    python3 benchmarks/scripts/saturation-xlsx.py \
        --output benchmarks/docs/sizing.xlsx

Requires: openpyxl
"""

import argparse, re, sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Border, Font, PatternFill, Side, Alignment, Protection
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    print("openpyxl required — pip install openpyxl"); sys.exit(1)

# ── Styles ────────────────────────────────────────────────────────────────────
B = Font(bold=True)
B14 = Font(bold=True, size=14)
B12 = Font(bold=True, size=12)
B11 = Font(bold=True, size=11)
IT = Font(italic=True, color="666666")
HF = Font(bold=True, color="FFFFFF", size=10)
SF = Font(bold=True, color="1F2937", size=10)
HB = PatternFill("solid", fgColor="2F5496")
YE = PatternFill("solid", fgColor="FFF2CC")  # editable
CSPF = PatternFill("solid", fgColor="D9EAF7")  # editable cloud provider selector
GR = PatternFill("solid", fgColor="F2F2F2")  # readonly
GN = PatternFill("solid", fgColor="C6EFCE")
AM = PatternFill("solid", fgColor="FFEB9C")
RD = PatternFill("solid", fgColor="FFC7CE")
T1 = PatternFill("solid", fgColor="D6E4F0")  # tier 1 - runtime blue
T2 = PatternFill("solid", fgColor="E2EFDA")  # tier 2 - search green
T3 = PatternFill("solid", fgColor="FCE4D6")  # tier 3 - workflow orange
T4 = PatternFill("solid", fgColor="E4DFEC")  # tier 4 - data purple
T5 = PatternFill("solid", fgColor="EDEDED")  # tier 5 - supporting gray
T6 = PatternFill("solid", fgColor="F8F8F8")  # tier 6 - operators
BD = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
SOFT_BD = Border(
    left=Side("thin", color="D9E2F3"),
    right=Side("thin", color="D9E2F3"),
    top=Side("thin", color="D9E2F3"),
    bottom=Side("thin", color="D9E2F3"),
)
AL_R = Alignment(horizontal="right", vertical="center")
AL_C = Alignment(horizontal="center", vertical="center")
AL_L = Alignment(horizontal="left", vertical="center")
AL_W = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Current Central US retail rates from Azure Retail Prices API, checked 2026-05-07.
HOURS_PER_MONTH = 730
D2S_V3_MONTHLY = round(0.110 * HOURS_PER_MONTH)
D4S_V5_MONTHLY = round(0.217 * HOURS_PER_MONTH)
D8S_V5_MONTHLY = round(0.434 * HOURS_PER_MONTH)
D16S_V5_MONTHLY = round(0.868 * HOURS_PER_MONTH)
D32S_V5_MONTHLY = round(1.736 * HOURS_PER_MONTH)
D48S_V5_MONTHLY = round(2.604 * HOURS_PER_MONTH)
E8S_V5_MONTHLY = round(0.566 * HOURS_PER_MONTH)
E16S_V5_MONTHLY = round(1.132 * HOURS_PER_MONTH)
E32S_V5_MONTHLY = round(2.278 * HOURS_PER_MONTH)
E48S_V5_MONTHLY = round(3.417 * HOURS_PER_MONTH)
NC4AS_T4_V3_MONTHLY = round(0.594 * HOURS_PER_MONTH)
WAF_V2_FIXED_1CU_MONTHLY = round((0.443 + 0.0144) * HOURS_PER_MONTH)
ACR_PREMIUM_MONTHLY = round(1.6666 * 30)
P10_ZRS_SUPPLEMENTAL_DISK_MONTHLY = round(29.565 + 1.119)
BLOB_HOT_LRS_GB_MONTH = 0.0184
AWS_M7I_XLARGE_MONTHLY = round(0.2016 * HOURS_PER_MONTH)
AWS_M7I_2XLARGE_MONTHLY = round(0.4032 * HOURS_PER_MONTH)
AWS_M7I_4XLARGE_MONTHLY = round(0.8064 * HOURS_PER_MONTH)
AWS_M7I_12XLARGE_MONTHLY = round(2.4192 * HOURS_PER_MONTH)
AWS_R7I_2XLARGE_MONTHLY = round(0.5292 * HOURS_PER_MONTH)
AWS_R7I_4XLARGE_MONTHLY = round(1.0584 * HOURS_PER_MONTH)
AWS_R7I_8XLARGE_MONTHLY = round(2.1168 * HOURS_PER_MONTH)
AWS_R7I_12XLARGE_MONTHLY = round(3.1752 * HOURS_PER_MONTH)
AWS_G4DN_XLARGE_MONTHLY = round(0.526 * HOURS_PER_MONTH)
AWS_ALB_WAF_BASELINE_MONTHLY = round((0.0225 + 0.008) * HOURS_PER_MONTH + 30)
AWS_GP3_128GB_MONTHLY = round(128 * 0.08)
AWS_ECR_BASELINE_MONTHLY = 10
S3_STANDARD_GB_MONTH = 0.023
# GCP us-central1 on-demand pricing (checked 2026-05-14)
GCP_N2_STD_4_MONTHLY = round(0.1943 * HOURS_PER_MONTH)    # n2-standard-4: 4 vCPU / 16 GiB
GCP_N2_STD_8_MONTHLY = round(0.3886 * HOURS_PER_MONTH)    # n2-standard-8: 8 vCPU / 32 GiB
GCP_N2_STD_16_MONTHLY = round(0.7772 * HOURS_PER_MONTH)   # n2-standard-16: 16 vCPU / 64 GiB
GCP_N2_STD_32_MONTHLY = round(1.5544 * HOURS_PER_MONTH)   # n2-standard-32: 32 vCPU / 128 GiB
GCP_N2_STD_48_MONTHLY = round(2.3316 * HOURS_PER_MONTH)   # n2-standard-48: 48 vCPU / 192 GiB
GCP_N2_HMEM_8_MONTHLY = round(0.5220 * HOURS_PER_MONTH)   # n2-highmem-8: 8 vCPU / 64 GiB
GCP_N2_HMEM_16_MONTHLY = round(1.0440 * HOURS_PER_MONTH)  # n2-highmem-16: 16 vCPU / 128 GiB
GCP_N2_HMEM_32_MONTHLY = round(2.0880 * HOURS_PER_MONTH)  # n2-highmem-32: 32 vCPU / 256 GiB
GCP_N2_HMEM_48_MONTHLY = round(3.1320 * HOURS_PER_MONTH)  # n2-highmem-48: 48 vCPU / 384 GiB
GCP_N1_STD_4_T4_MONTHLY = round((0.1900 + 0.35) * HOURS_PER_MONTH)  # n1-standard-4 + 1×T4 GPU
GCP_LB_WAF_BASELINE_MONTHLY = round(0.025 * HOURS_PER_MONTH + 25)  # Cloud Load Balancer + Cloud Armor
GCP_PD_SSD_128GB_MONTHLY = round(128 * 0.17)  # SSD persistent disk 128 GiB
GCP_AR_BASELINE_MONTHLY = 10  # Artifact Registry baseline
GCS_STANDARD_GB_MONTH = 0.020  # Cloud Storage standard class

# VM specifications: {azure_name: (vCPUs, memory_GiB, $/mo)}
# Used for dynamic node pool computation based on selected VM size.
AZURE_VM_SPECS = {
    "Standard_D4s_v5":  (4,  16,  D4S_V5_MONTHLY),
    "Standard_D8s_v5":  (8,  32,  D8S_V5_MONTHLY),
    "Standard_D16s_v5": (16, 64,  D16S_V5_MONTHLY),
    "Standard_D32s_v5": (32, 128, D32S_V5_MONTHLY),
    "Standard_D48s_v5": (48, 192, D48S_V5_MONTHLY),
    "Standard_E8s_v5":  (8,  64,  E8S_V5_MONTHLY),
    "Standard_E16s_v5": (16, 128, E16S_V5_MONTHLY),
    "Standard_E32s_v5": (32, 256, E32S_V5_MONTHLY),
    "Standard_E48s_v5": (48, 384, E48S_V5_MONTHLY),
    "Standard_NC4as_T4_v3": (4, 28, NC4AS_T4_V3_MONTHLY),
}
AWS_VM_SPECS = {
    "m7i.xlarge":    (4,  16,  AWS_M7I_XLARGE_MONTHLY),
    "m7i.2xlarge":   (8,  32,  AWS_M7I_2XLARGE_MONTHLY),
    "m7i.4xlarge":   (16, 64,  AWS_M7I_4XLARGE_MONTHLY),
    "m7i.12xlarge":  (48, 192, AWS_M7I_12XLARGE_MONTHLY),
    "r7i.2xlarge":   (8,  64,  AWS_R7I_2XLARGE_MONTHLY),
    "r7i.4xlarge":   (16, 128, AWS_R7I_4XLARGE_MONTHLY),
    "r7i.8xlarge":   (32, 256, AWS_R7I_8XLARGE_MONTHLY),
    "r7i.12xlarge":  (48, 384, AWS_R7I_12XLARGE_MONTHLY),
    "g4dn.xlarge":   (4,  16,  AWS_G4DN_XLARGE_MONTHLY),
}
GCP_VM_SPECS = {
    "n2-standard-4":  (4,  16,  GCP_N2_STD_4_MONTHLY),
    "n2-standard-8":  (8,  32,  GCP_N2_STD_8_MONTHLY),
    "n2-standard-16": (16, 64,  GCP_N2_STD_16_MONTHLY),
    "n2-standard-32": (32, 128, GCP_N2_STD_32_MONTHLY),
    "n2-standard-48": (48, 192, GCP_N2_STD_48_MONTHLY),
    "n2-highmem-8":   (8,  64,  GCP_N2_HMEM_8_MONTHLY),
    "n2-highmem-16":  (16, 128, GCP_N2_HMEM_16_MONTHLY),
    "n2-highmem-32":  (32, 256, GCP_N2_HMEM_32_MONTHLY),
    "n2-highmem-48":  (48, 384, GCP_N2_HMEM_48_MONTHLY),
    "n1-standard-4+T4": (4, 15, GCP_N1_STD_4_T4_MONTHLY),
}
# Pool → list of Azure VM options for dropdown (first = default)
POOL_VM_OPTIONS = {
    "user":              ["Standard_D8s_v5", "Standard_D16s_v5", "Standard_D32s_v5"],
    "database (mongo)":  ["Standard_E16s_v5", "Standard_E32s_v5"],
    "database (non-mongo)": ["Standard_E8s_v5", "Standard_E16s_v5", "Standard_D48s_v5"],
    "operator":          ["Standard_D4s_v5", "Standard_D8s_v5"],
    "system":            ["Standard_D8s_v5", "Standard_D16s_v5"],
    "gpu":               ["Standard_NC4as_T4_v3"],
}
# Pool → list of AWS VM options for dropdown (first = default)
POOL_VM_OPTIONS_AWS = {
    "user":              ["m7i.2xlarge", "m7i.4xlarge", "m7i.12xlarge"],
    "database (mongo)":  ["r7i.4xlarge", "r7i.8xlarge"],
    "database (non-mongo)": ["r7i.2xlarge", "r7i.4xlarge", "r7i.12xlarge"],
    "operator":          ["m7i.xlarge", "m7i.2xlarge"],
    "system":            ["m7i.2xlarge", "m7i.4xlarge"],
    "gpu":               ["g4dn.xlarge"],
}
# Pool → list of GCP VM options for dropdown (first = default)
POOL_VM_OPTIONS_GCP = {
    "user":              ["n2-standard-8", "n2-standard-16", "n2-standard-32"],
    "database (mongo)":  ["n2-highmem-16", "n2-highmem-32"],
    "database (non-mongo)": ["n2-highmem-8", "n2-highmem-16", "n2-highmem-48"],
    "operator":          ["n2-standard-4", "n2-standard-8"],
    "system":            ["n2-standard-8", "n2-standard-16"],
    "gpu":               ["n1-standard-4+T4"],
}
PREMIUM_SSD_LRS_DISK_PRICE = {
    "P10": 19.71,
    "P15": 38.012142,
    "P20": 73.22,
    "P30": 135.17,
    "P40": 259.0457,
}

def p(ws, r, c, v, font=None, fill=None, fmt=None, al=None):
    cl = ws.cell(row=r, column=c, value=v)
    if font: cl.font = font
    if fill: cl.fill = fill
    if fmt: cl.number_format = fmt
    if al: cl.alignment = al
    cl.border = BD; return cl

def hdr(ws, r, cols):
    for i, t in enumerate(cols, 1):
        p(ws, r, i, t, font=HF, fill=HB)

def wid(ws, ww):
    for i, v in enumerate(ww, 1): ws.column_dimensions[get_column_letter(i)].width = v

def sec(ws, r, t, fill=T1, ncols=8):
    for c in range(1, ncols+1):
        p(ws, r, c, t if c == 1 else "", font=B12, fill=fill, al=AL_L if c == 1 else AL_C)

def csp_text(csp_ref, aws_text, azure_text, gcp_text=None):
    if gcp_text is None:
        gcp_text = azure_text  # fallback
    return f'=IF({csp_ref}="AWS","{aws_text}",IF({csp_ref}="GCP","{gcp_text}","{azure_text}"))'

def csp_value(csp_ref, aws_value, azure_value, gcp_value=None):
    if gcp_value is None:
        gcp_value = azure_value  # fallback
    return f'=IF({csp_ref}="AWS",{aws_value},IF({csp_ref}="GCP",{gcp_value},{azure_value}))'

# ── Catalog helpers ──────────────────────────────────────────────────────────
def _cpu(v):
    return f"{v:g}" if isinstance(v, (int, float)) else (v or "—")

def _mem(v):
    return f"{v:g}Gi" if isinstance(v, (int, float)) else (v or "—")

def _container(name, cpu_req, cpu_lim, mem_req, mem_lim, image="catalog"):
    return {
        "name": name,
        "image": image,
        "cpuReq": _cpu(cpu_req),
        "cpuLim": _cpu(cpu_lim),
        "memReq": _mem(mem_req),
        "memLim": _mem(mem_lim),
    }

def _catalog_service(name, replicas, cpu_req, cpu_lim, mem_req, mem_lim, pool="user", kind="Deployment", max_replicas=None):
    return {
        "name": name,
        "kind": kind,
        "replicas": replicas,
        "max_replicas": max_replicas or replicas,
        "pool": pool,
        "containers": [_container(name, cpu_req, cpu_lim, mem_req, mem_lim, "catalog")],
    }

def default_service_catalog():
    """Built-in sizing catalog.

    Values mirror tested production resource limits (validated May 6-8 2026 load tests).
    CPU/memory limits for SearchAI pipeline services updated from enterprise performance reports:
    - File ingestion: docs/load-testing/enterprise-report/file-ingestion-enterprise-report-2026-05-08.md
    - Query pipeline: docs/load-testing/enterprise-report/searchai-unified-performance-report.md
    """
    return [
        # ── Runtime tier (min, cpu_req, cpu_lim, mem_req_gi, mem_lim_gi, max_replicas) ──
        _catalog_service("admin", 2, 0.25, 1, 0.256, 1, max_replicas=4),
        _catalog_service("runtime", 2, 1, 4, 2, 4, max_replicas=45),
        _catalog_service("studio", 2, 1, 2, 1, 2, max_replicas=12),
        # ── Search tier ──
        _catalog_service("search-ai", 2, 2, 6, 4, 8, max_replicas=8),
        _catalog_service("search-ai-runtime", 2, 1, 2, 2, 4, max_replicas=8),
        # ── Workflow tier ──
        _catalog_service("restate-workflows", 2, 0.1, 0.5, 0.5, 1),
        _catalog_service("workflow-engine", 2, 0.5, 2, 1, 4, max_replicas=5),
        # ── Pipeline tier ──
        _catalog_service("pipeline-engine", 2, 1, 2, 1, 4),
        _catalog_service("restate", 2, 1, 4, 2, 6),
        _catalog_service("restate-op", 2, 1, 4, 2, 6, kind="RestateCluster"),
        # ── Data stores (StatefulSet — fixed topology, no HPA) ──
        _catalog_service("clickhouse-keeper", 3, 0.25, 0.5, 0.5, 1, "system", "StatefulSet"),
        _catalog_service("clickhouse-shard-0", 6, 2, 4, 4, 8, "system", "StatefulSet"),
        _catalog_service("kafka-default", 3, 0.5, 2, 2, 4, "user", "KafkaNodePool"),
        _catalog_service("mongodb", 3, 8, 16, 64, 128, "database", "StatefulSet"),
        _catalog_service("mongodb-arb", 3, 0.1, 0.2, 0.256, 0.5, "database", "StatefulSet"),
        _catalog_service("redis-master", 3, 4, 8, 8, 16, "database", "StatefulSet"),
        _catalog_service("redis-replicas", 3, 2, 4, 8, 16, "database", "StatefulSet"),
        _catalog_service("rocksdb", 3, 0.5, 1, 2, 4, "database", "StatefulSet"),
        _catalog_service("neo4j", 3, 0.5, 2, 2, 4, "database", "StatefulSet"),
        # OpenSearch: 4 CPU / 8 lim, 32Gi / 64Gi (JVM 16g), 3 data nodes, 100Gi/node.
        _catalog_service("opensearch", 3, 4, 8, 32, 64, "database", "StatefulSet"),
        # ── Supporting / auxiliary ──
        _catalog_service("bge-m3", 2, 4, 8, 12, 24, "gpu", max_replicas=4),
        _catalog_service("bge-m3-cpu", 2, 2, 4, 4, 8, "user", max_replicas=4),
        _catalog_service("codetool-sandbox", 2, 0.25, 2, 0.5, 2),
        _catalog_service("crawler-go-worker", 2, 0.5, 2, 0.512, 2, max_replicas=5),
        _catalog_service("crawler-mcp-server", 2, 1, 4, 2, 4, max_replicas=4),
        _catalog_service("docling", 2, 4, 8, 12, 24, "gpu", max_replicas=4),
        _catalog_service("docling-cpu", 2, 0.5, 2, 1.5, 4, "user", max_replicas=4),
        _catalog_service("ingress-nginx-controller", 3, 0.5, 2, 0.5, 1),
        _catalog_service("livekit", 2, 0.1, 0.5, 0.125, 0.5),
        _catalog_service("minio", 2, 0.1, 0.5, 0.25, 0.5),
        _catalog_service("multimodal-service", 2, 0.5, 2, 1, 2, max_replicas=5),
        _catalog_service("preprocessing", 2, 0.5, 2, 2, 4, max_replicas=6),
        _catalog_service("restate-operator", 2, 0.05, 0.2, 0.0976562, 0.25),
        _catalog_service("external-secrets", 2, "—", "—", "—", "—"),
        _catalog_service("external-secrets-cert-controller", 2, "—", "—", "—", "—"),
        _catalog_service("external-secrets-webhook", 2, "—", "—", "—", "—"),
        _catalog_service("goldilocks-controller", 2, 0.025, 0.1, 0.0625, 0.125),
        _catalog_service("goldilocks-dashboard", 2, 0.025, 0.1, 0.0625, 0.125),
        _catalog_service("kafka-entity-operator", 2, 0.1, 0.5, 0.25, 0.5),
        _catalog_service("mongodb-kubernetes-operator", 2, 0.5, 1.1, 0.195312, 1),
        _catalog_service("strimzi-cluster-operator", 2, 0.1, 0.5, 0.25, 0.5),
        _catalog_service("vpa-recommender", 2, 0.05, "—", 0.488281, "—"),
    ]

def apply_catalog_placement(services, svc_pool, pool_vm):
    default_pool_vm = {
        "database (mongo)": "Standard_E16s_v5",
        "database (non-mongo)": "Standard_E8s_v5",
        "gpu": "Standard_NC4as_T4_v3",
        "operator": "Standard_D4s_v5",
        "system": "Standard_D8s_v5",
        "user": "Standard_D16s_v5",
    }
    for pool, vm in default_pool_vm.items():
        pool_vm.setdefault(pool, vm)
    for svc in services:
        pool = svc.get("pool")
        if pool:
            # Remap generic "database" pool to specific sub-pools
            if pool == "database":
                if "mongodb" in svc["name"].lower() or "mongo" in svc["name"].lower():
                    pool = "database (mongo)"
                else:
                    pool = "database (non-mongo)"
            svc_pool.setdefault(svc["name"], pool)

def classify_service(name):
    """Classify a service into a tier."""
    n = name.lower()
    if n in ("runtime","studio","admin"): return "runtime"
    if "search-ai" in n or "serach-ai" in n: return "search"
    if n in ("workflow-engine","restate-workflows"): return "workflow"
    if n in ("pipeline-engine","restate","restate-op"): return "pipeline"
    if any(x in n for x in ("mongodb-kubernetes","strimzi","external-secret","goldilocks","vpa","kafka-entity")): return "operator"
    if any(x in n for x in ("mongodb","redis","clickhouse","kafka","opensearch","neo4j","rocksdb")): return "data"
    return "supporting"

def cpu_to_cores(s):
    """Parse '500m' or '4' to float cores."""
    s = str(s).strip()
    if s in ("—","","None"): return 0
    if s.endswith("m"): return float(s[:-1])/1000
    return float(s)

def mem_to_gb(s):
    """Parse '512Mi' or '2Gi' to float GB."""
    s = str(s).strip()
    if s in ("—","","None"): return 0
    m = re.match(r"([\d.]+)\s*(Gi|Mi|G|M|Ki|K|Ti)?", s)
    if not m: return 0
    v = float(m.group(1)); u = m.group(2) or ""
    if u in ("Gi","G"): return v
    if u in ("Mi","M"): return v/1024
    if u in ("Ti",): return v*1024
    if u in ("Ki","K"): return v/(1024*1024)
    return v  # assume GB if no unit

# ── CLI ───────────────────────────────────────────────────────────────────────
parser = argparse.ArgumentParser()
parser.add_argument("--csp", choices=["Azure", "AWS", "GCP"], default="Azure", help="Cloud provider for VM/SKU labels and pricing")
parser.add_argument("--target", type=int, default=120, help="Runtime target msg/s")
parser.add_argument("--search-target", type=int, default=20, help="Search queries/s")
parser.add_argument("--embed-ratio", type=float, default=0.40, help="Fraction of search queries requiring embedding (hybrid+semantic)")
parser.add_argument("--ingest-target", type=float, default=5, help="File ingestion target files/min")
parser.add_argument("--workflow-target", type=int, default=10, help="Workflow executions/s")
parser.add_argument("--pipeline-target", type=int, default=10, help="Pipeline executions/s")
parser.add_argument("--util", type=float, default=0.70)
parser.add_argument("--per-pod", type=float, default=12, help="Runtime per-pod msg/s")
parser.add_argument("--runtime-active-sessions", type=int, default=0, help="Runtime target active sessions; 0 disables active-session scaling")
parser.add_argument("--runtime-active-sessions-per-pod", type=float, default=0, help="Safe runtime active sessions per pod; 0 disables active-session scaling")
parser.add_argument("--studio-runtime-ratio", type=float, default=0.25, help="Minimum Studio pods as a ratio of runtime pods")
parser.add_argument("--dirty-pages", type=float, default=8.2)
parser.add_argument("--node-price-mo", type=float, default=D16S_V5_MONTHLY, help="App pool Standard_D16s_v5 $/mo")
parser.add_argument("--db-node-price-mo", type=float, default=D48S_V5_MONTHLY, help="DB-pool Standard_D48s_v5 $/mo")
parser.add_argument("--gpu-node-price-mo", type=float, default=NC4AS_T4_V3_MONTHLY, help="GPU-pool Standard_NC4as_T4_v3 $/mo")
parser.add_argument("--system-node-price-mo", type=float, default=D8S_V5_MONTHLY, help="System pool Standard_D8s_v5 $/mo")
parser.add_argument("--hardware-discount", type=float, default=0, help="Hardware discount applied to node-pool MSRP")
parser.add_argument("--support-rate", type=float, default=0.10, help="Monthly support rate applied to node-pool MSRP")
parser.add_argument("--object-storage-tb", type=float, default=20, help="External object storage size in TiB")
parser.add_argument("--object-storage-gb-mo", type=float, default=BLOB_HOT_LRS_GB_MONTH, help="Object storage $/GB-month")
parser.add_argument("--mongo-disk-gb", type=int, default=256, help="MongoDB PVC disk size per node in GiB")
parser.add_argument("--opensearch-disk-gb", type=int, default=100, help="OpenSearch PVC disk size per node in GiB")
parser.add_argument("--neo4j-disk-gb", type=int, default=64, help="Neo4j PVC disk size per node in GiB")
parser.add_argument("--redis-disk-gb", type=int, default=32, help="Redis PVC disk size per node in GiB")
parser.add_argument("--clickhouse-disk-gb", type=int, default=256, help="ClickHouse PVC disk size per node in GiB")
parser.add_argument("--kafka-disk-gb", type=int, default=128, help="Kafka PVC disk size per node in GiB")
parser.add_argument("--rocksdb-disk-gb", type=int, default=20, help="RocksDB PVC disk size per node in GiB")
parser.add_argument("--mongo-license-per-node-mo", type=float, default=1800, help="MongoDB license $/node-month")
parser.add_argument("--redis-license-per-node-mo", type=float, default=410, help="Redis license $/node-month")
parser.add_argument("--clickhouse-license-per-node-mo", type=float, default=0, help="ClickHouse license $/node-month if applicable")
parser.add_argument("--kafka-license-per-node-mo", type=float, default=0, help="Kafka license $/node-month if applicable")
parser.add_argument("--pods-per-node", type=int, default=3)
parser.add_argument("--ha-min-app-pods", type=int, default=2, help="Minimum app replicas for HA")
parser.add_argument("--studio-min-pods", type=int, default=2, help="Minimum Studio replicas")
parser.add_argument("--search-min-pods", type=int, default=3, help="Minimum replicas for Search AI services")
parser.add_argument("--workflow-engine-min-pods", type=int, default=2, help="Minimum Workflow Engine replicas")
parser.add_argument("--redis-shards", type=int, default=3, help="Redis shard/master count")
parser.add_argument("--redis-replicas-per-shard", type=int, default=1, help="Redis replicas per shard")
parser.add_argument("--clickhouse-shards", type=int, default=3, help="ClickHouse shard count")
parser.add_argument("--clickhouse-replicas-per-shard", type=int, default=2, help="ClickHouse replicas per shard")
parser.add_argument("--output", required=True)
args = parser.parse_args()

# ── Built-in catalog ──────────────────────────────────────────────────────────
all_services = default_service_catalog()
svc_pool = {}  # service name → node pool
pool_vm = {}   # pool name → VM type
apply_catalog_placement(all_services, svc_pool, pool_vm)
print(f"[catalog] Loaded {len(all_services)} built-in services")

# Group by tier
tiers = {"runtime":[],"search":[],"workflow":[],"pipeline":[],"data":[],"supporting":[],"operator":[]}
for svc in all_services:
    tier = classify_service(svc["name"])
    tiers[tier].append(svc)
for svcs in tiers.values():
    svcs.sort(key=lambda s: s["name"])

node_pools = {
    "database (mongo)": {"count": 0, "cpu": 16, "mem": "128GB"},
    "database (non-mongo)": {"count": 0, "cpu": 8, "mem": "64GB"},
    "gpu": {"count": 0, "cpu": 4, "mem": "27GB"},
    "operator": {"count": 0, "cpu": 4, "mem": "16GB"},
    "system": {"count": 0, "cpu": 8, "mem": "32GB"},
    "user": {"count": 0, "cpu": 16, "mem": "64GB"},
}

# Set default pool VMs
pool_vm["database (mongo)"] = "Standard_E16s_v5"
pool_vm["database (non-mongo)"] = "Standard_E8s_v5"
pool_vm["operator"] = "Standard_D4s_v5"
pool_vm["system"] = "Standard_D8s_v5"
pool_vm["user"] = "Standard_D16s_v5"

per_pod = args.per_pod

# ══════════════════════════════════════════════════════════════════════════════
wb = Workbook()

# ═══ TAB 1: SIZING & COST ════════════════════════════════════════════════════
ws = wb.active; ws.title = "Sizing & Cost"
COLS = ["Service","Min","Max (HPA)","CPU Req","CPU Lim","Mem Req Gi","Mem Lim Gi","Total CPU","Total Mem Gi","Pool","VM Type","Req Nodes","Throughput/pod","p50","p90","p95","p99"]
wid(ws, [30, 9, 10, 13, 11, 11, 10, 12, 12, 24, 11, 15, 10, 10, 10, 10])

r = 1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=16)
p(ws,r,1,"Infrastructure Sizing & Cost Calculator",font=B14, fill=PatternFill("solid", fgColor="EAF2F8"))
r+=1
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=16)
p(ws,r,1,"Edit yellow cells only; calculated and catalog cells are protected",font=IT, fill=PatternFill("solid", fgColor="F8FBFD")); r+=1

# ── Global Inputs ──
sec(ws,r,"GLOBAL INPUTS",fill=HB,ncols=16); r+=1
hdr(ws,r,["Parameter","Value","Unit","","","","","","","","","","","","","Notes"]); r+=1

R = {}

R["csp"] = r
p(ws,r,1,"Cloud provider",font=B); p(ws,r,2,args.csp,font=B,fill=CSPF); p(ws,r,3,"CSP")
p(ws,r,10,"Switches VM/SKU labels and provider pricing",font=IT); r+=1
CSP = f"$B${R['csp']}"

R["ha_mode"] = r
p(ws,r,1,"Is this Prod env",font=B); p(ws,r,2,"Yes",fill=YE); p(ws,r,3,"Yes/No")
p(ws,r,10,"No = 1 replica all services, no zone-spread (dev/test); Yes = HA, zone-aware (prod/staging)",font=IT); r+=1

R["zones"] = r
# B7 shows "N/A" when non-prod, zones value when prod.
# Editable zones input lives in Q column (hidden helper column).
# All formulas referencing ZONES are already HA-guarded so "N/A" string is never evaluated.
p(ws,r,1,"Availability Zones",font=B)
p(ws,r,2,f'=IF(B{R["ha_mode"]}="No","N/A",Q{r})',fill=YE)
p(ws,r,3,"zones")
p(ws,r,17,2,fill=YE,fmt="0")  # Q column (hidden): editable zones input (default 2)
p(ws,r,10,"Prod only — min nodes per pool = zones (pod spread across AZs). Shows N/A when non-prod.",font=IT)
r+=1

R["target"] = r
p(ws,r,1,"Runtime target msg/s",font=B); p(ws,r,2,args.target,fill=YE,fmt="0"); p(ws,r,3,"msg/s")
p(ws,r,10,"← Primary scaling driver",font=IT); r+=1

R["runtime_active_sessions"] = r
p(ws,r,1,"Runtime target active sessions",font=B); p(ws,r,2,args.runtime_active_sessions,fill=YE,fmt="0"); p(ws,r,3,"sessions")
p(ws,r,10,"Runtime replicas use max(msg/s pods, active-session pods)",font=IT); r+=1

R["runtime_active_sessions_per_pod"] = r
p(ws,r,1,"Runtime safe active sessions/pod",font=B); p(ws,r,2,args.runtime_active_sessions_per_pod,fill=YE,fmt="0.0"); p(ws,r,3,"sessions/pod")
p(ws,r,10,"Safe concurrent active-session capacity per runtime pod",font=IT); r+=1

R["studio_runtime_ratio"] = r
p(ws,r,1,"Studio/runtime pod ratio",font=B); p(ws,r,2,args.studio_runtime_ratio,fill=YE,fmt="0%"); p(ws,r,3,"ratio")
p(ws,r,10,"Studio pods are at least this share of runtime pods",font=IT); r+=1

R["search_target"] = r
p(ws,r,1,"Search target queries/s",font=B); p(ws,r,2,args.search_target,fill=YE,fmt="0"); p(ws,r,3,"qps")
p(ws,r,10,"Total query load (keyword + hybrid + semantic + aggregation)",font=IT); r+=1

R["embed_ratio"] = r
p(ws,r,1,"Embedding query ratio",font=B); p(ws,r,2,args.embed_ratio,fill=YE,fmt="0%"); p(ws,r,3,"ratio")
p(ws,r,10,"Fraction needing BGE-M3 (hybrid+semantic); 60% bypass at keyword/agg",font=IT); r+=1

R["ingest_target"] = r
p(ws,r,1,"Ingestion target files/min",font=B); p(ws,r,2,args.ingest_target,fill=YE,fmt="0.0"); p(ws,r,3,"files/min")
p(ws,r,10,"File upload pipeline throughput target (SearchAI + Docling + BGE-M3)",font=IT); r+=1

R["knowledge_enabled"] = r
p(ws,r,1,"Search (Knowledge) enabled",font=B); p(ws,r,2,"Yes",fill=YE); p(ws,r,3,"Yes/No")
p(ws,r,10,"No = no file ingestion (Docling=0, ingestion pods=0); Yes = full search pipeline",font=IT); r+=1

KNOWLEDGE_EN_EARLY = f"B{R['knowledge_enabled']}"

R["bge_m3_enabled"] = r
p(ws,r,1,"BGE-M3 enabled",font=B); p(ws,r,2,f'=IF({KNOWLEDGE_EN_EARLY}="Yes","Yes","No")',fill=YE); p(ws,r,3,"Yes/No")
p(ws,r,10,"Auto No if Knowledge disabled; set to No if keyword/aggregation only (no embedding needed)",font=IT); r+=1

# Early refs for conditional mode display
BGE_EN_EARLY = f"B{R['bge_m3_enabled']}"

R["bge_m3_mode"] = r
p(ws,r,1,"BGE-M3 mode",font=B); p(ws,r,2,f'=IF(AND({KNOWLEDGE_EN_EARLY}="Yes",{BGE_EN_EARLY}="Yes"),"GPU","N/A")',fill=YE); p(ws,r,3,"GPU/CPU/N/A")
p(ws,r,10,"GPU = T4 (15 rps/pod); CPU = user pool (4.4 rps/pod); N/A if Knowledge or BGE-M3 disabled",font=IT); r+=1

R["docling_mode"] = r
p(ws,r,1,"Docling mode",font=B); p(ws,r,2,f'=IF({KNOWLEDGE_EN_EARLY}="Yes","GPU","N/A")',fill=YE); p(ws,r,3,"GPU/CPU/N/A")
p(ws,r,10,"GPU = T4 (43 files/min); CPU = user pool; N/A if Knowledge disabled",font=IT); r+=1

R["wf_target"] = r
p(ws,r,1,"Workflow target exec/s",font=B); p(ws,r,2,args.workflow_target,fill=YE,fmt="0"); p(ws,r,3,"exec/s")
p(ws,r,10,"Reference only; edit workflow replicas below",font=IT); r+=1

R["pl_target"] = r
p(ws,r,1,"Pipeline target exec/s",font=B); p(ws,r,2,args.pipeline_target,fill=YE,fmt="0"); p(ws,r,3,"exec/s")
p(ws,r,10,"Reference only; edit pipeline replicas below",font=IT); r+=1

R["util"] = r
p(ws,r,1,"Utilization headroom",font=B); p(ws,r,2,args.util,fill=YE,fmt="0%")
p(ws,r,10,"0.70 = 30% buffer",font=IT); r+=1

R["ppn"] = r
p(ws,r,1,"Pods per node",font=B); p(ws,r,2,f'=IF(B{R["ha_mode"]}="No",4,{args.pods_per_node})',fill=YE,fmt="0")
p(ws,r,10,"Non-prod default: 4; Prod default: 3",font=IT); r+=1

# ── Node Pool VM Selection (editable dropdowns with derived specs) ──
# Helper: build nested IF formula to derive CPU/Mem/Price from selected VM name
def _vm_lookup_formula(vm_cell, specs_dict, field_idx):
    """Build nested IF(vm="X",val,...) formula. field_idx: 0=cpu, 1=mem, 2=price."""
    items = list(specs_dict.items())
    if not items:
        return "0"
    # Build from inside out
    formula = str(items[-1][1][field_idx])
    for name, spec in reversed(items[:-1]):
        formula = f'IF({vm_cell}="{name}",{spec[field_idx]},{formula})'
    return formula

def _vm_lookup_formula_csp(vm_cell, azure_specs, aws_specs, field_idx, csp_ref, gcp_specs=None):
    """CSP-aware lookup: merges ALL CSP specs into one nested IF chain.

    This ensures that regardless of which CSP is selected, the VM name
    from ANY CSP will resolve correctly. When the CSP switches, the VM cell
    formula (csp_text) auto-changes the VM name to the correct CSP's default,
    and this merged lookup finds it.  If the VM name doesn't match any known
    entry, the formula falls through to 0 (safe fallback).
    """
    # Merge all specs into one dict — no duplicates because names are CSP-specific
    merged = {}
    merged.update(azure_specs)
    merged.update(aws_specs)
    if gcp_specs:
        merged.update(gcp_specs)
    return _vm_lookup_formula(vm_cell, merged, field_idx)

# User pool VM
R["user_vm"] = r
default_user_vm_az = POOL_VM_OPTIONS["user"][0]
default_user_vm_aws = POOL_VM_OPTIONS_AWS["user"][0]
default_user_vm_gcp = POOL_VM_OPTIONS_GCP["user"][0]
p(ws,r,1,"User node VM",font=B); p(ws,r,2,csp_text(CSP,default_user_vm_aws,default_user_vm_az,default_user_vm_gcp),fill=YE)
p(ws,r,10,"Select VM size for user/app workload pool",font=IT); r+=1
R["user_vm_cpu"] = r
user_vm_ref = f"B{R['user_vm']}"
az_user = {k: AZURE_VM_SPECS[k] for k in POOL_VM_OPTIONS["user"] if k in AZURE_VM_SPECS}
aws_user = {k: AWS_VM_SPECS[k] for k in POOL_VM_OPTIONS_AWS["user"] if k in AWS_VM_SPECS}
gcp_user = {k: GCP_VM_SPECS[k] for k in POOL_VM_OPTIONS_GCP["user"] if k in GCP_VM_SPECS}
p(ws,r,1,"  vCPU / node",font=IT); p(ws,r,2,f'={_vm_lookup_formula_csp(user_vm_ref, az_user, aws_user, 0, CSP, gcp_user)}',fill=GR,fmt="0")
p(ws,r,3,"vCPU"); r+=1
R["user_vm_mem"] = r
p(ws,r,1,"  Memory / node",font=IT); p(ws,r,2,f'={_vm_lookup_formula_csp(user_vm_ref, az_user, aws_user, 1, CSP, gcp_user)}',fill=GR,fmt="0")
p(ws,r,3,"GiB"); r+=1
R["node_price"] = r
p(ws,r,1,"  $/month",font=IT); p(ws,r,2,f'={_vm_lookup_formula_csp(user_vm_ref, az_user, aws_user, 2, CSP, gcp_user)}',fill=GR,fmt='"$"#,##0')
p(ws,r,10,"Provider on-demand compute × 730h",font=IT); r+=1

# Database (mongo) pool VM
R["mongo_vm"] = r
default_mongo_vm_az = POOL_VM_OPTIONS["database (mongo)"][0]
default_mongo_vm_aws = POOL_VM_OPTIONS_AWS["database (mongo)"][0]
default_mongo_vm_gcp = POOL_VM_OPTIONS_GCP["database (mongo)"][0]
p(ws,r,1,"Database (mongo) node VM",font=B); p(ws,r,2,csp_text(CSP,default_mongo_vm_aws,default_mongo_vm_az,default_mongo_vm_gcp),fill=YE)
p(ws,r,10,"Dedicated MongoDB pool — r7i.4xlarge equivalent (16 CPU / 128 GiB); 1 node per replica",font=IT); r+=1
R["mongo_vm_cpu"] = r
mongo_vm_ref = f"B{R['mongo_vm']}"
az_mongo = {k: AZURE_VM_SPECS[k] for k in POOL_VM_OPTIONS["database (mongo)"] if k in AZURE_VM_SPECS}
aws_mongo = {k: AWS_VM_SPECS[k] for k in POOL_VM_OPTIONS_AWS["database (mongo)"] if k in AWS_VM_SPECS}
gcp_mongo = {k: GCP_VM_SPECS[k] for k in POOL_VM_OPTIONS_GCP["database (mongo)"] if k in GCP_VM_SPECS}
p(ws,r,1,"  vCPU / node",font=IT); p(ws,r,2,f'={_vm_lookup_formula_csp(mongo_vm_ref, az_mongo, aws_mongo, 0, CSP, gcp_mongo)}',fill=GR,fmt="0")
p(ws,r,3,"vCPU"); r+=1
R["mongo_vm_mem"] = r
p(ws,r,1,"  Memory / node",font=IT); p(ws,r,2,f'={_vm_lookup_formula_csp(mongo_vm_ref, az_mongo, aws_mongo, 1, CSP, gcp_mongo)}',fill=GR,fmt="0")
p(ws,r,3,"GiB"); r+=1
R["mongo_node_price"] = r
p(ws,r,1,"  $/month",font=IT); p(ws,r,2,f'={_vm_lookup_formula_csp(mongo_vm_ref, az_mongo, aws_mongo, 2, CSP, gcp_mongo)}',fill=GR,fmt='"$"#,##0')
p(ws,r,10,"Dedicated MongoDB nodes (1 replica per node with headroom)",font=IT); r+=1

# Database (non-mongo) pool VM
R["db_vm"] = r
default_db_vm_az = POOL_VM_OPTIONS["database (non-mongo)"][0]
default_db_vm_aws = POOL_VM_OPTIONS_AWS["database (non-mongo)"][0]
default_db_vm_gcp = POOL_VM_OPTIONS_GCP["database (non-mongo)"][0]
p(ws,r,1,"Database (non-mongo) node VM",font=B); p(ws,r,2,csp_text(CSP,default_db_vm_aws,default_db_vm_az,default_db_vm_gcp),fill=YE)
p(ws,r,10,"Select VM size for Redis/OpenSearch/Neo4j/RocksDB pool",font=IT); r+=1
R["db_vm_cpu"] = r
db_vm_ref = f"B{R['db_vm']}"
az_db = {k: AZURE_VM_SPECS[k] for k in POOL_VM_OPTIONS["database (non-mongo)"] if k in AZURE_VM_SPECS}
aws_db = {k: AWS_VM_SPECS[k] for k in POOL_VM_OPTIONS_AWS["database (non-mongo)"] if k in AWS_VM_SPECS}
gcp_db = {k: GCP_VM_SPECS[k] for k in POOL_VM_OPTIONS_GCP["database (non-mongo)"] if k in GCP_VM_SPECS}
p(ws,r,1,"  vCPU / node",font=IT); p(ws,r,2,f'={_vm_lookup_formula_csp(db_vm_ref, az_db, aws_db, 0, CSP, gcp_db)}',fill=GR,fmt="0")
p(ws,r,3,"vCPU"); r+=1
R["db_vm_mem"] = r
p(ws,r,1,"  Memory / node",font=IT); p(ws,r,2,f'={_vm_lookup_formula_csp(db_vm_ref, az_db, aws_db, 1, CSP, gcp_db)}',fill=GR,fmt="0")
p(ws,r,3,"GiB"); r+=1
R["db_node_price"] = r
p(ws,r,1,"  $/month",font=IT); p(ws,r,2,f'={_vm_lookup_formula_csp(db_vm_ref, az_db, aws_db, 2, CSP, gcp_db)}',fill=GR,fmt='"$"#,##0')
p(ws,r,10,"Non-mongo DB pool (Redis, OpenSearch, Neo4j, RocksDB)",font=IT); r+=1

# GPU pool VM (fixed — only T4 option per CSP)
R["gpu_vm"] = r
p(ws,r,1,"GPU node VM",font=B); p(ws,r,2,csp_text(CSP,"g4dn.xlarge","Standard_NC4as_T4_v3","n1-standard-4+T4"),fill=GR)
p(ws,r,10,"T4 GPU worker node (fixed)",font=IT); r+=1
R["gpu_node_price"] = r
p(ws,r,1,"  $/month",font=IT); p(ws,r,2,csp_value(CSP,AWS_G4DN_XLARGE_MONTHLY,NC4AS_T4_V3_MONTHLY,GCP_N1_STD_4_T4_MONTHLY),fill=GR,fmt='"$"#,##0')
p(ws,r,10,"Provider GPU compute × 730h",font=IT); r+=1

# System pool VM
R["system_vm"] = r
default_sys_vm_az = POOL_VM_OPTIONS["system"][0]
default_sys_vm_aws = POOL_VM_OPTIONS_AWS["system"][0]
default_sys_vm_gcp = POOL_VM_OPTIONS_GCP["system"][0]
p(ws,r,1,"System node VM",font=B); p(ws,r,2,csp_text(CSP,default_sys_vm_aws,default_sys_vm_az,default_sys_vm_gcp),fill=YE)
p(ws,r,10,"Select VM size for system workloads (ClickHouse, Kafka)",font=IT); r+=1
R["system_vm_cpu"] = r
sys_vm_ref = f"B{R['system_vm']}"
az_sys = {k: AZURE_VM_SPECS[k] for k in POOL_VM_OPTIONS["system"] if k in AZURE_VM_SPECS}
aws_sys = {k: AWS_VM_SPECS[k] for k in POOL_VM_OPTIONS_AWS["system"] if k in AWS_VM_SPECS}
gcp_sys = {k: GCP_VM_SPECS[k] for k in POOL_VM_OPTIONS_GCP["system"] if k in GCP_VM_SPECS}
p(ws,r,1,"  vCPU / node",font=IT); p(ws,r,2,f'={_vm_lookup_formula_csp(sys_vm_ref, az_sys, aws_sys, 0, CSP, gcp_sys)}',fill=GR,fmt="0")
p(ws,r,3,"vCPU"); r+=1
R["system_vm_mem"] = r
p(ws,r,1,"  Memory / node",font=IT); p(ws,r,2,f'={_vm_lookup_formula_csp(sys_vm_ref, az_sys, aws_sys, 1, CSP, gcp_sys)}',fill=GR,fmt="0")
p(ws,r,3,"GiB"); r+=1
R["system_node_price"] = r
p(ws,r,1,"  $/month",font=IT); p(ws,r,2,f'={_vm_lookup_formula_csp(sys_vm_ref, az_sys, aws_sys, 2, CSP, gcp_sys)}',fill=GR,fmt='"$"#,##0')
p(ws,r,10,"Provider system compute × 730h",font=IT); r+=1

# Operator pool VM
R["operator_vm"] = r
default_op_vm_az = POOL_VM_OPTIONS["operator"][0]
default_op_vm_aws = POOL_VM_OPTIONS_AWS["operator"][0]
default_op_vm_gcp = POOL_VM_OPTIONS_GCP["operator"][0]
p(ws,r,1,"Operator node VM",font=B); p(ws,r,2,csp_text(CSP,default_op_vm_aws,default_op_vm_az,default_op_vm_gcp),fill=YE)
p(ws,r,10,"Small VMs for operator/controller workloads",font=IT); r+=1
R["operator_vm_cpu"] = r
op_vm_ref = f"B{R['operator_vm']}"
az_op = {k: AZURE_VM_SPECS[k] for k in POOL_VM_OPTIONS["operator"] if k in AZURE_VM_SPECS}
aws_op = {k: AWS_VM_SPECS[k] for k in POOL_VM_OPTIONS_AWS["operator"] if k in AWS_VM_SPECS}
gcp_op = {k: GCP_VM_SPECS[k] for k in POOL_VM_OPTIONS_GCP["operator"] if k in GCP_VM_SPECS}
p(ws,r,1,"  vCPU / node",font=IT); p(ws,r,2,f'={_vm_lookup_formula_csp(op_vm_ref, az_op, aws_op, 0, CSP, gcp_op)}',fill=GR,fmt="0")
p(ws,r,3,"vCPU"); r+=1
R["operator_vm_mem"] = r
p(ws,r,1,"  Memory / node",font=IT); p(ws,r,2,f'={_vm_lookup_formula_csp(op_vm_ref, az_op, aws_op, 1, CSP, gcp_op)}',fill=GR,fmt="0")
p(ws,r,3,"GiB"); r+=1
R["operator_node_price"] = r
p(ws,r,1,"  $/month",font=IT); p(ws,r,2,f'={_vm_lookup_formula_csp(op_vm_ref, az_op, aws_op, 2, CSP, gcp_op)}',fill=GR,fmt='"$"#,##0')
p(ws,r,10,"Operator pool — small, spread across zones",font=IT); r+=1

R["hardware_discount"] = r
p(ws,r,1,"Hardware discount",font=B); p(ws,r,2,args.hardware_discount,fill=YE,fmt="0%")
p(ws,r,10,"Applied to node-pool MSRP",font=IT); r+=1
R["support_rate"] = r
p(ws,r,1,"Support rate",font=B); p(ws,r,2,args.support_rate,fill=YE,fmt="0%")
p(ws,r,10,"Applied to node-pool MSRP",font=IT); r+=1
R["object_storage_tb"] = r
p(ws,r,1,"Object storage",font=B); p(ws,r,2,args.object_storage_tb,fill=YE,fmt="0.0"); p(ws,r,3,"TiB"); r+=1
R["object_storage_price"] = r
p(ws,r,1,"Object storage $/GB-month",font=B); p(ws,r,2,csp_value(CSP,S3_STANDARD_GB_MONTH,args.object_storage_gb_mo,GCS_STANDARD_GB_MONTH),fill=GR,fmt='"$"0.0000')
p(ws,r,10,"S3 Standard / Azure Blob Hot LRS / GCS Standard",font=IT); r+=1
# ── Data Store Configuration (replicas + disk per node) ──
R["disk_price_gb"] = r
# Block storage $/GiB-month: Azure Premium SSD LRS ~$0.132, AWS gp3 ~$0.08, GCP PD-SSD ~$0.17
p(ws,r,1,"Block storage $/GiB-month",font=B); p(ws,r,2,csp_value(CSP,0.08,0.132,0.17),fill=GR,fmt='"$"0.000')
p(ws,r,10,"EBS gp3 / Azure Premium SSD / GCP PD-SSD",font=IT); r+=1
# Header for the data store config mini-table
p(ws,r,1,"Data Store",font=B); p(ws,r,2,"Replicas",font=B); p(ws,r,3,"Disk/Node GiB",font=B); p(ws,r,4,"Deployment",font=B)
p(ws,r,6,"Deployment legend ↓",font=IT); r+=1

# Replicas input represents the configured instance count:
#   - Enterprise  → instances purchased from the vendor (drives license qty only; no self-host pods/PVC)
#   - Self-Hosted → replicas on our cluster (drives pod CPU/mem, PVC, and node pool)
#   - N/A         → forced to 0 across the board
# Disk input is per-instance: meaningful for both Enterprise (informational/sizing) and Self-Hosted (PVC cost).
# Defaults below preserve the calculator's historical "self-hosted" baseline so totals stay unchanged
# unless the operator opts into Enterprise SaaS or disables a service.
HA_EARLY = f'B{R["ha_mode"]}'

# One consolidated legend, vertically merged across F:J of the first three datastore rows
# (MongoDB / OpenSearch / Neo4j). No extra rows consumed; per-row B/C/D inputs remain editable.
_LEGEND_FILL = PatternFill("solid", fgColor="E8F4F8")  # NOT in EDITABLE_FILL_RGBS
_LEGEND_TEXT = (
    "Deployment legend\n"
    "• Enterprise  → vendor-managed SaaS: license/pricing only, no hardware on our cluster.\n"
    "• Self-Hosted → runs on our cluster: hardware/PVC only, no commercial license.\n"
    "• N/A         → service not required: may rely on shared resources; all costs 0."
)
R["datastore_legend_top"] = r  # MongoDB row = top of the merged legend block

R["mongo_replicas"] = r
p(ws,r,1,"  MongoDB",font=B)
p(ws,r,2,f'=IF(D{r}="N/A",0,IF({HA_EARLY}="No",1,3))',fill=YE,fmt="0")
p(ws,r,3,f'=IF(D{r}="N/A",0,{args.mongo_disk_gb})',fill=YE,fmt="0")
p(ws,r,4,"Self-Hosted",fill=YE,al=AL_C)
r+=1
R["opensearch_replicas"] = r
p(ws,r,1,"  OpenSearch",font=B)
p(ws,r,2,f'=IF(D{r}="N/A",0,IF({HA_EARLY}="No",1,3))',fill=YE,fmt="0")
p(ws,r,3,f'=IF(D{r}="N/A",0,{args.opensearch_disk_gb})',fill=YE,fmt="0")
p(ws,r,4,"Self-Hosted",fill=YE,al=AL_C)
r+=1
R["neo4j_replicas"] = r
p(ws,r,1,"  Neo4j",font=B)
p(ws,r,2,f'=IF(D{r}="N/A",0,IF({HA_EARLY}="No",1,3))',fill=YE,fmt="0")
p(ws,r,3,f'=IF(D{r}="N/A",0,{args.neo4j_disk_gb})',fill=YE,fmt="0")
p(ws,r,4,"Self-Hosted",fill=YE,al=AL_C)
r+=1

# Place the merged legend block now that all three rows exist.
R["datastore_legend_bottom"] = r - 1
ws.merge_cells(
    start_row=R["datastore_legend_top"],    start_column=6,
    end_row=R["datastore_legend_bottom"],   end_column=10,
)
_legend_cell = ws.cell(row=R["datastore_legend_top"], column=6, value=_LEGEND_TEXT)
_legend_cell.font = IT
_legend_cell.fill = _LEGEND_FILL
_legend_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
_legend_cell.border = BD
R["redis_replicas"] = r
_redis_prod_default = args.redis_shards + args.redis_shards * args.redis_replicas_per_shard
p(ws,r,1,"  Redis",font=B)
p(ws,r,2,f'=IF(D{r}="N/A",0,IF({HA_EARLY}="No",1,{_redis_prod_default}))',fill=YE,fmt="0")
p(ws,r,3,f'=IF(D{r}="N/A",0,{args.redis_disk_gb})',fill=YE,fmt="0")
p(ws,r,4,"Self-Hosted",fill=YE,al=AL_C)
p(ws,r,10,"Replicas & PVC per Redis node (non-prod default: 1); Deployment gates row.",font=IT); r+=1
R["clickhouse_replicas"] = r
_ch_prod_default = args.clickhouse_shards * args.clickhouse_replicas_per_shard
p(ws,r,1,"  ClickHouse",font=B)
p(ws,r,2,f'=IF(D{r}="N/A",0,IF({HA_EARLY}="No",1,{_ch_prod_default}))',fill=YE,fmt="0")
p(ws,r,3,f'=IF(D{r}="N/A",0,{args.clickhouse_disk_gb})',fill=YE,fmt="0")
p(ws,r,4,"Self-Hosted",fill=YE,al=AL_C)
p(ws,r,10,"Replicas & PVC per ClickHouse shard replica (non-prod default: 1); Deployment gates row.",font=IT); r+=1
R["kafka_replicas"] = r
p(ws,r,1,"  Kafka",font=B)
p(ws,r,2,f'=IF(D{r}="N/A",0,IF({HA_EARLY}="No",1,3))',fill=YE,fmt="0")
p(ws,r,3,f'=IF(D{r}="N/A",0,{args.kafka_disk_gb})',fill=YE,fmt="0")
p(ws,r,4,"Self-Hosted",fill=YE,al=AL_C)
p(ws,r,10,"Replicas & PVC per Kafka broker (non-prod default: 1); Deployment gates row.",font=IT); r+=1
R["rocksdb_replicas"] = r
p(ws,r,1,"  RocksDB",font=B)
p(ws,r,2,f'=IF(D{r}="N/A",0,IF({HA_EARLY}="No",1,3))',fill=YE,fmt="0")
p(ws,r,3,f'=IF(D{r}="N/A",0,{args.rocksdb_disk_gb})',fill=YE,fmt="0")
# RocksDB is an embedded KV store — no vendor SaaS exists, so only Self-Hosted / N/A are valid.
p(ws,r,4,"Self-Hosted",fill=YE,al=AL_C)
p(ws,r,10,"Embedded KV (no SaaS option); Self-Hosted only. Replicas & PVC per node.",font=IT); r+=1

R["dpt"] = r
p(ws,r,1,"Dirty pages per turn (IOPS)",font=B); p(ws,r,2,args.dirty_pages,fill=YE,fmt="0.0")
p(ws,r,10,"OTel: 0.78 writes × index amplification",font=IT); r+=1

csp_validation = DataValidation(type="list", formula1='"Azure,AWS,GCP"', allow_blank=False)
ws.add_data_validation(csp_validation)
csp_validation.add(ws[f"B{R['csp']}"])

# Per-datastore deployment dropdown — gates replicas, disk, and licensing per row.
#   Most stores: Enterprise (SaaS) | Self-Hosted | N/A
#   RocksDB:     Self-Hosted | N/A   (embedded KV; no vendor SaaS exists)
datastore_deployment_validation = DataValidation(type="list", formula1='"Enterprise,Self-Hosted,N/A"', allow_blank=False)
ws.add_data_validation(datastore_deployment_validation)
for _ds_key in ("mongo_replicas", "opensearch_replicas", "neo4j_replicas", "redis_replicas", "clickhouse_replicas", "kafka_replicas"):
    datastore_deployment_validation.add(ws[f"D{R[_ds_key]}"])

rocksdb_deployment_validation = DataValidation(type="list", formula1='"Self-Hosted,N/A"', allow_blank=False)
ws.add_data_validation(rocksdb_deployment_validation)
rocksdb_deployment_validation.add(ws[f"D{R['rocksdb_replicas']}"])

# Dropdowns for Knowledge, BGE-M3, and Docling selectors
knowledge_enabled_validation = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
ws.add_data_validation(knowledge_enabled_validation)
knowledge_enabled_validation.add(ws[f"B{R['knowledge_enabled']}"])

bge_enabled_validation = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
ws.add_data_validation(bge_enabled_validation)
bge_enabled_validation.add(ws[f"B{R['bge_m3_enabled']}"])

gpu_cpu_validation = DataValidation(type="list", formula1='"GPU,CPU,N/A"', allow_blank=False)
ws.add_data_validation(gpu_cpu_validation)
gpu_cpu_validation.add(ws[f"B{R['bge_m3_mode']}"])
gpu_cpu_validation.add(ws[f"B{R['docling_mode']}"])

ha_validation = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
ws.add_data_validation(ha_validation)
ha_validation.add(ws[f"B{R['ha_mode']}"])

zones_validation = DataValidation(type="whole", operator="between", formula1="1", formula2="3", allow_blank=False)
ws.add_data_validation(zones_validation)
zones_validation.add(ws[f"Q{R['zones']}"])  # Validation on Q column (hidden editable input), B shows formula

# VM dropdown validations — include ALL CSPs so validation passes regardless of CSP selection
user_vm_options = ",".join(POOL_VM_OPTIONS["user"] + POOL_VM_OPTIONS_AWS["user"] + POOL_VM_OPTIONS_GCP["user"])
user_vm_validation = DataValidation(type="list", formula1=f'"{user_vm_options}"', allow_blank=False)
ws.add_data_validation(user_vm_validation)
user_vm_validation.add(ws[f"B{R['user_vm']}"])

mongo_vm_options = ",".join(POOL_VM_OPTIONS["database (mongo)"] + POOL_VM_OPTIONS_AWS["database (mongo)"] + POOL_VM_OPTIONS_GCP["database (mongo)"])
mongo_vm_validation = DataValidation(type="list", formula1=f'"{mongo_vm_options}"', allow_blank=False)
ws.add_data_validation(mongo_vm_validation)
mongo_vm_validation.add(ws[f"B{R['mongo_vm']}"])

db_vm_options = ",".join(POOL_VM_OPTIONS["database (non-mongo)"] + POOL_VM_OPTIONS_AWS["database (non-mongo)"] + POOL_VM_OPTIONS_GCP["database (non-mongo)"])
db_vm_validation = DataValidation(type="list", formula1=f'"{db_vm_options}"', allow_blank=False)
ws.add_data_validation(db_vm_validation)
db_vm_validation.add(ws[f"B{R['db_vm']}"])

op_vm_options = ",".join(POOL_VM_OPTIONS["operator"] + POOL_VM_OPTIONS_AWS["operator"] + POOL_VM_OPTIONS_GCP["operator"])
op_vm_validation = DataValidation(type="list", formula1=f'"{op_vm_options}"', allow_blank=False)
ws.add_data_validation(op_vm_validation)
op_vm_validation.add(ws[f"B{R['operator_vm']}"])

sys_vm_options = ",".join(POOL_VM_OPTIONS["system"] + POOL_VM_OPTIONS_AWS["system"] + POOL_VM_OPTIONS_GCP["system"])
sys_vm_validation = DataValidation(type="list", formula1=f'"{sys_vm_options}"', allow_blank=False)
ws.add_data_validation(sys_vm_validation)
sys_vm_validation.add(ws[f"B{R['system_vm']}"])

positive_number_validation = DataValidation(type="decimal", operator="greaterThan", formula1="0", allow_blank=False)
nonnegative_number_validation = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1="0", allow_blank=False)
ws.add_data_validation(positive_number_validation)
ws.add_data_validation(nonnegative_number_validation)
for key in ("target", "util", "ppn"):
    positive_number_validation.add(ws[f"B{R[key]}"])
for key in ("runtime_active_sessions", "runtime_active_sessions_per_pod", "studio_runtime_ratio"):
    nonnegative_number_validation.add(ws[f"B{R[key]}"])

for key in ("runtime_active_sessions", "runtime_active_sessions_per_pod"):
    ws.row_dimensions[R[key]].hidden = True

# Hide columns Q (zones helper) and R (throughput override) — user unhides or uses Name Box
ws.column_dimensions['Q'].hidden = True
ws.column_dimensions['R'].hidden = True

# Formula refs
T = f"B{R['target']}"; ST = f"B{R['search_target']}"; ER = f"B{R['embed_ratio']}"; IT_REF = f"B{R['ingest_target']}"; WT = f"B{R['wf_target']}"; PT = f"B{R['pl_target']}"
KNOWLEDGE_EN = f"B{R['knowledge_enabled']}"; BGE_EN = f"B{R['bge_m3_enabled']}"; BGE_MODE = f"B{R['bge_m3_mode']}"; DOC_MODE = f"B{R['docling_mode']}"; HA_REF = f"B{R['ha_mode']}"
ZONES = f"Q{R['zones']}"  # Q column (hidden) holds the numeric zones input (B shows "N/A" when non-prod)
ACTIVE_SESSIONS = f"B{R['runtime_active_sessions']}"; SESSIONS_PER_POD = f"B{R['runtime_active_sessions_per_pod']}"; STUDIO_RATIO = f"B{R['studio_runtime_ratio']}"
U = f"B{R['util']}"; PPN = f"B{R['ppn']}"; HAM = ZONES; DPT = f"B{R['dpt']}"
STUDIO_MIN = ZONES; SEARCH_MIN = ZONES; WFE_MIN = ZONES
# Data store replica refs from editable input cells
MGN = f"B{R['mongo_replicas']}"
OSRN = f"B{R['opensearch_replicas']}"
N4JN = f"B{R['neo4j_replicas']}"
RDTOTAL = f"B{R['redis_replicas']}"
CHTOTAL = f"B{R['clickhouse_replicas']}"
KAFKAN = f"B{R['kafka_replicas']}"
ROCKSN = f"B{R['rocksdb_replicas']}"
# Deployment dropdown refs (Enterprise/Self-Hosted/N/A) — used to gate licensing and dependent capacity.
MGN_DEPLOY = f"D{R['mongo_replicas']}"
OSRN_DEPLOY = f"D{R['opensearch_replicas']}"
N4JN_DEPLOY = f"D{R['neo4j_replicas']}"
RDTOTAL_DEPLOY = f"D{R['redis_replicas']}"
CHTOTAL_DEPLOY = f"D{R['clickhouse_replicas']}"
KAFKAN_DEPLOY = f"D{R['kafka_replicas']}"
ROCKSN_DEPLOY = f"D{R['rocksdb_replicas']}"
# Self-host replica refs: only "Self-Hosted" generates pod CPU/memory & PVC on our cluster.
# Enterprise (vendor-managed SaaS) → 0 self-host footprint (license still billed below).
# N/A → 0 (the underlying input cell already collapses to 0 too).
MGN_SH = f'IF({MGN_DEPLOY}="Self-Hosted",{MGN},0)'
OSRN_SH = f'IF({OSRN_DEPLOY}="Self-Hosted",{OSRN},0)'
N4JN_SH = f'IF({N4JN_DEPLOY}="Self-Hosted",{N4JN},0)'
RDTOTAL_SH = f'IF({RDTOTAL_DEPLOY}="Self-Hosted",{RDTOTAL},0)'
CHTOTAL_SH = f'IF({CHTOTAL_DEPLOY}="Self-Hosted",{CHTOTAL},0)'
KAFKAN_SH = f'IF({KAFKAN_DEPLOY}="Self-Hosted",{KAFKAN},0)'
ROCKSN_SH = f'IF({ROCKSN_DEPLOY}="Self-Hosted",{ROCKSN},0)'
NP = f"B{R['node_price']}"; DNP = f"B{R['db_node_price']}"; GNP = f"B{R['gpu_node_price']}"; SNP = f"B{R['system_node_price']}"
MNP = f"B{R['mongo_node_price']}"; ONP = f"B{R['operator_node_price']}"
USER_VM = f"B{R['user_vm']}"; DB_VM = f"B{R['db_vm']}"; GPU_VM = f"B{R['gpu_vm']}"; SYSTEM_VM = f"B{R['system_vm']}"
MONGO_VM = f"B{R['mongo_vm']}"; OPERATOR_VM = f"B{R['operator_vm']}"
USER_VM_CPU = f"B{R['user_vm_cpu']}"; USER_VM_MEM = f"B{R['user_vm_mem']}"
DB_VM_CPU = f"B{R['db_vm_cpu']}"; DB_VM_MEM = f"B{R['db_vm_mem']}"
MONGO_VM_CPU = f"B{R['mongo_vm_cpu']}"; MONGO_VM_MEM = f"B{R['mongo_vm_mem']}"
OP_VM_CPU = f"B{R['operator_vm_cpu']}"; OP_VM_MEM = f"B{R['operator_vm_mem']}"
SYS_VM_CPU = f"B{R['system_vm_cpu']}"; SYS_VM_MEM = f"B{R['system_vm_mem']}"
HD = f"B{R['hardware_discount']}"; SR = f"B{R['support_rate']}"
OST = f"B{R['object_storage_tb']}"; OSP = f"B{R['object_storage_price']}"

capacity_items = []
service_rows_by_name = {}

def vm_ref_for_pool(pool):
    if pool == "database (mongo)":
        return MONGO_VM
    if pool == "database (non-mongo)":
        return DB_VM
    if pool == "gpu":
        return GPU_VM
    if pool == "operator":
        return OPERATOR_VM
    if pool == "system":
        return SYSTEM_VM
    if pool == "user":
        return USER_VM
    return None

def vm_value_for_pool(pool, fallback="—"):
    ref = vm_ref_for_pool(pool)
    return f"={ref}" if ref else fallback

def add_capacity_item(tier, svc_name, pool, vm, replicas_ref, cpu_req, mem_req, source, cpu_ref=None, mem_ref=None):
    cpu = cpu_to_cores(cpu_req)
    mem = mem_to_gb(mem_req)
    capacity_items.append({
        "tier": tier,
        "service": svc_name,
        "pool": pool,
        "vm": vm,
        "replicas_ref": replicas_ref,
        "cpu": cpu,
        "mem": mem,
        "cpu_formula": f"({replicas_ref}*{cpu_ref})" if cpu_ref else f"({replicas_ref}*{cpu:g})",
        "mem_formula": f"({replicas_ref}*{mem_ref})" if mem_ref else f"({replicas_ref}*{mem:g})",
        "source": source,
    })

def throughput_default(tier_key, svc_name, default_perpod=None):
    """Per-pod throughput defaults.

    Measured values sourced from enterprise load test reports (May 6-8 2026):
    - File ingestion: file-ingestion-enterprise-report-2026-05-08.md
    - Query pipeline: searchai-unified-performance-report.md (§11.1 Saturation Matrix)

    Unmeasured services use conservative estimates marked with (est).
    """
    name = svc_name.lower()
    if tier_key == "runtime":
        if name == "runtime":
            return default_perpod if default_perpod is not None else per_pod
        if name == "studio":
            return 20
        if name == "admin":
            return 20
    if tier_key == "search":
        if name == "search-ai":
            # File ingestion throughput (May 8 enterprise report):
            # 5 VUs / 2 pods: 8.4 files/min → 4.2 files/min/pod
            # 30 VUs / 3 pods: 43 files/min → 14.3 files/min/pod
            # Pipeline saturates at 5 VUs (p95>56s). Bottleneck: API+workers co-located.
            return 14.3
        if name == "search-ai-runtime":
            # Measured §11.1: 30+ rps/pod (not saturated, event loop not contended).
            # Keyword/aggregation: 30+ rps. Hybrid GPU: 15 rps (BGE-M3 is bottleneck, not runtime).
            # Runtime itself never saturated at any concurrency level tested.
            return 30.0
        return default_perpod if default_perpod is not None else 7.5
    if tier_key in ("workflow", "pipeline"):
        return default_perpod if default_perpod is not None else 4
    if "redis" in name:
        return 5000
    if "mongodb" in name:
        return 1000
    if "clickhouse" in name:
        return 500
    if "kafka" in name:
        return 1000
    if name == "bge-m3-cpu":
        # Measured: 4.4 rps/pod (CPU 4-core), saturates at 10 concurrent. May 6-7.
        return 4.4
    if "bge" in name:
        # Measured: 15.0 rps/pod (T4 GPU), saturates at 30 concurrent. May 6-7.
        return 15.0
    if name == "docling-cpu":
        # CPU-only — not load-tested. Estimate based on no GPU acceleration.
        return 5.0
    if "docling" in name:
        # Measured: 1 GPU pod handled 43 files/min at 30 VUs (May 8).
        # Docling was mostly idle at lower VUs — SearchAI couldn't feed it fast enough.
        # True ceiling not reached; 43 files/min/pod is a floor, not a ceiling.
        return 43.0
    if "preprocessing" in name:
        # Not independently measured — estimate from downstream of docling (est)
        return 5
    if "opensearch" in name:
        # Measured: 130+ rps BM25, 100+ rps kNN — never saturated at 50 VUs
        return 130
    if "neo4j" in name:
        return 100  # (est) — not load-tested
    if "ingress" in name:
        return 1000
    if "livekit" in name:
        return 100
    if "minio" in name:
        return 500
    return 10

# ── Helper: render a scaling tier ──
def render_scaling_tier(ws, r, title, fill, tier_key, target_ref, default_perpod, services, latencies=None):
    """Render a tier with consistent columns, latency percentiles, and auto-scaling formulas.
    latencies: dict of {service_name: {p50, p95, p99}} or None for readonly catalog placeholders."""
    sec(ws, r, title, fill=fill, ncols=17); r+=1
    hdr(ws, r, COLS); r+=1
    svc_rows = []
    if not services:
        p(ws,r,1,"No services configured",font=IT,fill=GR)
        p(ws,r,14,"Add services to default_service_catalog()",font=IT)
        r+=1
    for svc in services:
        c0 = svc["containers"][0] if svc["containers"] else {}
        pool = svc_pool.get(svc["name"], "user")
        vm = pool_vm.get(pool, "—")
        lat = (latencies or {}).get(svc["name"], {})
        cpu_req = c0.get("cpuReq","—")
        mem_req = c0.get("memReq","—")
        max_rep = svc.get("max_replicas", svc.get("replicas", 1))
        scales_with_target = tier_key == "runtime" and svc["name"] == "runtime"
        min_ref = HAM
        if tier_key == "search":
            min_ref = SEARCH_MIN
        elif tier_key == "workflow" and svc["name"] == "workflow-engine":
            min_ref = WFE_MIN
        elif tier_key == "runtime" and svc["name"] == "studio":
            min_ref = STUDIO_MIN
        p(ws,r,1,svc["name"],font=B)
        pp_ref = f"M{r}"  # throughput is now col 13
        throughput_value = throughput_default(tier_key, svc["name"], default_perpod)
        if scales_with_target:
            p(ws,r,13,throughput_value,fill=YE,fmt="0.0")
            tps_pods = f"CEILING({target_ref}/({pp_ref}*{U}),1)"
            active_session_pods = f"IF(AND({ACTIVE_SESSIONS}>0,{SESSIONS_PER_POD}>0),CEILING({ACTIVE_SESSIONS}/({SESSIONS_PER_POD}*{U}),1),0)"
            ha_formula = f"MIN(C{r},MAX({min_ref},{tps_pods},{active_session_pods}))"
            # Non-prod: still scale for throughput (perf-test envs) but no HA floor
            nonha_formula = f"MAX(1,{tps_pods})"
            p(ws,r,2,f'=IF({HA_REF}="No",{nonha_formula},{ha_formula})',fill=GR,fmt="0")
        else:
            p(ws,r,13,throughput_value,fill=YE,fmt="0.0")
            if tier_key == "runtime" and svc["name"] == "studio" and "runtime" in service_rows_by_name:
                runtime_ref = f"B{service_rows_by_name['runtime']}"
                ha_formula = f"MIN(C{r},MAX({STUDIO_MIN},CEILING({runtime_ref}*{STUDIO_RATIO},1)))"
                # Non-prod: 1 studio pod (or scale if runtime scaled)
                nonha_formula = f"MAX(1,CEILING({runtime_ref}*{STUDIO_RATIO},1))"
                p(ws,r,2,f'=IF({HA_REF}="No",{nonha_formula},{ha_formula})',fill=GR,fmt="0")
            elif tier_key == "runtime":
                cat_replicas = int(svc.get("replicas", 2))
                ha_formula = f"MIN(C{r},MAX({ZONES},{cat_replicas}))"
                p(ws,r,2,f'=IF({HA_REF}="No",1,{ha_formula})',fill=GR,fmt="0")
            elif tier_key == "search" and svc["name"] == "search-ai":
                ha_formula = f"MIN(C{r},MAX({ZONES},CEILING({IT_REF}/({pp_ref}*{U}),1)))"
                # Non-prod: still scale for ingestion target but no HA floor
                nonha_formula = f"MAX(1,CEILING({IT_REF}/({pp_ref}*{U}),1))"
                p(ws,r,2,f'=IF({HA_REF}="No",{nonha_formula},{ha_formula})',fill=GR,fmt="0")
            elif tier_key == "search" and svc["name"] == "search-ai-runtime":
                ha_formula = f"MIN(C{r},MAX({ZONES},CEILING({ST}/({pp_ref}*{U}),1)))"
                # Non-prod: still scale for search target but no HA floor
                nonha_formula = f"MAX(1,CEILING({ST}/({pp_ref}*{U}),1))"
                p(ws,r,2,f'=IF({HA_REF}="No",{nonha_formula},{ha_formula})',fill=GR,fmt="0")
            elif tier_key == "workflow" and svc["name"] == "workflow-engine":
                cat_replicas = svc.get("replicas", 4)
                ha_formula = f"MIN(C{r},MAX({ZONES},{cat_replicas}))"
                p(ws,r,2,f'=IF({HA_REF}="No",1,{ha_formula})',fill=GR,fmt="0")
            elif tier_key == "pipeline" and svc["name"] == "pipeline-engine":
                cat_replicas = svc.get("replicas", 4)
                ha_formula = f"MIN(C{r},MAX({ZONES},{cat_replicas}))"
                p(ws,r,2,f'=IF({HA_REF}="No",1,{ha_formula})',fill=GR,fmt="0")
            else:
                cat_replicas = int(svc.get("replicas", 2))
                ha_formula = f"MIN(C{r},MAX({ZONES},{cat_replicas}))"
                p(ws,r,2,f'=IF({HA_REF}="No",1,{ha_formula})',fill=GR,fmt="0")
        # Col 3: Max (HPA) — editable
        p(ws,r,3,max_rep,fill=YE,fmt="0")
        # Col 4-7: CPU/Mem req/lim
        p(ws,r,4,cpu_to_cores(cpu_req),fill=GR,fmt="0.000")
        p(ws,r,5,cpu_to_cores(c0.get("cpuLim","—")),fill=GR,fmt="0.000")
        p(ws,r,6,mem_to_gb(mem_req),fill=GR,fmt="0.000")
        p(ws,r,7,mem_to_gb(c0.get("memLim","—")),fill=GR,fmt="0.000")
        # Col 8-9: Total CPU/Mem (replicas × Requests — K8s schedules on requests)
        p(ws,r,8,f"=B{r}*D{r}",fill=GR,fmt="0.00",font=B,al=AL_C)
        p(ws,r,9,f"=B{r}*F{r}",fill=GR,fmt="0.00",font=B,al=AL_C)
        # Col 10-11: Pool, VM
        p(ws,r,10,pool,fill=GR)
        p(ws,r,11,vm_value_for_pool(pool, vm),fill=GR)
        # Col 12: Req Nodes
        p(ws,r,12,f"=CEILING(B{r}/{PPN},1)",fill=GR,fmt="0",font=B,al=AL_C)
        add_capacity_item(tier_key, svc["name"], pool, vm, f"B{r}", cpu_req, mem_req, "target", cpu_ref=f"D{r}", mem_ref=f"F{r}")
        # Col 14-17: Latencies
        p50_est = lat.get("p50_est", "")
        p90 = lat.get("p90", "")
        p95 = lat.get("p95", "")
        p99 = lat.get("p99", "")
        lat_fill = YE
        p(ws,r,14,p50_est,fill=lat_fill,fmt="0" if isinstance(p50_est,(int,float)) else None)
        p(ws,r,15,p90,fill=lat_fill,fmt="0" if isinstance(p90,(int,float)) else None)
        p(ws,r,16,p95,fill=lat_fill,fmt="0" if isinstance(p95,(int,float)) else None)
        p(ws,r,17,p99,fill=lat_fill,fmt="0" if isinstance(p99,(int,float)) else None)
        svc_rows.append(r)
        service_rows_by_name[svc["name"]] = r
        r+=1

    nodes_formula = "+".join(f"L{x}" for x in svc_rows) if svc_rows else "0"
    return r, nodes_formula, None, svc_rows

# ── Tier 1: Runtime ──
r+=1
rt_latencies = {}

r, rt_nodes_f, rt_cost_r, rt_rows = render_scaling_tier(ws, r, "RUNTIME (chat agents)", T1, "runtime", T, per_pod, tiers.get("runtime",[]), latencies=rt_latencies)
R["rt_nodes"] = rt_nodes_f
R["rt_cost"] = rt_cost_r

# ── Tier 2: Search AI ──
# Measured latencies from enterprise load test reports (May 6-8 2026):
# - search-ai: file upload accept latency (file-ingestion-enterprise-report §Results)
# - search-ai-runtime: query latency hybrid GPU mode at 20 VUs (unified-performance-report §4.1)
search_latencies = {
    "search-ai": {
        "p50_est": 871,    # Upload accept p50 (ms) — May 8 file ingestion report
        "p90": 1750,       # Upload accept p90 (ms)
        "p95": 2420,       # Upload accept p95 (ms)
        "p99": 10010,      # Upload accept max (ms) — tail from large PDFs
    },
    "search-ai-runtime": {
        "p50_est": 938,    # Hybrid GPU avg at 10 VUs (ms) — May 6 saturation report §4.1
        "p90": 1113,       # Hybrid GPU avg at 20 VUs (approaching saturation)
        "p95": 1570,       # Hybrid GPU p95 at 20 VUs
        "p99": 1958,       # Hybrid GPU p95 at 50 VUs (saturated)
    },
}
r+=1
r, search_nodes_f, search_cost_r, search_rows = render_scaling_tier(ws, r, "SEARCH AI", T2, "search", ST, 7.5, tiers.get("search",[]), latencies=search_latencies)
R["search_nodes"] = search_nodes_f
R["search_cost"] = search_cost_r


# ── Tier 3: Workflow ──
r+=1
r, wf_nodes_f, wf_cost_r, wf_rows = render_scaling_tier(ws, r, "WORKFLOW", T3, "workflow", WT, 4, tiers.get("workflow",[]))
R["wf_nodes"] = wf_nodes_f
R["wf_cost"] = wf_cost_r

# ── Tier 4: Pipeline ──
r+=1
TP = PatternFill("solid", fgColor="FFF0E0")  # pipeline peach
r, pl_nodes_f, pl_cost_r, pl_rows = render_scaling_tier(ws, r, "PIPELINE ENGINE", TP, "pipeline", PT, 4, tiers.get("pipeline",[]))
R["pl_nodes"] = pl_nodes_f
R["pl_cost"] = pl_cost_r

# ── Tier 5: Data Stores (StatefulSet — fixed topology, no HPA) ──
r+=1; sec(ws,r,"DATA STORES",fill=T4,ncols=13); r+=1
hdr(ws,r,["Service","Replicas","","CPU Req","CPU Lim","Mem Req Gi","Mem Lim Gi","Total CPU","Total Mem Gi","Pool","VM Type","Throughput/pod","Notes"]); r+=1

data_costs = []
mongo_main_row = None
for svc in tiers.get("data",[]):
    for c0 in svc["containers"][:1]:
        pool = svc_pool.get(svc["name"],"user")
        if "clickhouse" in svc["name"].lower():
            pool = "system"
        elif "mongodb" in svc["name"].lower() or "mongo" in svc["name"].lower():
            pool = "database (mongo)"
        elif any(component in svc["name"].lower() for component in ("redis", "rocksdb", "opensearch", "neo4j")):
            pool = "database (non-mongo)"
        vm = pool_vm.get(pool,"—")
        cpu_req = c0.get("cpuReq","—")
        mem_req = c0.get("memReq","—")
        replicas_value = svc["replicas"]
        # Self-host capacity only counts when Deployment = "Self-Hosted".
        # Enterprise/N/A → 0 replicas (no pods on our cluster).
        if svc["name"] == "mongodb":
            replicas_ref = MGN_SH
        elif svc["name"] == "mongodb-arb":
            # Arbiter count tracks MongoDB self-host replicas; 0 when mongo ≤ 1
            # (a single-node mongo set has no quorum to arbitrate).
            replicas_ref = f"IF({MGN_SH}<=1,0,{MGN_SH})"
        elif svc["name"] == "redis-master":
            replicas_ref = f"CEILING({RDTOTAL_SH}/2,1)"
        elif "redis" in svc["name"]:
            replicas_ref = f"{RDTOTAL_SH}-CEILING({RDTOTAL_SH}/2,1)"
        elif "clickhouse-keeper" in svc["name"]:
            # Keepers only when ClickHouse is Self-Hosted: 3 in prod, 1 in non-prod.
            # Enterprise/N/A → 0 keeper pods.
            replicas_ref = f'IF({CHTOTAL_DEPLOY}="Self-Hosted",IF({HA_EARLY}="No",1,3),0)'
        elif "clickhouse" in svc["name"]:
            replicas_ref = CHTOTAL_SH
        elif "opensearch" in svc["name"]:
            replicas_ref = OSRN_SH
        elif "neo4j" in svc["name"]:
            replicas_ref = N4JN_SH
        elif "kafka" in svc["name"]:
            replicas_ref = KAFKAN_SH
        elif "rocksdb" in svc["name"]:
            replicas_ref = ROCKSN_SH
        else:
            replicas_ref = str(replicas_value)
        p(ws,r,1,svc["name"],font=B)
        # Data store replicas always follow user input — no prod/non-prod override.
        # User controls replica counts via editable input cells in "Data Store" section.
        if svc["name"] == "mongodb-arb":
            # Arbiter matches MongoDB self-host replicas; 0 when mongo ≤ 1 (single-node has no quorum)
            # or when MongoDB is Enterprise/N/A.
            p(ws,r,2,f'=IF({MGN_SH}<=1,0,{MGN_SH})',fill=GR,fmt="0")
        elif "clickhouse-keeper" in svc["name"]:
            # Keepers: 3 in prod (consensus quorum), 1 in non-prod (single-node dev)
            p(ws,r,2,f'={replicas_ref}',fill=GR,fmt="0")
        else:
            p(ws,r,2,f'={replicas_ref}',fill=GR,fmt="0")
        p(ws,r,4,cpu_to_cores(cpu_req),fill=GR,fmt="0.000")
        p(ws,r,5,cpu_to_cores(c0.get("cpuLim","—")),fill=GR,fmt="0.000")
        p(ws,r,6,mem_to_gb(mem_req),fill=GR,fmt="0.000")
        p(ws,r,7,mem_to_gb(c0.get("memLim","—")),fill=GR,fmt="0.000")
        p(ws,r,8,f"=B{r}*D{r}",fill=GR,fmt="0.00",font=B,al=AL_C)
        p(ws,r,9,f"=B{r}*F{r}",fill=GR,fmt="0.00",font=B,al=AL_C)
        p(ws,r,10,pool,fill=GR)
        p(ws,r,11,vm_value_for_pool(pool, vm),fill=GR)
        p(ws,r,12,throughput_default("data", svc["name"]),fill=YE,fmt="0.0")
        add_capacity_item("data", svc["name"], pool, vm, f"B{r}", cpu_req, mem_req, "catalog topology", cpu_ref=f"D{r}", mem_ref=f"F{r}")
        note = ""
        if svc["name"] == "mongodb-arb":
            note = "Arbiter: matches MongoDB replica count; 0 when mongo ≤ 1 or Enterprise/N/A"
        elif "mongodb" in svc["name"] and "arb" not in svc["name"]:
            note = "MongoDB replica set (editable replicas above)"
            mongo_main_row = r
        elif "redis-master" in svc["name"]: note = "Redis Cluster masters/shards on database pool"
        elif "redis-replica" in svc["name"]: note = "Redis replicas on database pool"
        elif "clickhouse" in svc["name"] and "keeper" not in svc["name"]: note = "OLAP analytics on system pool"
        elif "keeper" in svc["name"]: note = "ClickHouse consensus on system pool"
        elif "kafka" in svc["name"]: note = "Event streaming"
        elif "opensearch" in svc["name"]: note = "OpenSearch cluster (HA, 100Gi/node)"
        elif "neo4j" in svc["name"]: note = "Neo4j core cluster (HA)"
        elif "rocksdb" in svc["name"]: note = "Embedded KV store (HA replicated)"
        p(ws,r,13,note,font=IT)
        data_costs.append((r, pool, svc["name"]))
        r+=1
if not data_costs:
    p(ws,r,1,"No data services configured",font=IT,fill=GR)
    p(ws,r,13,"Add data services to default_service_catalog()",font=IT)
    r+=1

# MongoDB IOPS pressure signal
IOPS_REF = f"{T}*{DPT}"
mg_rep_ref = f"B{mongo_main_row}" if mongo_main_row else "3"

p(ws,r,1,"  MongoDB IOPS",font=B)
p(ws,r,3,"Projected",fill=GR)
p(ws,r,4,f"={IOPS_REF}",fill=GR,fmt="0")
p(ws,r,5,"IOPS",fill=GR)
p(ws,r,8,"Ceiling",fill=GR)
p(ws,r,9,5000,fill=GR,fmt="0")
p(ws,r,10,f'=IF({IOPS_REF}/5000>0.9,"RED",IF({IOPS_REF}/5000>0.7,"AMBER","OK")) & " (" & TEXT({IOPS_REF}/5000,"0%") & ")"')
R["iops_row"] = r
r+=1

# Mongo node sizing note (self-host only; Enterprise SaaS → 0 nodes on our cluster)
p(ws,r,1,"  Mongo node sizing",font=B)
p(ws,r,3,"Pool",fill=GR); p(ws,r,4,f"={MONGO_VM}",fill=GR)
p(ws,r,5,"Nodes",fill=GR); p(ws,r,6,f"={MGN_SH}",fill=GR,fmt="0")
p(ws,r,7,"$/node",fill=GR)
p(ws,r,8,f"={MNP}",fill=GR,fmt='"$"#,##0')
p(ws,r,10,"Dedicated MongoDB pool: 1 node per self-hosted replica (Enterprise SaaS → 0)",font=IT)
R["db_vm_row"] = r
r+=1

# Kafka partition scaling
rt_pods_ref = f"B{service_rows_by_name['runtime']}" if "runtime" in service_rows_by_name else "2"
p(ws,r,1,"  Kafka partitions",font=B)
p(ws,r,3,"Current",fill=GR); p(ws,r,4,"3 per topic",fill=GR)
p(ws,r,5,"Required",fill=GR)
p(ws,r,6,f"=MAX({rt_pods_ref},3)",fill=GR,fmt="0")
p(ws,r,7,"Status",fill=GR)
p(ws,r,8,f'=IF({rt_pods_ref}>3,"INCREASE to "&{rt_pods_ref},"OK (3≥pods)")',fill=GR)
p(ws,r,10,"partitions >= consumer pods for parallel consumption",font=IT)
r+=1

# Restate RocksDB journal
p(ws,r,1,"  Restate RocksDB",font=B)
p(ws,r,3,"PVC",fill=GR); p(ws,r,4,"10Gi",fill=GR)
p(ws,r,5,"Used",fill=GR); p(ws,r,6,"<1%",fill=GR)
p(ws,r,7,"Topology",fill=GR); p(ws,r,8,"1 leader (standalone)",fill=GR)
p(ws,r,10,"Scale: add followers for HA, increase PVC for journal growth",font=IT)
r+=1

# ── Tier 6: Supporting ──
r+=1; sec(ws,r,"SUPPORTING SERVICES",fill=T5,ncols=13); r+=1
hdr(ws,r,["Service","Min","Max (HPA)","CPU Req","CPU Lim","Mem Req Gi","Mem Lim Gi","Total CPU","Total Mem Gi","Pool","VM Type","Throughput/pod","Notes"]); r+=1

gpu_replica_refs = []
for svc in tiers.get("supporting",[]):
    for c0 in svc["containers"][:1]:
        pool = svc_pool.get(svc["name"],"user")
        vm = pool_vm.get(pool,"—")
        cpu_req = c0.get("cpuReq","—")
        mem_req = c0.get("memReq","—")
        max_rep = svc.get("max_replicas", svc.get("replicas", 1))
        p(ws,r,1,svc["name"])
        # Formula-driven replicas (col 2 = Min, col 3 = Max HPA)
        # Non-prod: 1 replica for all supporting services (no HA needed)
        # GPU services: 0 when disabled regardless of HA mode
        if svc["name"] == "bge-m3":
            ha_formula = f'IF(AND({BGE_EN}="Yes",{BGE_MODE}="GPU"),IF({HA_REF}="No",1,MIN(C{r},MAX({ZONES},CEILING(({ST}*{ER}+{IT_REF}*50/60)/15/{U},1)))),0)'
            p(ws,r,2,f'={ha_formula}',fill=GR,fmt="0")
        elif svc["name"] == "bge-m3-cpu":
            ha_formula = f'IF(AND({BGE_EN}="Yes",{BGE_MODE}="CPU"),IF({HA_REF}="No",1,MIN(C{r},MAX({ZONES},CEILING(({ST}*{ER}+{IT_REF}*50/60)/4.4/{U},1)))),0)'
            p(ws,r,2,f'={ha_formula}',fill=GR,fmt="0")
        elif svc["name"] == "docling":
            ha_formula = f'IF(AND({KNOWLEDGE_EN}="Yes",{DOC_MODE}="GPU"),IF({HA_REF}="No",1,MIN(C{r},MAX({ZONES},CEILING({IT_REF}/43/{U},1)))),0)'
            p(ws,r,2,f'={ha_formula}',fill=GR,fmt="0")
        elif svc["name"] == "docling-cpu":
            ha_formula = f'IF(AND({KNOWLEDGE_EN}="Yes",{DOC_MODE}="CPU"),IF({HA_REF}="No",1,MIN(C{r},MAX({ZONES},CEILING({IT_REF}/5/{U},1)))),0)'
            p(ws,r,2,f'={ha_formula}',fill=GR,fmt="0")
        else:
            cat_replicas = int(svc["replicas"])
            ha_formula = f"MIN(C{r},MAX({ZONES},{cat_replicas}))"
            p(ws,r,2,f'=IF({HA_REF}="No",1,{ha_formula})',fill=GR,fmt="0")
        # Col 3: Max (HPA) — 0 for disabled GPU services, else editable max
        if svc["name"] in ("bge-m3", "bge-m3-cpu"):
            p(ws,r,3,f'=IF({BGE_EN}="Yes",{max_rep},0)',fill=YE,fmt="0")
        elif svc["name"] in ("docling", "docling-cpu"):
            p(ws,r,3,f'=IF({KNOWLEDGE_EN}="Yes",{max_rep},0)',fill=YE,fmt="0")
        else:
            p(ws,r,3,max_rep,fill=YE,fmt="0")
        # Col 4-7: CPU/Mem
        p(ws,r,4,cpu_to_cores(cpu_req),fill=GR,fmt="0.000")
        p(ws,r,5,cpu_to_cores(c0.get("cpuLim","—")),fill=GR,fmt="0.000")
        p(ws,r,6,mem_to_gb(mem_req),fill=GR,fmt="0.000")
        p(ws,r,7,mem_to_gb(c0.get("memLim","—")),fill=GR,fmt="0.000")
        # Col 8-9: Total CPU/Mem (replicas × Requests — K8s schedules on requests)
        p(ws,r,8,f"=B{r}*D{r}",fill=GR,fmt="0.00",font=B,al=AL_C)
        p(ws,r,9,f"=B{r}*F{r}",fill=GR,fmt="0.00",font=B,al=AL_C)
        # Col 10-11: Pool, VM
        p(ws,r,10,pool,fill=GR)
        p(ws,r,11,vm_value_for_pool(pool, vm),fill=GR)
        p(ws,r,12,throughput_default("supporting", svc["name"]),fill=YE,fmt="0.0")
        note = ""
        if svc["name"] == "bge-m3": note = "GPU: 15 rps/pod (T4), saturates 30 conc"
        elif svc["name"] == "bge-m3-cpu": note = "CPU: 4.4 rps/pod (4-core), saturates 10 conc"
        elif svc["name"] == "docling": note = "GPU: 43 files/min/pod (floor, not saturated)"
        elif svc["name"] == "docling-cpu": note = "CPU: not load-tested"
        elif "preprocessing" in svc["name"]: note = "Document chunking/OCR"
        elif "ingress" in svc["name"]: note = "NGINX ingress controller"
        elif "livekit" in svc["name"]: note = "WebRTC/voice server"
        elif "multimodal" in svc["name"]: note = "Image/file processing"
        elif "codetool" in svc["name"]: note = "Sandboxed code execution"
        elif "crawler" in svc["name"]: note = "Web crawler worker"
        p(ws,r,13,note,font=IT)
        add_capacity_item("supporting", svc["name"], pool, vm, f"B{r}", cpu_req, mem_req, "catalog replicas", cpu_ref=f"D{r}", mem_ref=f"F{r}")
        if pool == "gpu":
            gpu_replica_refs.append(f"B{r}")
        r+=1
if not tiers.get("supporting",[]):
    p(ws,r,1,"No supporting services configured",font=IT,fill=GR)
    p(ws,r,13,"Add supporting services to default_service_catalog()",font=IT)
    r+=1

gpu_svcs = [s for s in tiers.get("supporting",[]) if svc_pool.get(s["name"]) == "gpu"]
gpu_nodes_ref = "+".join(gpu_replica_refs) if gpu_replica_refs else "0"
R["supp_cost"] = r
p(ws,r,1,"  GPU node context",font=B)
p(ws,r,13,f"=({gpu_nodes_ref})*{GNP}",fill=GR,fmt='"$"#,##0')
p(ws,r,10,f"{len(gpu_svcs)} GPU services; GPU pods included in node-pool capacity cost",font=IT); r+=1

# ── Tier 7: Operators ──
r+=1; sec(ws,r,"OPERATORS (fixed)",fill=T6,ncols=13); r+=1
hdr(ws,r,["Service","Replicas","","CPU Req","CPU Lim","Mem Req Gi","Mem Lim Gi","Total CPU","Total Mem Gi","Pool","VM Type","Throughput/pod","Notes"]); r+=1

for svc in tiers.get("operator",[]):
    for c0 in svc["containers"][:1]:
        pool = "operator"
        vm = pool_vm.get(pool,"—")
        cpu_req = c0.get("cpuReq","—")
        mem_req = c0.get("memReq","—")
        p(ws,r,1,svc["name"])
        # Operators: non-prod = 1 (leader-elected, no HA needed); prod = MAX(zones, catalog)
        op_replicas = svc["replicas"]
        if op_replicas > 1:
            p(ws,r,2,f'=IF({HA_REF}="No",1,MAX({ZONES},{op_replicas}))',fill=GR,fmt="0")
        else:
            p(ws,r,2,f'=IF({HA_REF}="No",1,{op_replicas})',fill=GR,fmt="0")
        p(ws,r,4,cpu_to_cores(cpu_req),fill=GR,fmt="0.000")
        p(ws,r,5,cpu_to_cores(c0.get("cpuLim","—")),fill=GR,fmt="0.000")
        p(ws,r,6,mem_to_gb(mem_req),fill=GR,fmt="0.000")
        p(ws,r,7,mem_to_gb(c0.get("memLim","—")),fill=GR,fmt="0.000")
        p(ws,r,8,f"=B{r}*D{r}",fill=GR,fmt="0.00",font=B,al=AL_C)
        p(ws,r,9,f"=B{r}*F{r}",fill=GR,fmt="0.00",font=B,al=AL_C)
        p(ws,r,10,pool,fill=GR)
        p(ws,r,11,vm_value_for_pool(pool, vm),fill=GR)
        p(ws,r,12,throughput_default("operator", svc["name"]),fill=YE,fmt="0.0")
        note = ""
        if "strimzi" in svc["name"]: note = "Kafka operator"
        elif "mongodb-kubernetes" in svc["name"]: note = "MongoDB operator"
        elif "external-secret" in svc["name"]: note = "Secret sync from Azure KV"
        elif "goldilocks" in svc["name"]: note = "VPA recommendations"
        elif "kafka-entity" in svc["name"]: note = "Kafka topic/user mgmt"
        p(ws,r,13,note,font=IT)
        add_capacity_item("operator", svc["name"], pool, vm, f"B{r}", cpu_req, mem_req, "fixed", cpu_ref=f"D{r}", mem_ref=f"F{r}")
        r+=1
if not tiers.get("operator",[]):
    p(ws,r,1,"No operators configured",font=IT,fill=GR)
    p(ws,r,13,"Operators are fixed topology components in the built-in catalog",font=IT)
    r+=1

# ── Node Pools Summary ──
r+=1; sec(ws,r,"NODE POOLS",fill=HB,ncols=13); r+=1
hdr(ws,r,["Pool","VM Type","CPU/node","Mem/node Gi","CPU Req","Mem Req Gi","Pod Nodes","CPU Nodes","Mem Nodes","Req Nodes","Unit $/mo","Monthly $","Notes"]); r+=1

node_pool_rows = {}
node_pool_cost_rows = []
for pn, pi in node_pools.items():
    # VM type reference — dynamic from input selectors
    if pn == "database (mongo)":
        vm = f"={MONGO_VM}"
        cpu_per_node_ref = MONGO_VM_CPU
        mem_per_node_ref = MONGO_VM_MEM
        unit_cost_ref = MNP
        unit_note = "Mongo node price (from selected VM)"
        note = "Dedicated MongoDB pool; 1 node per replica with headroom"
    elif pn == "database (non-mongo)":
        vm = f"={DB_VM}"
        cpu_per_node_ref = DB_VM_CPU
        mem_per_node_ref = DB_VM_MEM
        unit_cost_ref = DNP
        unit_note = "Database node price (from selected VM)"
        note = "Redis/OpenSearch/Neo4j/RocksDB; zone-aware min nodes"
    elif pn == "user":
        vm = f"={USER_VM}"
        cpu_per_node_ref = USER_VM_CPU
        mem_per_node_ref = USER_VM_MEM
        unit_cost_ref = NP
        unit_note = "User node price (from selected VM)"
        note = "App workloads; zone-aware min nodes"
    elif pn == "operator":
        vm = f"={OPERATOR_VM}"
        cpu_per_node_ref = OP_VM_CPU
        mem_per_node_ref = OP_VM_MEM
        unit_cost_ref = ONP
        unit_note = "Operator node price (from selected VM)"
        note = "Small VMs for controllers/operators; zone-spread"
    elif pn == "system":
        vm = f"={SYSTEM_VM}"
        cpu_per_node_ref = SYS_VM_CPU
        mem_per_node_ref = SYS_VM_MEM
        unit_cost_ref = SNP
        unit_note = "System node price (from selected VM)"
        note = "ClickHouse + Kafka; zone-aware min nodes"
    elif pn == "gpu":
        vm = f"={GPU_VM}"
        cpu_per_node_ref = None  # GPU nodes: 1 pod per node, not CPU/mem driven
        mem_per_node_ref = None
        unit_cost_ref = GNP
        unit_note = "GPU node price"
        note = "T4 GPU nodes (1 pod/node); zone-aware min nodes"
    else:
        vm = pool_vm.get(pn, "—")
        cpu_per_node_ref = None
        mem_per_node_ref = None
        unit_cost_ref = NP
        unit_note = "Fallback"
        note = ""

    p(ws,r,1,pn,font=B)
    node_pool_rows[pn] = r
    p(ws,r,2,vm,fill=GR)

    # CPU/node and Mem/node — reference dynamic VM specs from input section
    if cpu_per_node_ref:
        p(ws,r,3,f"={cpu_per_node_ref}",fill=GR,fmt="0.0")
    else:
        p(ws,r,3,4,fill=GR,fmt="0.0")  # GPU: fixed 4 vCPU
    if mem_per_node_ref:
        p(ws,r,4,f"={mem_per_node_ref}",fill=GR,fmt="0.0")
    elif pn == "gpu":
        p(ws,r,4,csp_value(CSP,16,28,15),fill=GR,fmt="0.0")  # GPU mem: AWS g4dn=16, Azure NC4as=28, GCP n1-std-4=15
    else:
        p(ws,r,4,0,fill=GR,fmt="0.0")

    pool_cpu_formula = "+".join(item["cpu_formula"] for item in capacity_items if item["pool"] == pn) or "0"
    pool_mem_formula = "+".join(item["mem_formula"] for item in capacity_items if item["pool"] == pn) or "0"
    pool_replicas_formula = "+".join(
        item["replicas_ref"]
        for item in capacity_items
        if item["pool"] == pn
    ) or "0"
    pool_pod_nodes_formula = f"CEILING(({pool_replicas_formula})/{PPN},1)" if pool_replicas_formula != "0" else "0"
    p(ws,r,5,f"={pool_cpu_formula}",fill=GR,fmt="0.00",font=B,al=AL_C)
    p(ws,r,6,f"={pool_mem_formula}",fill=GR,fmt="0.00",font=B,al=AL_C)
    p(ws,r,7,f"={pool_pod_nodes_formula}",fill=GR,fmt="0",font=B,al=AL_C)
    p(ws,r,8,f'=IF(C{r}>0,CEILING(E{r}/(C{r}*{U}),1),0)',fill=GR,fmt="0",font=B,al=AL_C)
    p(ws,r,9,f'=IF(D{r}>0,CEILING(F{r}/(D{r}*{U}),1),0)',fill=GR,fmt="0",font=B,al=AL_C)
    # Req Nodes: pool-specific logic
    if pn == "gpu":
        # GPU pool = 0 if no GPU pods; otherwise zone-aware
        p(ws,r,10,f'=IF(G{r}=0,0,IF({HA_REF}="No",MAX(1,G{r},H{r},I{r}),MAX({ZONES},G{r},H{r},I{r})))',fill=GR,fmt="0",font=B,al=AL_C)
    elif pn == "database (mongo)":
        # MongoDB: 1 dedicated node per self-hosted replica (Enterprise SaaS / N/A → 0 nodes).
        p(ws,r,10,f'={MGN_SH}',fill=GR,fmt="0",font=B,al=AL_C)
    else:
        p(ws,r,10,f'=IF({HA_REF}="No",MAX(1,G{r},H{r},I{r}),MAX({ZONES},G{r},H{r},I{r}))',fill=GR,fmt="0",font=B,al=AL_C)
    p(ws,r,11,f"={unit_cost_ref}",fill=GR,fmt='"$"#,##0',font=B,al=AL_R)
    p(ws,r,12,f"=J{r}*K{r}",fill=GR,fmt='"$"#,##0',font=B,al=AL_R)
    p(ws,r,13,f"{note}; {unit_note}",font=IT)
    node_pool_cost_rows.append(r)
    r+=1

R["node_pool_cost"] = r
p(ws,r,1,"  Total node pool cost",font=B)
node_pool_cost = "+".join(f"L{x}" for x in node_pool_cost_rows) if node_pool_cost_rows else "0"
p(ws,r,10,f"=SUM({','.join(f'J{x}' for x in node_pool_cost_rows)})" if node_pool_cost_rows else 0,fill=GR,fmt="0",font=B,al=AL_C)
p(ws,r,12,f"={node_pool_cost}",fill=GR,fmt='"$"#,##0',font=B,al=AL_R)
p(ws,r,13,"Source of truth for workload hardware in grand total",font=IT)
r+=1

# ── Fixed Infrastructure Costs ──
r+=1; sec(ws,r,"FIXED INFRASTRUCTURE COSTS",fill=HB,ncols=12); r+=1
hdr(ws,r,["Item","SKU / Basis","Qty","Unit $/mo","Monthly $","Notes","","","","","",""]); r+=1

fixed_infra_rows = []
fixed_cost_items = [
    ("LoadBalancer / WAF / Advanced Shield", csp_text(CSP,"ALB + AWS WAF + Shield Advanced","Application Gateway WAF v2 + DDoS Protection","Cloud LB + Cloud Armor + DDoS"), 1, 3500, "LB + WAF + Advanced Shield combined; editable"),
    ("Object Storage", csp_text(CSP,"S3 Standard / external object store","Blob Hot LRS / external object store","GCS Standard / external object store"), f"={OST}*1099.51", f"={OSP}", "Editable TiB → GB (1 TiB = 1099.51 GB); provider $/GB-month from CSP selector"),
    ("Block Storage", csp_text(CSP,"EBS gp3 128 GiB supplemental","P10 ZRS supplemental","PD-SSD 128 GiB supplemental"), 1, csp_value(CSP,AWS_GP3_128GB_MONTHLY,P10_ZRS_SUPPLEMENTAL_DISK_MONTHLY,GCP_PD_SSD_128GB_MONTHLY), "Additional managed disk outside service PVCs"),
    ("Container Registry", csp_text(CSP,"ECR private registry baseline","ACR Premium registry unit","Artifact Registry baseline"), 1, csp_value(CSP,AWS_ECR_BASELINE_MONTHLY,ACR_PREMIUM_MONTHLY,GCP_AR_BASELINE_MONTHLY), "Provider container registry baseline"),
]
for item, sku, qty, unit_cost, note in fixed_cost_items:
    p(ws,r,1,item,font=B,al=AL_L)
    p(ws,r,2,sku,fill=GR,al=AL_L)
    qty_fill = GR if isinstance(qty, str) and qty.startswith("=") else YE
    unit_fill = GR if isinstance(unit_cost, str) and unit_cost.startswith("=") else YE
    p(ws,r,3,qty,fill=qty_fill,fmt="0",font=B,al=AL_C)
    p(ws,r,4,unit_cost,fill=unit_fill,fmt='"$"#,##0',font=B,al=AL_R)
    p(ws,r,5,f"=C{r}*D{r}",fill=GR,fmt='"$"#,##0',font=B,al=AL_R)
    p(ws,r,6,note,font=IT,al=AL_L)
    fixed_infra_rows.append(r)
    r+=1

R["fixed_infra_cost"] = r
p(ws,r,1,"  Fixed infrastructure cost",font=B)
fixed_infra_cost = "+".join(f"E{x}" for x in fixed_infra_rows) if fixed_infra_rows else "0"
p(ws,r,5,f"={fixed_infra_cost}",fill=GR,fmt='"$"#,##0',font=B,al=AL_R)
r+=1

# ── PVC Storage Costs (Database Disk Per Node) ──
r+=1; sec(ws,r,"PVC STORAGE COSTS (per-node disk × replicas)",fill=HB,ncols=12); r+=1
hdr(ws,r,["Data Store","Disk/Node GiB","Replicas","Total GiB","$/GiB-month","Monthly $","Notes","","","","",""]); r+=1

DISK_PRICE_REF = f"B{R['disk_price_gb']}"
# Replicas use the self-host (_SH) refs so Enterprise (SaaS) and N/A both contribute 0 GiB / $0
# to our PVC bill. Disk-per-node remains the configured value (informational for Enterprise).
pvc_storage_rows = []
pvc_items = [
    ("MongoDB",    f"C{R['mongo_replicas']}",      MGN_SH,     "Data + indexes + oplog + headroom (self-host only)"),
    ("OpenSearch", f"C{R['opensearch_replicas']}", OSRN_SH,    "Indices + segment merge headroom (self-host only)"),
    ("Neo4j",      f"C{R['neo4j_replicas']}",      N4JN_SH,    "Graph store + transaction logs (self-host only)"),
    ("Redis",      f"C{R['redis_replicas']}",      RDTOTAL_SH, "AOF persistence + RDB snapshots (self-host only)"),
    ("ClickHouse", f"C{R['clickhouse_replicas']}", CHTOTAL_SH, "Column store + merge tree parts (self-host only)"),
    ("Kafka",      f"C{R['kafka_replicas']}",      KAFKAN_SH,  "Log segments + retention window (self-host only)"),
    ("RocksDB",    f"C{R['rocksdb_replicas']}",    ROCKSN_SH,  "Embedded KV store data (self-host only)"),
]
for store_name, disk_ref, replicas_ref, note in pvc_items:
    p(ws,r,1,store_name,font=B,al=AL_L)
    p(ws,r,2,f"={disk_ref}",fill=GR,fmt="0",font=B,al=AL_C)
    p(ws,r,3,f"={replicas_ref}",fill=GR,fmt="0",font=B,al=AL_C)  # Always follow user input
    p(ws,r,4,f"=B{r}*C{r}",fill=GR,fmt="#,##0",font=B,al=AL_C)
    p(ws,r,5,f"={DISK_PRICE_REF}",fill=GR,fmt='"$"0.000',al=AL_R)
    p(ws,r,6,f"=D{r}*E{r}",fill=GR,fmt='"$"#,##0',font=B,al=AL_R)
    p(ws,r,7,note,font=IT,al=AL_L)
    pvc_storage_rows.append(r)
    r+=1

R["pvc_storage_cost"] = r
p(ws,r,1,"  PVC storage cost",font=B)
pvc_storage_total = "+".join(f"F{x}" for x in pvc_storage_rows) if pvc_storage_rows else "0"
p(ws,r,6,f"={pvc_storage_total}",fill=GR,fmt='"$"#,##0',font=B,al=AL_R)
p(ws,r,7,"Total monthly block storage for all data store PVCs",font=IT)
r+=1

# ── Software / Licensing Costs ──
r+=1; sec(ws,r,"SOFTWARE / LICENSING COSTS",fill=HB,ncols=10); r+=1
hdr(ws,r,["Component","Basis","Qty","Monthly Unit $","Monthly $","Notes","","","",""]); r+=1

TOTAL_NODES_REF = f"J{R['node_pool_cost']}"  # sum of Req Nodes across all pools

software_rows = []
# Licensing Qty per datastore:
#   - Enterprise (vendor SaaS) → MAX(1, configured instances) — vendor bills per provisioned instance.
#   - Self-Hosted              → 0 (no commercial license — OSS / self-managed).
#   - N/A                      → 0.
software_items = [
    ("MongoDB",    "instances purchased", f'=IF({MGN_DEPLOY}="Enterprise",MAX(1,{MGN}),0)',         args.mongo_license_per_node_mo,      "Monthly $/instance; charged only for Enterprise SaaS (e.g. Atlas); Self-Hosted/N/A → 0"),
    ("Redis",      "instances purchased", f'=IF({RDTOTAL_DEPLOY}="Enterprise",MAX(1,{RDTOTAL}),0)', args.redis_license_per_node_mo,      "Monthly $/instance; charged only for Enterprise SaaS (e.g. Redis Cloud); Self-Hosted/N/A → 0"),
    ("ClickHouse", "instances purchased", f'=IF({CHTOTAL_DEPLOY}="Enterprise",MAX(1,{CHTOTAL}),0)', args.clickhouse_license_per_node_mo, "Monthly $/instance; charged only for Enterprise SaaS (e.g. ClickHouse Cloud); Self-Hosted/N/A → 0"),
    ("Kafka",      "instances purchased", f'=IF({KAFKAN_DEPLOY}="Enterprise",MAX(1,{KAFKAN}),0)',   args.kafka_license_per_node_mo,      "Monthly $/instance; charged only for Enterprise SaaS (e.g. Confluent Cloud); Self-Hosted/N/A → 0"),
    ("Groundcover", "per base node", f"={TOTAL_NODES_REF}", 17, "Groundcover monitoring agent $17/node-month"),
    ("CrowdStrike", "per base node", f"={TOTAL_NODES_REF}", 10, "CrowdStrike Falcon agent $10/node-month; editable"),
]
for component, basis, qty, unit_cost, note in software_items:
    p(ws,r,1,component,font=B)
    p(ws,r,2,basis,fill=GR)
    p(ws,r,3,qty,fill=GR,fmt="0",font=B,al=AL_C)
    p(ws,r,4,unit_cost,fill=YE,fmt='"$"#,##0',font=B,al=AL_R)
    p(ws,r,5,f"=C{r}*D{r}",fill=GR,fmt='"$"#,##0',font=B,al=AL_R)
    p(ws,r,6,note,font=IT)
    software_rows.append(r)
    r+=1

R["software_cost"] = r
p(ws,r,1,"  Software/licensing cost",font=B)
software_total = "+".join(f"E{x}" for x in software_rows) if software_rows else "0"
p(ws,r,5,f"={software_total}",fill=GR,fmt='"$"#,##0',font=B,al=AL_R)
r+=1

# ── Total Cost ──
r+=1; sec(ws,r,"TOTAL COST SUMMARY",fill=HB,ncols=12); r+=1
hdr(ws,r,["Cost Bucket","Monthly $","Yearly $","Included in Total","Notes","","","","","","",""]); r+=1

cost_rows = [
    ("Node pool MSRP", f"L{R['node_pool_cost']}", False, "Reference before discount"),
    ("Hardware discount", f"-L{R['node_pool_cost']}*{HD}", False, "Shown separately; net hardware is included"),
    ("Net node pool hardware", f"L{R['node_pool_cost']}*(1-{HD})", True, "Discounted compute capacity"),
    ("Fixed Infrastructure", f"E{R['fixed_infra_cost']}", True, "WAF, object storage, supplemental disk, ACR"),
    ("PVC Storage", f"F{R['pvc_storage_cost']}", True, "Database disk (per-node × replicas)"),
    ("Software / licensing", f"E{R['software_cost']}", True, "Monthly conversion of applicable licenses"),
    ("Support", f"L{R['node_pool_cost']}*{SR}", True, "Monthly support on node-pool MSRP"),
]
monthly_total_terms = []
yearly_total_terms = []

for tier_name, cost_ref, include_in_total, note in cost_rows:
    p(ws,r,1,tier_name,font=B if include_in_total else None)
    p(ws,r,2,f"={cost_ref}",fill=GR,fmt='"$"#,##0',font=B,al=AL_R)
    p(ws,r,3,f"=B{r}*12",fill=GR,fmt='"$"#,##0',font=B,al=AL_R)
    p(ws,r,4,"Yes" if include_in_total else "No",fill=GR,al=AL_C)
    p(ws,r,5,note,font=IT)
    if include_in_total:
        monthly_total_terms.append(f"B{r}")
        yearly_total_terms.append(f"C{r}")
    r+=1

r+=1
R["total_month"] = r
p(ws,r,1,"GRAND TOTAL",font=B14)
monthly_costs = "+".join(monthly_total_terms)
p(ws,r,2,f"={monthly_costs}",fill=GN,fmt='"$"#,##0',font=B14,al=AL_R)
yearly_costs = "+".join(yearly_total_terms)
p(ws,r,3,f"={yearly_costs}",fill=GN,fmt='"$"#,##0',font=B14,al=AL_R)
p(ws,r,4,"Final monthly and annualized run-rate",font=IT); r+=1

p(ws,r,1,"Cost per 1M runtime messages")
p(ws,r,2,f'=IF({T}=0,"N/A",B{R["total_month"]}/({T}*3600*24*30/1000000))',fill=GR,fmt='"$"#,##0.00',font=B,al=AL_R)
p(ws,r,3,f'=IF({T}=0,"N/A",C{R["total_month"]}/({T}*3600*24*365/1000000))',fill=GR,fmt='"$"#,##0.00',font=B,al=AL_R)
p(ws,r,4,"Monthly and annualized basis",font=IT); r+=1



# ═══ TAB 3: REVIEW ═══════════════════════════════════════════════════════════
ws_r = wb.create_sheet("Review")
wid(ws_r, [18, 34, 10, 12, 14, 12, 14, 12, 22, 52])

r = 1
ws_r.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
p(ws_r,r,1,"Sizing Review — Services and Node Pools",font=B14, fill=PatternFill("solid", fgColor="EAF2F8")); r+=2

sec(ws_r,r,"SERVICE REVIEW",fill=HB,ncols=10); r+=1
hdr(ws_r,r,["Tier","Service","Replicas","CPU Req/pod","Mem Req/pod","Total CPU","Total Mem Gi","Pool","VM Type","Review"]); r+=1

service_sections = {
    "RUNTIME (chat agents)",
    "SEARCH AI",
    "WORKFLOW",
    "PIPELINE ENGINE",
    "DATA STORES",
    "SUPPORTING SERVICES",
    "OPERATORS (fixed)",
}
stop_sections = {
    "NODE POOLS",
    "FIXED INFRASTRUCTURE COSTS",
    "SOFTWARE / LICENSING COSTS",
    "TOTAL COST SUMMARY",
}
section = None
for source_row in range(1, ws.max_row + 1):
    value = ws.cell(source_row, 1).value
    if isinstance(value, str) and value in service_sections:
        section = value
        continue
    if isinstance(value, str) and value in stop_sections:
        section = None
        continue
    if (
        not section
        or not isinstance(value, str)
        or value in ("Service", "No services configured", "No data services configured", "No supporting services configured", "No operators configured")
        or value.startswith("  ")
    ):
        continue

    ref = f"'Sizing & Cost'!"
    p(ws_r,r,1,section.replace(" (chat agents)", ""),fill=GR)
    p(ws_r,r,2,f"={ref}A{source_row}",fill=GR)
    p(ws_r,r,3,f"={ref}B{source_row}",fill=GR,fmt="0",font=B,al=AL_C)
    p(ws_r,r,4,f"={ref}C{source_row}",fill=GR)
    p(ws_r,r,5,f"={ref}E{source_row}",fill=GR)
    p(ws_r,r,6,f"={ref}G{source_row}",fill=GR,fmt="0.00",font=B,al=AL_C)
    p(ws_r,r,7,f"={ref}H{source_row}",fill=GR,fmt="0.00",font=B,al=AL_C)
    p(ws_r,r,8,f"={ref}I{source_row}",fill=GR)
    p(ws_r,r,9,f"={ref}J{source_row}",fill=GR)
    service_name = value.lower()
    if section in ("DATA STORES", "OPERATORS (fixed)"):
        if "clickhouse-shard" in service_name:
            review_note = "OK: ClickHouse shards and replicas are modeled on the D8s v5 system pool."
        elif service_name == "redis-master":
            review_note = "OK: Redis master count uses shard input; validate Redis Cluster slots and failover policy."
        elif service_name == "redis-replicas":
            review_note = "OK: Redis replica count uses shards × replicas/shard."
        elif service_name == "mongodb":
            review_note = "OK: MongoDB uses 3-node D48s v5 sizing; database nodes are sized by CPU and memory."
        elif section == "OPERATORS (fixed)":
            review_note = "Fixed operator component; keep single replica unless operator availability requires HA."
        else:
            review_note = "Stateful/fixed component; HA is topology-specific, not forced by app pod minimum."
    else:
        review_note = "OK: HA minimum app pods enforced; verify PDB/HPA max in deploy repo before production rollout."
    p(ws_r,r,10,review_note,font=IT)
    r+=1

r+=1
sec(ws_r,r,"NODE POOL REVIEW",fill=HB,ncols=10); r+=1
hdr(ws_r,r,["Pool","VM Type","CPU/node","Mem/node Gi","CPU Req","Mem Req Gi","Pod Nodes","Req Nodes","Monthly $","Review"]); r+=1

valid_pools = ("database (mongo)", "database (non-mongo)", "gpu", "operator", "system", "user")
for source_row in range(1, ws.max_row + 1):
    pool_name = ws.cell(source_row, 1).value
    if pool_name not in valid_pools:
        continue
    ref = f"'Sizing & Cost'!"
    p(ws_r,r,1,f"={ref}A{source_row}",fill=GR,font=B)
    p(ws_r,r,2,f"={ref}B{source_row}",fill=GR)
    p(ws_r,r,3,f"={ref}C{source_row}",fill=GR,fmt="0.0",font=B,al=AL_C)
    p(ws_r,r,4,f"={ref}D{source_row}",fill=GR,fmt="0.0",font=B,al=AL_C)
    p(ws_r,r,5,f"={ref}E{source_row}",fill=GR,fmt="0.00",font=B,al=AL_C)
    p(ws_r,r,6,f"={ref}F{source_row}",fill=GR,fmt="0.00",font=B,al=AL_C)
    p(ws_r,r,7,f"={ref}G{source_row}",fill=GR,fmt="0",font=B,al=AL_C)
    p(ws_r,r,8,f"={ref}J{source_row}",fill=GR,fmt="0",font=B,al=AL_C)
    p(ws_r,r,9,f"={ref}L{source_row}",fill=GR,fmt='"$"#,##0',font=B,al=AL_R)
    if pool_name == "user":
        note = "CPU-driven after HA floor; user pool should scale to required nodes."
    elif pool_name == "gpu":
        note = "GPU app pod count is used as a node floor; 1 GPU pod per GPU-node."
    elif pool_name == "database (mongo)":
        note = "Dedicated MongoDB pool; 1 node per replica with headroom."
    elif pool_name == "database (non-mongo)":
        note = "Redis/OpenSearch/Neo4j/RocksDB; node count must fit DB CPU/memory."
    elif pool_name == "operator":
        note = "Small VMs for controllers/operators; spread across zones."
    else:
        note = "System pool: ClickHouse + Kafka capacity."
    p(ws_r,r,10,note,font=IT)
    r+=1

def beautify_workbook(wb):
    """Apply workbook-wide presentation polish after all cells are written."""
    CENTER_COLUMNS = {2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16}
    TEXT_COLUMNS = {1}
    EDITABLE_FILL_RGBS = {"00FFF2CC", "00D9EAF7"}
    SECTION_TITLES = {
        "GLOBAL INPUTS",
        "RUNTIME (CHAT AGENTS)",
        "SEARCH AI",
        "WORKFLOW",
        "PIPELINE ENGINE",
        "DATA STORES",
        "SUPPORTING SERVICES",
        "OPERATORS (FIXED)",
        "NODE POOLS",
        "FIXED INFRASTRUCTURE COSTS",
        "SOFTWARE / LICENSING COSTS",
        "TOTAL COST SUMMARY",
    }
    TABLE_HEADERS = {
        "Parameter",
        "Service",
        "Pool",
        "Item",
        "Component",
        "Cost Bucket",
    }

    def is_money_cell(cell):
        return isinstance(cell.number_format, str) and '"$"' in cell.number_format

    def is_percent_cell(cell):
        return isinstance(cell.number_format, str) and "%" in cell.number_format

    def is_editable_cell(cell):
        return cell.fill.fgColor.rgb in EDITABLE_FILL_RGBS

    def is_section_row(row_cells):
        first_value = row_cells[0].value
        if not isinstance(first_value, str):
            return False
        return first_value.upper() in SECTION_TITLES

    def is_table_header_row(row_cells):
        first_value = row_cells[0].value
        return isinstance(first_value, str) and first_value in TABLE_HEADERS

    def merge_section_rows(ws_obj):
        for row in range(1, ws_obj.max_row + 1):
            row_cells = [ws_obj.cell(row=row, column=col) for col in range(1, ws_obj.max_column + 1)]
            if not is_section_row(row_cells):
                continue
            end_col = ws_obj.max_column
            for merged_range in ws_obj.merged_cells.ranges:
                if row in range(merged_range.min_row, merged_range.max_row + 1):
                    break
            else:
                ws_obj.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_col)

    def estimated_cell_width(cell):
        value = cell.value
        if value in (None, ""):
            return 0
        if isinstance(value, str) and value.startswith("="):
            if cell.number_format.startswith('"$"'):
                return 15
            return 11
        text = str(value)
        if cell.number_format.startswith('"$"'):
            return max(12, min(16, len(text) + 5))
        return max(len(part) for part in text.splitlines()) + 2

    def autosize_columns(ws_obj, bounds):
        for col in range(1, ws_obj.max_column + 1):
            letter = get_column_letter(col)
            min_width, max_width = bounds.get(letter, (9, 34))
            measured = min_width
            for row in range(1, ws_obj.max_row + 1):
                measured = max(measured, estimated_cell_width(ws_obj.cell(row=row, column=col)))
            ws_obj.column_dimensions[letter].width = min(max(measured, min_width), max_width)

    def merge_long_text_cells(ws_obj):
        merged_cells = {
            coord
            for merged_range in ws_obj.merged_cells.ranges
            for row in ws_obj[merged_range.coord]
            for cell in row
            for coord in [cell.coordinate]
        }
        for row in range(1, ws_obj.max_row + 1):
            for col in range(1, ws_obj.max_column):
                cell = ws_obj.cell(row=row, column=col)
                value = cell.value
                if cell.coordinate in merged_cells:
                    continue
                if not isinstance(value, str) or value.startswith("=") or len(value) < 28:
                    continue
                end_col = col
                for next_col in range(col + 1, ws_obj.max_column + 1):
                    next_cell = ws_obj.cell(row=row, column=next_col)
                    if next_cell.value not in (None, "") or next_cell.coordinate in merged_cells:
                        break
                    end_col = next_col
                if end_col <= col:
                    continue
                ws_obj.merge_cells(start_row=row, start_column=col, end_row=row, end_column=end_col)
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                merged_cells.update(
                    ws_obj.cell(row=row, column=merged_col).coordinate
                    for merged_col in range(col, end_col + 1)
                )

    def is_notes_cell(ws_obj, cell):
        for row in range(cell.row - 1, max(0, cell.row - 25), -1):
            value = ws_obj.cell(row=row, column=cell.column).value
            if value == "Notes":
                return True
            if isinstance(value, str) and value.isupper() and cell.column == 1:
                return False
        return False

    tab_colors = {
        "Sizing & Cost": "2F5496",
    }
    for ws_obj in wb.worksheets:
        merge_section_rows(ws_obj)
        ws_obj.sheet_view.showGridLines = False
        ws_obj.freeze_panes = "A5"
        ws_obj.sheet_properties.tabColor = tab_colors.get(ws_obj.title, "A5A5A5")
        ws_obj.sheet_format.defaultRowHeight = 21
        for row_idx in range(1, ws_obj.max_row + 1):
            ws_obj.row_dimensions[row_idx].height = 21
        ws_obj.row_dimensions[1].height = 30
        ws_obj.row_dimensions[2].height = 21
        for row_cells in ws_obj.iter_rows():
            row_has_value = any(cell.value is not None for cell in row_cells)
            if not row_has_value:
                ws_obj.row_dimensions[row_cells[0].row].height = 10
                continue
            first_value = row_cells[0].value
            section_row = is_section_row(row_cells)
            table_header_row = is_table_header_row(row_cells)
            is_cost_row = isinstance(first_value, str) and first_value.strip().lower().endswith("cost")
            is_total_row = isinstance(first_value, str) and first_value.startswith("GRAND TOTAL")
            for cell in row_cells:
                if cell.value is None:
                    continue
                cell.border = SOFT_BD
                if cell.row == 1:
                    cell.border = BD
                    cell.fill = PatternFill("solid", fgColor="EAF2F8")
                    ws_obj.row_dimensions[cell.row].height = 30
                    if cell.column == 1:
                        cell.font = B14
                        cell.alignment = Alignment(horizontal="left", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                elif cell.row == 2:
                    cell.border = SOFT_BD
                    cell.fill = PatternFill("solid", fgColor="F8FBFD")
                    cell.font = IT
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                elif is_total_row:
                    cell.border = BD
                    ws_obj.row_dimensions[cell.row].height = 28
                    if cell.column in (1, 2, 3):
                        cell.font = B14
                    if cell.column == 1:
                        cell.fill = PatternFill("solid", fgColor="EAF2F8")
                        cell.alignment = AL_L
                    elif cell.column in (2, 3):
                        cell.fill = GN
                        cell.alignment = AL_C
                    else:
                        cell.alignment = AL_L
                elif section_row:
                    cell.alignment = AL_L
                    cell.font = HF if cell.fill == HB else SF
                    ws_obj.row_dimensions[cell.row].height = 24
                elif table_header_row:
                    cell.alignment = AL_C
                    cell.fill = HB
                    ws_obj.row_dimensions[cell.row].height = 24
                    if cell.value:
                        cell.font = HF
                elif is_cost_row:
                    cell.alignment = AL_L if cell.column == 1 else AL_C
                    ws_obj.row_dimensions[cell.row].height = 23
                elif is_editable_cell(cell):
                    cell.alignment = AL_C
                    ws_obj.row_dimensions[cell.row].height = max(ws_obj.row_dimensions[cell.row].height or 21, 22)
                elif (
                    isinstance(cell.value, str)
                    and not cell.value.startswith("=")
                    and (len(cell.value) > 28 or is_notes_cell(ws_obj, cell))
                ):
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    ws_obj.row_dimensions[cell.row].height = max(ws_obj.row_dimensions[cell.row].height or 21, 23)
                elif is_money_cell(cell):
                    cell.alignment = AL_C
                elif is_percent_cell(cell):
                    cell.alignment = AL_C
                elif cell.column in TEXT_COLUMNS:
                    cell.alignment = AL_L
                elif cell.column in CENTER_COLUMNS:
                    cell.alignment = AL_C
                else:
                    cell.alignment = AL_L
                cell.protection = Protection(locked=not is_editable_cell(cell))
        merge_long_text_cells(ws_obj)
        for row_cells in ws_obj.iter_rows():
            for cell in row_cells:
                if is_editable_cell(cell):
                    cell.protection = Protection(locked=False)
                else:
                    cell.protection = Protection(locked=True)
        if ws_obj.title == "Sizing & Cost":
            autosize_columns(ws_obj, {
                "A": (30, 46),
                "B": (16, 24),
                "C": (10, 14),
                "D": (10, 14),
                "E": (12, 18),
                "F": (12, 18),
                "G": (11, 14),
                "H": (12, 15),
                "I": (10, 14),
                "J": (18, 26),
                "K": (10, 13),
                "L": (15, 22),
                "M": (10, 13),
                "N": (10, 13),
                "O": (10, 13),
                "P": (10, 13),
            })
        if ws_obj.max_row >= 4 and ws_obj.max_column >= 2:
            ws_obj.auto_filter.ref = ws_obj.dimensions
        ws_obj.protection.sheet = True
        ws_obj.protection.password = "abl"
        ws_obj.protection.selectLockedCells = True
        ws_obj.protection.selectUnlockedCells = True
        ws_obj.protection.formatCells = False
        ws_obj.protection.formatColumns = False
        ws_obj.protection.formatRows = False
        ws_obj.protection.insertRows = False
        ws_obj.protection.deleteRows = False
        ws_obj.protection.sort = False
        ws_obj.protection.autoFilter = False
        ws_obj.protection.pivotTables = False
        ws_obj.protection.objects = True
        ws_obj.protection.scenarios = True
        ws_obj.protection.enable()

# ═══ TAB: THROUGHPUT CALCULATOR ══════════════════════════════════════════════
# Converts sessions/year into required msg/sec throughput.
# Output links to Sizing & Cost via Option C (auto-pull + manual override).
ws_tp = wb.create_sheet("Throughput Calculator")
wid(ws_tp, [32, 14, 10, 28])

tp_r = 1
ws_tp.merge_cells(start_row=tp_r, start_column=1, end_row=tp_r, end_column=4)
p(ws_tp,tp_r,1,"Throughput Calculator",font=B14, fill=PatternFill("solid", fgColor="EAF2F8")); tp_r+=2

sec(ws_tp,tp_r,"INPUTS",fill=HB,ncols=4); tp_r+=1
hdr(ws_tp,tp_r,["Parameter","Value","Unit","Notes"]); tp_r+=1

# Channel type reference (for user info)
TP = {}  # row registry

TP["channel_type"] = tp_r
p(ws_tp,tp_r,1,"Channel type",font=B); p(ws_tp,tp_r,2,"Webhook",fill=YE)
p(ws_tp,tp_r,3,"type"); p(ws_tp,tp_r,4,"Socket=4s, Webhook=5s, Voice=6s think time",font=IT); tp_r+=1

TP["think_time"] = tp_r
p(ws_tp,tp_r,1,"Think time (delay between messages)",font=B); p(ws_tp,tp_r,2,5.0,fill=YE,fmt="0.0")
p(ws_tp,tp_r,3,"sec"); p(ws_tp,tp_r,4,"User pause before next message (channel dependent)",font=IT); tp_r+=1

TP["session_duration"] = tp_r
p(ws_tp,tp_r,1,"Session duration",font=B); p(ws_tp,tp_r,2,300,fill=YE,fmt="0")
p(ws_tp,tp_r,3,"sec"); p(ws_tp,tp_r,4,"Average conversation length (5 min = 300 sec)",font=IT); tp_r+=1

TP["response_time"] = tp_r
p(ws_tp,tp_r,1,"Average response time per message",font=B); p(ws_tp,tp_r,2,2.0,fill=YE,fmt="0.0")
p(ws_tp,tp_r,3,"sec"); p(ws_tp,tp_r,4,"Bot processing + LLM latency per turn",font=IT); tp_r+=1

TP["working_days"] = tp_r
p(ws_tp,tp_r,1,"Working days per month",font=B); p(ws_tp,tp_r,2,22,fill=YE,fmt="0")
p(ws_tp,tp_r,3,"days"); p(ws_tp,tp_r,4,"Business days; use 30 for 24/7 operations",font=IT); tp_r+=1

TP["peak_pct"] = tp_r
p(ws_tp,tp_r,1,"% of transactions in peak hours",font=B); p(ws_tp,tp_r,2,80,fill=YE,fmt="0")
p(ws_tp,tp_r,3,"%"); p(ws_tp,tp_r,4,"Traffic concentration — 80% is typical enterprise",font=IT); tp_r+=1

TP["peak_hours"] = tp_r
p(ws_tp,tp_r,1,"Peak hours per day",font=B); p(ws_tp,tp_r,2,4,fill=YE,fmt="0")
p(ws_tp,tp_r,3,"hours"); p(ws_tp,tp_r,4,"How many hours carry the peak traffic",font=IT); tp_r+=1

TP["sessions_year"] = tp_r
p(ws_tp,tp_r,1,"Number of sessions per year",font=B); p(ws_tp,tp_r,2,10000000,fill=YE,fmt="#,##0")
p(ws_tp,tp_r,3,"sessions"); p(ws_tp,tp_r,4,"← PRIMARY INPUT: total annual bot session volume",font=IT); tp_r+=1

# ── Derivation chain ──
tp_r+=1
sec(ws_tp,tp_r,"DERIVATION",fill=T4,ncols=4); tp_r+=1
hdr(ws_tp,tp_r,["Step","Value","Unit","Formula"]); tp_r+=1

# Cell references for formulas
_SY = f"B{TP['sessions_year']}"
_WD = f"B{TP['working_days']}"
_PP = f"B{TP['peak_pct']}"
_PH = f"B{TP['peak_hours']}"
_SD = f"B{TP['session_duration']}"
_TT = f"B{TP['think_time']}"
_RT = f"B{TP['response_time']}"

TP["sessions_month"] = tp_r
p(ws_tp,tp_r,1,"Sessions per month",font=B)
p(ws_tp,tp_r,2,f"={_SY}/12",fill=GR,fmt="#,##0")
p(ws_tp,tp_r,3,"sessions"); p(ws_tp,tp_r,4,"sessions_year / 12",font=IT); tp_r+=1

TP["sessions_day"] = tp_r
p(ws_tp,tp_r,1,"Sessions per day",font=B)
p(ws_tp,tp_r,2,f"=B{TP['sessions_month']}/{_WD}",fill=GR,fmt="#,##0")
p(ws_tp,tp_r,3,"sessions"); p(ws_tp,tp_r,4,"sessions_month / working_days",font=IT); tp_r+=1

TP["sessions_peak_hour"] = tp_r
p(ws_tp,tp_r,1,"Sessions per peak hour",font=B)
p(ws_tp,tp_r,2,f"=(B{TP['sessions_day']}*{_PP}/100)/{_PH}",fill=GR,fmt="#,##0")
p(ws_tp,tp_r,3,"sessions"); p(ws_tp,tp_r,4,"(sessions_day x peak%) / peak_hours",font=IT); tp_r+=1

TP["concurrent"] = tp_r
p(ws_tp,tp_r,1,"Concurrent sessions",font=B)
p(ws_tp,tp_r,2,f"=B{TP['sessions_peak_hour']}*({_SD}/3600)",fill=GR,fmt="#,##0")
p(ws_tp,tp_r,3,"sessions"); p(ws_tp,tp_r,4,"sessions_hour x (session_duration / 3600)",font=IT); tp_r+=1

TP["total_messages"] = tp_r
p(ws_tp,tp_r,1,"Messages in concurrent window",font=B)
p(ws_tp,tp_r,2,f"=(B{TP['concurrent']}*{_SD})/({_TT}+{_RT})",fill=GR,fmt="#,##0")
p(ws_tp,tp_r,3,"messages"); p(ws_tp,tp_r,4,"(concurrent x duration) / (think + response)",font=IT); tp_r+=1

TP["throughput"] = tp_r
p(ws_tp,tp_r,1,"Required Throughput (msg/sec)",font=B14)
p(ws_tp,tp_r,2,f"=B{TP['total_messages']}/{_SD}",fill=PatternFill("solid", fgColor="C6EFCE"),fmt="#,##0.0")
p(ws_tp,tp_r,3,"msg/sec"); p(ws_tp,tp_r,4,"← OUTPUT: feeds into Sizing & Cost sheet",font=B); tp_r+=1

# ── Summary box ──
tp_r+=1
sec(ws_tp,tp_r,"QUICK REFERENCE",fill=HB,ncols=4); tp_r+=1
p(ws_tp,tp_r,1,"Simplified formula:",font=B)
p(ws_tp,tp_r,2,"sessions_year × peak% / (12 × working_days × peak_hours × 3600 / session_duration × (think + response))",font=IT)
tp_r+=1
p(ws_tp,tp_r,1,"Channel think times:",font=B)
p(ws_tp,tp_r,2,"Socket (RTM, Liveperson) = 4s | Webhook (WhatsApp, Slack, Teams) = 5s | Voice (IVR) = 6s",font=IT)
tp_r+=1

# Store the throughput output cell reference for cross-sheet linking
THROUGHPUT_OUTPUT_CELL = f"B{TP['throughput']}"

# ── Option C: Patch Sizing & Cost B8 (target) with auto-pull + manual override ──
# B8 = IF(R8>0, R8, 'Throughput Calculator'!B<output>)
# R column (col 18) = hidden manual override cell; B8 = formula that pulls from throughput
# If user types a number in R8, that wins. Otherwise auto-pull from Throughput Calculator.
target_row = R["target"]
ws_sc = wb["Sizing & Cost"]
# Put manual override in column R (col 18) — same hidden column as zones helper
ws_sc.cell(row=target_row, column=18, value=args.target)
ws_sc.cell(row=target_row, column=18).fill = YE
ws_sc.cell(row=target_row, column=18).number_format = "0"
# B8 formula: IF override > 0, use override; else pull from Throughput Calculator
ws_sc.cell(row=target_row, column=2, value=f"=IF(R{target_row}>0,R{target_row},'Throughput Calculator'!{THROUGHPUT_OUTPUT_CELL})")
ws_sc.cell(row=target_row, column=2).fill = YE
ws_sc.cell(row=target_row, column=2).number_format = "0"
# Update the notes column to explain
ws_sc.cell(row=target_row, column=10, value="Auto from Throughput Calculator; override in R column (set to 0 for auto)")
ws_sc.cell(row=target_row, column=10).font = IT

# ═══ TAB: ASSUMPTIONS & NOTES ════════════════════════════════════════════════
ws_as = wb.create_sheet("Assumptions")
wid(ws_as, [8, 80])

as_r = 1
ws_as.merge_cells(start_row=as_r, start_column=1, end_row=as_r, end_column=2)
p(ws_as,as_r,1,"Assumptions and Notes",font=B14, fill=PatternFill("solid", fgColor="EAF2F8")); as_r+=2

sec(ws_as,as_r,"ASSUMPTIONS",fill=HB,ncols=2); as_r+=1
hdr(ws_as,as_r,["S.No","Assumption and Notes"]); as_r+=1

assumptions = [
    "The Agent Platform V2 hosting cost does not include any hosted models. V2 does not support NLP without an LLM, unlike XO. All agents must have an LLM.",
    "Software licenses for MongoDB and Redis are not included.",
    "Hosting Support cost is not included.",
    "External / Commercial models are required for agent reasoning and Search AI. External / commercial models cost is not included and borne by customer.",
    "Not considering the SearchAI in the sizing calculator.",
]
for idx, text in enumerate(assumptions, 1):
    p(ws_as,as_r,1,idx,font=B,al=AL_C)
    p(ws_as,as_r,2,text,font=Font(name="Segoe UI", size=10))
    as_r+=1

# ── Finalize — Assumptions (sheet 1), Throughput Calculator (sheet 2), Sizing & Cost (sheet 3) ──
wb._sheets = [wb["Assumptions"], wb["Throughput Calculator"], wb["Sizing & Cost"]]
beautify_workbook(wb)

# Set active tab to Assumptions (index 0) so all tabs are visible on open
wb.active = wb["Assumptions"]
# Ensure sheet tab bar is visible with enough width for all tab names
from openpyxl.workbook.views import BookView
wb.views = [BookView(tabRatio=800, showSheetTabs=True, activeTab=0)]

# ── Force recalculation on open ──
wb.calculation.calcMode = "auto"
wb.calculation.fullCalcOnLoad = True
wb.calculation.calcCompleted = False

# ── Post-postprocess fixup: re-pin wrap_text on the merged datastore legend block ──
# The global post-processor walks every cell and can override alignment based on row context.
# Since the legend spans F:J across the MongoDB/OpenSearch/Neo4j rows (which also contain
# editable B/C/D inputs), we re-assert wrap_text=True on the top-left of the merged region.
if "datastore_legend_top" in R:
    _sizing_ws = wb["Sizing & Cost"]
    _tl = _sizing_ws.cell(row=R["datastore_legend_top"], column=6)
    _tl.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

# ── Save and patch cached formula values ──
out = Path(args.output); out.parent.mkdir(parents=True, exist_ok=True)
wb.save(str(out))

# Patch the xlsx XML to add cached <v> values for formula cells.
# Google Sheets sometimes doesn't auto-evaluate openpyxl formulas on import.
# Adding a <v> element with a pre-calculated value ensures correct display.
import math, zipfile, shutil, tempfile
try:
    from lxml import etree
except ImportError:
    etree = None

def eval_formula(ws_obj, formula, cache):
    """Simple formula evaluator for CEILING, MAX, +, *, / expressions."""
    f = str(formula)
    if not f.startswith("="): return None
    f = f[1:]
    def resolve(m):
        ref = m.group(0)
        if ref in cache: return str(cache[ref])
        try:
            cell = ws_obj[ref]
            v = cell.value
            if v is None: return "0"
            if isinstance(v, str) and v.startswith("="):
                ev = eval_formula(ws_obj, v, cache)
                if ev is not None: cache[ref] = ev; return str(ev)
                return "0"
            cache[ref] = float(v) if isinstance(v, (int, float)) else 0
            return str(cache[ref])
        except: return "0"
    try:
        f = re.sub(r'[A-Z]{1,3}\d+', resolve, f)
        f = re.sub(r'CEILING\(([^,]+),\s*1\)', lambda m: f'__import__("math").ceil({m.group(1)})', f)
        f = re.sub(r'MAX\(([^)]+)\)', lambda m: f'max({m.group(1)})', f)
        f = re.sub(r'OR\(([^)]+)\)', lambda m: f'any([{m.group(1)}])', f)
        if "IF(" in f or '"' in f or "&" in f: return None
        return float(eval(f, {"max":max,"min":min,"any":any,"__builtins__":{}, "__import__": __import__}))
    except: return None

# Build cache of all formula values
if etree is None:
    print("[cache] lxml not installed; saved workbook without cached formula values")
else:
    for sheet_name in wb.sheetnames:
        ws_c = wb[sheet_name]
        cache = {}
        for row in ws_c.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    ref = f"{cell.column_letter}{cell.row}"
                    ev = eval_formula(ws_c, cell.value, cache)
                    if ev is not None: cache[ref] = ev

        if not cache: continue

        # Patch the XML
        tmp = tempfile.mkdtemp()
        try:
            with zipfile.ZipFile(str(out), 'r') as zin:
                zin.extractall(tmp)

            # Find the sheet XML
            sheet_idx = list(wb.sheetnames).index(sheet_name) + 1
            sheet_xml = Path(tmp) / f"xl/worksheets/sheet{sheet_idx}.xml"
            if not sheet_xml.exists(): continue

            tree = etree.parse(str(sheet_xml))
            ns = {'s': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}

            patched = 0
            for row_el in tree.findall('.//s:row', ns):
                for c_el in row_el.findall('s:c', ns):
                    ref = c_el.get('r', '')
                    f_el = c_el.find('s:f', ns)
                    if f_el is not None and ref in cache:
                        v_el = c_el.find('s:v', ns)
                        if v_el is None:
                            v_el = etree.SubElement(c_el, '{http://schemas.openxmlformats.org/spreadsheetml/2006/main}v')
                        v_el.text = str(cache[ref])
                        # Set type to number
                        c_el.set('t', 'n') if 't' in c_el.attrib else None
                        if c_el.get('t') == 's': del c_el.attrib['t']
                        patched += 1

            tree.write(str(sheet_xml), xml_declaration=True, encoding='UTF-8', standalone=True)

            # Re-zip
            with zipfile.ZipFile(str(out), 'w', zipfile.ZIP_DEFLATED) as zout:
                for root_dir, dirs, files in __import__('os').walk(tmp):
                    for file in files:
                        file_path = Path(root_dir) / file
                        arc_name = str(file_path.relative_to(tmp))
                        zout.write(str(file_path), arc_name)

            if patched: print(f"[cache] Patched {patched} formula cells in {sheet_name} with pre-calculated values")
        finally:
            shutil.rmtree(tmp, ignore_errors=True)
print(f"XLSX: {out} ({out.stat().st_size/1024:.1f} KB)")
print(f"Tabs: {', '.join(wb.sheetnames)}")
print(f"Tiers: runtime={len(tiers['runtime'])} search={len(tiers['search'])} workflow={len(tiers['workflow'])} pipeline={len(tiers['pipeline'])} data={len(tiers['data'])} supporting={len(tiers['supporting'])} operator={len(tiers['operator'])}")
